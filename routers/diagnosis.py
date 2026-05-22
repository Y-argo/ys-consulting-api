# api/routers/diagnosis.py
import datetime
_JST = datetime.timezone(datetime.timedelta(hours=9))
def _now_jst(): return datetime.datetime.now(_JST)
import uuid
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Body, Form, Request
from pydantic import BaseModel
from typing import List, Optional
from google.cloud import firestore as fs

from api.routers.auth import verify_token
from api.core.firestore_client import get_db, DEFAULT_TENANT
from api.core.llm_client import call_llm

router = APIRouter(prefix="/api/diagnosis", tags=["diagnosis"])

def _load_rank_config(tenant_id: str) -> dict:
    try:
        db = get_db()
        doc = db.collection("tenant_settings").document(tenant_id).get()
        if doc.exists:
            d = doc.to_dict() or {}
            return {
                "rank_1_name": d.get("rank_1_name", "D"),
                "rank_2_name": d.get("rank_2_name", "C"),
                "rank_3_name": d.get("rank_3_name", "B"),
                "rank_4_name": d.get("rank_4_name", "A"),
            }
    except Exception:
        pass
    return {"rank_1_name": "D", "rank_2_name": "C", "rank_3_name": "B", "rank_4_name": "A"}

def _load_chat_history_across_sessions(uid: str, tenant_id: str, limit: int = 30) -> list:
    db = get_db()
    try:
        sessions_raw = list(
            db.collection("chat_sessions")
            .where("uid", "==", uid)
            .limit(50)
            .stream()
        )
        sessions_raw.sort(key=lambda s: str((s.to_dict() or {}).get("updated_at", "")), reverse=True)
        sessions = sessions_raw[:10]
        msgs = []
        for s in sessions:
            ref = db.collection("chat_sessions").document(s.id).collection("messages")
            for m in ref.order_by("ts").limit_to_last(10).get():
                d = m.to_dict() or {}
                msgs.append({"role": d.get("role", "user"), "content": d.get("content", "")})
            if len(msgs) >= limit:
                break
        return msgs[-limit:]
    except Exception:
        return []

def _load_score_config(tenant_id: str) -> dict:
    db = get_db()
    DEFAULT = {
        "struct_words": "構造,資本,市場,制度,最適,期待値,確率,アーキテクチャ,設計,フレームワーク",
        "strategy_words": "戦略,施策,優先,差別化,競合,ポジショニング,KPI,ROI,目標",
        "exec_words": "実行,手順,タスク,スケジュール,チェック,改善,運用,効率",
        "emotion_words": "不安,ムカつく,なぜ俺,怖い,どうせ,無理,クソ,無能,イライラ,最悪",
    }
    try:
        for tid in [tenant_id, "default"]:
            doc = db.collection("system_settings").document(f"score_config_{tid}").get()
            if doc.exists:
                d = doc.to_dict() or {}
                return {**DEFAULT, **d}
    except Exception:
        pass
    return DEFAULT

def _generate_diagnosis(uid: str, tenant_id: str, n_chats: int = 30) -> str:
    rank_cfg = _load_rank_config(tenant_id)
    msgs = _load_chat_history_across_sessions(uid, tenant_id, limit=n_chats)
    chat_text = "\n".join([f"{m['role']}: {m.get('content','')}" for m in msgs if m.get("content")])
    if not chat_text.strip():
        return ""
    timestamp_str = _now_jst().strftime("%Y-%m-%d %H:%M")
    dr1 = rank_cfg["rank_1_name"]
    dr2 = rank_cfg["rank_2_name"]
    dr3 = rank_cfg["rank_3_name"]
    dr4 = rank_cfg["rank_4_name"]
    prompt = f"""以下はユーザー「{uid}」の直近チャット履歴（{n_chats}件）です。

{chat_text}

---
【ランク体系（低→高）】: {dr1} → {dr2} → {dr3} → {dr4}

上記の履歴を踏まえ、以下のフォーマットを厳守して「現状課題診断レポート」を生成してください。

# 現状課題診断レポート
生成日時: {timestamp_str}
対象ユーザー: {uid}
解析範囲: 直近チャット {n_chats} 件
---
## 総合評価
評価ランク: [S/A/B/C/D]
現状状態: [1〜3行で要約]
優先改善度: [低/中/高/緊急]
---
## 主要課題（最大3件）
### 課題1: [課題名]
影響度: [高/中/低]
原因: [構造的原因]
推奨行動: [具体行動]
---
### 課題2: [課題名]
影響度: [高/中/低]
原因: [構造的原因]
推奨行動: [具体行動]
---
### 課題3: [課題名]
影響度: [高/中/低]
原因: [構造的原因]
推奨行動: [具体行動]
---
## 強み・弱点
強み: [維持すべき行動]
弱点: [改善が必要な行動]
"""
    try:
        return call_llm(
            system_prompt="あなたは戦略コンサルタントです。与えられたチャット履歴を分析し、構造的な課題診断レポートを生成してください。",
            messages=[{"role": "user", "content": prompt}],
            ai_tier="core",
            max_tokens=4096,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"診断生成エラー: {e}")

def _save_diagnosis(uid: str, tenant_id: str, report_md: str, rank: str = None, n_chats: int = None) -> str:
    db = get_db()
    doc_id = str(uuid.uuid4())
    data = {
        "uid": uid,
        "tenant_id": tenant_id,
        "report_md": report_md,
        "created_at": fs.SERVER_TIMESTAMP,
    }
    if rank: data["rank"] = rank
    if n_chats: data["n_chats"] = n_chats
    db.collection("user_diagnoses").document(doc_id).set(data)
    return doc_id

def _load_diagnoses(uid: str, tenant_id: str, limit: int = 5) -> list:
    db = get_db()
    try:
        docs = list(
            db.collection("user_diagnoses")
            .where("uid", "==", uid)
            .limit(50)
            .stream()
        )
        result = []
        for d in docs:
            data = d.to_dict() or {}
            result.append({
                "doc_id": d.id,
                "report_md": data.get("report_md", ""),
                "created_at": str(data.get("created_at", "")),
                "rank": data.get("rank"),
                "n_chats": data.get("n_chats"),
            })
        result.sort(key=lambda x: x["created_at"], reverse=True)
        return result[:limit]
    except Exception:
        return []

class DiagnosisRequest(BaseModel):
    n_chats: int = 30

@router.post("/generate")
def generate_diagnosis(req: DiagnosisRequest, payload: dict = Depends(verify_token)):
    from api.core.features import is_feature_enabled
    uid = payload["uid"]; tenant_id = payload.get("tenant_id","default")
    if not is_feature_enabled(uid, "current_issue_diagnosis"):
        raise HTTPException(status_code=403, detail="この機能は利用できません")
    uid = payload["uid"]
    tenant_id = payload.get("tenant_id", DEFAULT_TENANT)
    report_md = _generate_diagnosis(uid, tenant_id, req.n_chats)
    if not report_md:
        raise HTTPException(status_code=400, detail="チャット履歴が不足しています")
    # ランク取得
    try:
        from api.routers.user_stats import _load_rank_config, _score_to_rank
        from api.core.firestore_client import get_db as _gdb2
        _snap = _gdb2().collection("users").document(uid).get()
        _score = int((_snap.to_dict() or {}).get("level_score", 0)) if _snap.exists else 0
        _cfg = _load_rank_config(tenant_id)
        _rank = _score_to_rank(_score, _cfg)
    except Exception:
        _rank = None
    doc_id = _save_diagnosis(uid, tenant_id, report_md, rank=_rank, n_chats=req.n_chats)
    return {"doc_id": doc_id, "report_md": report_md}

@router.get("/list")
def list_diagnoses(payload: dict = Depends(verify_token)):
    uid = payload["uid"]
    tenant_id = payload.get("tenant_id", DEFAULT_TENANT)
    return {"diagnoses": _load_diagnoses(uid, tenant_id)}

@router.get("/thought_map")
def get_thought_map(payload: dict = Depends(verify_token)):
    from api.core.features import is_feature_enabled
    uid = payload["uid"]
    tenant_id = payload.get("tenant_id", DEFAULT_TENANT)
    if not is_feature_enabled(uid, "diag_graph"):
        raise HTTPException(status_code=403, detail="会話の可視化は現在未開放のため使用できません。")
    db = get_db()
    try:
        sessions_raw = list(db.collection("chat_sessions").where("uid","==",uid).limit(200).stream())
        sessions_raw.sort(key=lambda s: str((s.to_dict() or {}).get("updated_at","")), reverse=True)

        TOPICS = {
            "戦略・競合":   ["戦略","競合","差別化","ポジション","市場","シェア","ブランド","参入","撤退","優位"],
            "集客・SNS":   ["集客","SNS","Instagram","Twitter","広告","フォロワー","投稿","認知","バズ","LP"],
            "売上・財務":   ["売上","収益","利益","資金","コスト","価格","単価","資産","融資","キャッシュ"],
            "組織・人材":   ["採用","チーム","組織","スタッフ","教育","マネジメント","育成","離職","評価"],
            "投資・株":     ["投資","株","銘柄","シグナル","相場","資産","ポートフォリオ","底打ち","売り"],
            "診断・分析":   ["診断","分析","レポート","スコア","評価","課題","ボトルネック","原因"],
            "指名・接客":   ["指名","接客","お客様","キャスト","リピート","コミュニケーション"],
            "マーケ・集客": ["マーケ","宣材","写真","予約","ネット予約","口コミ","レビュー","集患"],
            "計画・実行":   ["計画","ロードマップ","スケジュール","手順","ステップ","期限","タスク"],
            "リスク・危機": ["リスク","危機","失敗","損失","トラブル","クレーム","炎上","最悪"],
            "成長・習慣":   ["成長","スキル","習慣","訓練","学習","キャリア","目標","継続"],
            "交渉・説得":   ["交渉","説得","条件","合意","契約","提案","伝え方","折衝"],
            "思考・構造":   ["構造","フレームワーク","ロジック","仮説","論点","因果","MECE"],
        }

        def classify(text):
            for topic, keywords in TOPICS.items():
                if any(k in text for k in keywords):
                    return topic
            return "その他"

        theme_count = {}
        theme_sessions = {}
        all_prompts = []
        raw_nodes = []
        node_set = set()

        for s in sessions_raw[:20]:
            try:
                s_data = s.to_dict() or {}
                s_date = str(s_data.get("updated_at",""))[:10]
                msgs = list(db.collection("chat_sessions").document(s.id).collection("messages").limit(40).stream())
                msgs.sort(key=lambda m: str((m.to_dict() or {}).get("ts","")))
                for m in msgs:
                    d = m.to_dict() or {}
                    if d.get("role") == "user":
                        content = (d.get("content","") or "").strip()
                        if not content or len(content) <= 3: continue
                        all_prompts.append(content)
                        topic = classify(content)
                        theme_count[topic] = theme_count.get(topic, 0) + 1
                        if topic not in theme_sessions:
                            theme_sessions[topic] = []
                        if s_date not in theme_sessions[topic]:
                            theme_sessions[topic].append(s_date)
                        label = content[:35]
                        if label not in node_set:
                            node_set.add(label)
                            raw_nodes.append({"id": label, "label": label, "group": topic, "full_text": content[:200]})
            except Exception:
                continue

        raw_nodes = raw_nodes[:40]

        unresolved_alerts = [
            {"topic": t, "count": c, "message": f"「{t}」に関する相談が{c}回繰り返されています。根本解決が必要な可能性があります。"}
            for t, c in theme_count.items() if c >= 3 and t != "その他"
        ]
        unresolved_alerts.sort(key=lambda x: x["count"], reverse=True)

        growth_trend = [
            {"topic": t, "session_count": len(dates), "last_date": max(dates) if dates else ""}
            for t, dates in theme_sessions.items() if t != "その他"
        ]
        growth_trend.sort(key=lambda x: x["session_count"], reverse=True)

        topic_last = {}
        edges = []
        for n in raw_nodes:
            g = n["group"]
            if g in topic_last:
                edges.append({"from": topic_last[g], "to": n["id"], "topic": g})
            topic_last[g] = n["id"]

        topics_used = list(set(n["group"] for n in raw_nodes))
        center_nodes = [{"id": f"__topic_{t}__", "label": t, "group": t, "is_center": True} for t in topics_used]
        center_edges = [{"from": f"__topic_{n['group']}__", "to": n["id"], "topic": n["group"]} for n in raw_nodes]
        all_nodes = center_nodes + raw_nodes
        all_edges = center_edges

        issue_tree = {}
        try:
            if all_prompts:
                from api.core.llm_client import call_llm as _cllm
                import json as _json, re as _re
                _sample = "\n".join(all_prompts[:30])[:3000]
                _tree_prompt = f"""以下はユーザーの直近チャット相談内容です。コンサルタントとして課題構造を分析せよ。JSONのみ出力。

{_sample}

以下のJSONスキーマで返せ:
{{
  "root_issues": ["根本課題1", "根本課題2"],
  "surface_issues": ["表面的課題1", "表面的課題2", "表面的課題3"],
  "recurring_patterns": ["繰り返しパターン1", "繰り返しパターン2"],
  "growth_opportunities": ["成長機会1", "成長機会2"],
  "priority_action": "最優先で取り組むべきこと（1文）"
}}"""
                _raw = _cllm(
                    system_prompt="戦略コンサルタント。JSONのみ出力。",
                    messages=[{"role":"user","content":_tree_prompt}],
                    ai_tier="core", max_tokens=1024
                )
                _m = _re.search(r"\{.*\}", _raw, _re.DOTALL)
                if _m:
                    issue_tree = _json.loads(_m.group(0))
        except Exception:
            pass

        return {
            "nodes": all_nodes,
            "edges": all_edges,
            "topics": topics_used,
            "unresolved_alerts": unresolved_alerts[:5],
            "growth_trend": growth_trend[:8],
            "issue_tree": issue_tree,
            "theme_count": theme_count,
        }
    except Exception as e:
        return {"nodes": [], "edges": [], "error": str(e)}

class ConsultRequest(BaseModel):
    analysis_type: str
    input_text: str
    supplement: str = ""
    options: str = ""
    strategy: str = ""
    policy: str = ""

@router.post("/consult")
def run_consult(req: ConsultRequest, payload: dict = Depends(verify_token)):
    from api.core.llm_client import call_llm
    from api.core.features import is_feature_enabled
    uid = payload["uid"]
    tenant_id = payload.get("tenant_id", "default")
    _ANALYSIS_FLAG_MAP = {
        "structure": "diag_structure",
        "issue": "diag_issue",
        "comparison": "diag_comparison",
        "contradiction": "diag_contradiction",
        "execution": "diag_execution",
        "investment": "diag_investment",
        "decision_metrics": "decision_metrics",
    }
    _flag_id = _ANALYSIS_FLAG_MAP.get(req.analysis_type)
    if _flag_id and not is_feature_enabled(uid, _flag_id):
        raise HTTPException(status_code=403, detail=f"この機能は現在未開放のため使用できません。")
    db = get_db()

    frameworks = []
    try:
        docs = list(db.collection("tenants").document(tenant_id).collection("consulting_frameworks").where("active","==",True).stream())
        frameworks = [d.to_dict().get("name","") for d in docs if d.to_dict().get("name")]
    except Exception:
        pass

    fw_str = "、".join(frameworks[:5]) if frameworks else "MECE・SWOT・3C・ロジックツリー・Issue Tree"

    if req.analysis_type == "structure":
        import re as _re_s, json as _json_s

        STRUCTURE_REQUIRED_KEYS = [
            "issue_summary", "observations", "structure_map",
            "surface_causes", "root_causes", "causal_chains",
            "bottlenecks", "current_structure", "ideal_structure",
            "transition_barriers", "priority_points", "recommended_actions",
            "risks", "verification_plan", "missing_information"
        ]
        STRUCTURE_MAP_KEYS = [
            "core_system", "actors", "resources",
            "constraints", "feedback_loops", "failure_points"
        ]

        def _parse_structure_json(raw: str):
            text = str(raw or "").strip()
            text = _re_s.sub(r"^```[a-zA-Z]*\n?", "", text)
            text = _re_s.sub(r"```$", "", text).strip()
            s = text.find("{")
            e = text.rfind("}")
            if s == -1 or e == -1 or e <= s:
                return None
            try:
                return _json_s.loads(text[s:e+1])
            except Exception:
                return None

        def _repair_structure_keys(data: dict) -> dict:
            for k in STRUCTURE_REQUIRED_KEYS:
                if k not in data:
                    if k == "structure_map":
                        data[k] = {sk: [] if sk != "core_system" else "" for sk in STRUCTURE_MAP_KEYS}
                    elif k in ("issue_summary", "current_structure", "ideal_structure"):
                        data[k] = ""
                    else:
                        data[k] = []
            if isinstance(data.get("structure_map"), dict):
                for sk in STRUCTURE_MAP_KEYS:
                    if sk not in data["structure_map"]:
                        data["structure_map"][sk] = [] if sk != "core_system" else ""
            return data

        structure_prompt = f"""あなたは戦略コンサルタントです。以下の相談内容を構造診断してください。
適用フレームワーク: {fw_str}

【相談内容】
{req.input_text}

【補足情報】
{req.supplement or "（なし）"}

【分解順序】
必ず「観測事実 → 表層原因 → 根因 → 因果連鎖 → ボトルネック → 移行障害 → 初手」の順で分解すること。

【厳守ルール】
- 感想・印象・一般論は一切禁止
- 入力にない事実の断定は禁止
- 観測事実に基づかない推測はmissing_informationに格納すること
- confidence は根拠量に応じて必ず「高/中/低」で分けること
- recommended_actions は必ず bottlenecks と causal_chains に紐付けること
- JSONのみ返すこと。説明文・前置き・Markdownコードブロック・余計なテキストは一切禁止
- 最初の文字は必ず{{で始めること

以下のJSONスキーマで返すこと:
{{
  "issue_summary": "問題の要約（1〜2文）",
  "observations": ["観測事実1", "観測事実2"],
  "structure_map": {{
    "core_system": "中心システム・仕組みの説明",
    "actors": ["関係者・主体1", "関係者2"],
    "resources": ["リソース1", "リソース2"],
    "constraints": ["制約1", "制約2"],
    "feedback_loops": ["フィードバック構造1"],
    "failure_points": ["破綻ポイント1"]
  }},
  "surface_causes": ["表層原因1", "表層原因2"],
  "root_causes": ["根因1", "根因2"],
  "causal_chains": [
    {{
      "root": "根因",
      "mechanism": "伝播メカニズム",
      "symptom": "表出している症状",
      "business_impact": "事業への影響",
      "evidence": ["根拠となる事実"],
      "confidence": "高/中/低"
    }}
  ],
  "bottlenecks": [
    {{
      "point": "ボトルネック箇所",
      "why_bottleneck": "なぜボトルネックか",
      "affected_area": "影響範囲",
      "priority": "高/中/低",
      "first_fix": "最初に手をつけるべき施策"
    }}
  ],
  "current_structure": "現状の構造を一文で",
  "ideal_structure": "理想の構造を一文で",
  "transition_barriers": ["移行障害1", "移行障害2"],
  "priority_points": ["優先論点1", "優先論点2"],
  "recommended_actions": ["打ち手1（bottlenecks/causal_chainsに紐付け）", "打ち手2"],
  "risks": ["リスク1", "リスク2"],
  "verification_plan": ["検証ステップ1", "検証ステップ2"],
  "missing_information": ["不足情報1", "不足情報2"]
}}"""

        quality_prompt_tpl = """以下の構造診断JSONを品質レビューし、修正済みJSONのみ返してください。
【レビュー観点】
- causal_chains の confidence が根拠なく「高」になっていないか → 根拠が薄ければ「中」か「低」に下げる
- bottlenecks の priority が全て「高」になっていないか → 優先度に差をつける
- verification_plan が空または抽象的すぎないか → 具体的な検証ステップを補完する
- recommended_actions が bottlenecks / causal_chains に紐付いていない場合 → 紐付けコメントを追加する
- 修正不要な項目はそのまま維持すること
- JSONのみ返すこと

【対象JSON】
{result_json}"""

        try:
            # Step1: 初回生成
            _s_raw = call_llm(
                system_prompt="あなたは戦略コンサルタントです。必ず有効なJSONオブジェクトのみ返してください。説明文・前置き・Markdownコードブロック・余計なテキストは一切禁止です。最初の文字は必ず{で始めてください。",
                messages=[{"role": "user", "content": structure_prompt}],
                ai_tier="core", max_tokens=6000
            )
            _s_data = _parse_structure_json(_s_raw)

            # Step2: parse失敗時 → repair prompt
            if _s_data is None:
                _repair_prompt = f"""以下のテキストは壊れたJSONです。下記スキーマに従って有効なJSONオブジェクトのみに修復して返してください。
余計な説明は禁止。JSONのみ返すこと。最初の文字は必ず{{で始めること。

【必須スキーマ】
{{
  "issue_summary": "",
  "observations": [],
  "structure_map": {{
    "core_system": "",
    "actors": [],
    "resources": [],
    "constraints": [],
    "feedback_loops": [],
    "failure_points": []
  }},
  "surface_causes": [],
  "root_causes": [],
  "causal_chains": [{{"root":"","mechanism":"","symptom":"","business_impact":"","evidence":[],"confidence":"高/中/低"}}],
  "bottlenecks": [{{"point":"","why_bottleneck":"","affected_area":"","priority":"高/中/低","first_fix":""}}],
  "current_structure": "",
  "ideal_structure": "",
  "transition_barriers": [],
  "priority_points": [],
  "recommended_actions": [],
  "risks": [],
  "verification_plan": [],
  "missing_information": []
}}

【壊れたテキスト】
{_s_raw[:3000]}"""
                _s_raw2 = call_llm(
                    system_prompt="有効なJSONオブジェクトのみ返してください。",
                    messages=[{"role": "user", "content": _repair_prompt}],
                    ai_tier="core", max_tokens=6000
                )
                _s_data = _parse_structure_json(_s_raw2)
                if _s_data is None:
                    return {"ok": False, "error": "構造診断JSONの生成・修復に失敗しました", "analysis_type": "structure"}

            # Step3: 必須キー補完
            _s_data = _repair_structure_keys(_s_data)

            # Step4: 品質監査
            _q_prompt = quality_prompt_tpl.format(result_json=_json_s.dumps(_s_data, ensure_ascii=False))
            _q_raw = call_llm(
                system_prompt="有効なJSONオブジェクトのみ返してください。",
                messages=[{"role": "user", "content": _q_prompt}],
                ai_tier="core", max_tokens=6000
            )
            _q_data = _parse_structure_json(_q_raw)
            if _q_data is not None:
                _q_data = _repair_structure_keys(_q_data)
                _s_data = _q_data

            # Step5: Firestore保存
            db.collection("tenants").document(tenant_id).collection("consulting_analyses").add({
                "uid": uid, "tenant_id": tenant_id,
                "analysis_type": "structure",
                "input_text": req.input_text[:500],
                "result": _s_data,
                "created_at": _now_jst().isoformat(),
            })
            return {"ok": True, "result": _s_data, "analysis_type": "structure"}

        except Exception as _se:
            return {"ok": False, "error": str(_se), "analysis_type": "structure"}

    elif req.analysis_type == "issue":
        # ── issue_v2 構造化エンジン（独立ブロック） ──
        import re as _iss_re, json as _iss_json

        _input_text = (req.input_text or "").strip()
        _supplement_text = (req.supplement or "（なし）").strip()
        _options_text = (req.options or "（なし）").strip()
        _strategy_text = (req.strategy or "（なし）").strip()
        _policy_text = (req.policy or "（なし）").strip()

        if not _input_text:
            return {"ok": False, "error": "課題仮説生成には状況・背景の入力が必要です。"}

        def _parse_issue_json(raw, label=""):
            text = str(raw).strip()
            text = _iss_re.sub(r"^```[a-zA-Z]*\n?", "", text)
            text = _iss_re.sub(r"```$", "", text).strip()
            s = text.find("{")
            e = text.rfind("}")
            if s == -1 or e == -1 or e <= s:
                raise ValueError(f"JSON not found [{label}]: raw_length={len(text)} raw_preview={text[:300]}")
            return _iss_json.loads(text[s:e+1])

        _issue_prompt1 = f"""あなたは戦略コンサルタント兼Issue Tree設計者です。
以下の入力から、課題仮説を構造化してください。
感想・一般論・努力論は禁止です。
入力にない事実は断定せず、不明点は missing_information に入れてください。

適用フレームワーク: {fw_str}

【入力内容】
{_input_text}

【補足情報】
{_supplement_text}

【比較案・選択肢】
{_options_text}

【戦略前提】
{_strategy_text}

【方針・制約】
{_policy_text}

分析順序:
1. 観測事実を整理する
2. 表面的課題と根本課題を分ける
3. 主要論点を最大5件に絞る
4. 課題仮説を最大5件生成する
5. 各仮説に根拠・影響・検証方法・必要データ・反証条件・初手を付ける
6. 因果連鎖を root → mechanism → symptom → business_impact で整理する

JSONのみ出力。最初の文字は必ず {{ にすること。

{{
  "issue_summary": "問題の要約（1〜2文）",
  "main_issues": ["主要論点1", "主要論点2", "主要論点3"],
  "root_issues": ["根本課題1", "根本課題2"],
  "surface_issues": ["表面的課題1", "表面的課題2"],
  "hypotheses": [
    {{
      "title": "仮説名",
      "description": "仮説の内容",
      "priority": "高",
      "confidence": "中",
      "evidence": ["根拠1", "根拠2"],
      "expected_impact": "この仮説が正しい場合の影響",
      "verification_method": "検証方法",
      "required_data": ["必要データ1", "必要データ2"],
      "falsification_condition": "この条件なら仮説を棄却する",
      "first_action": "最初に取るべき行動"
    }}
  ],
  "causal_chains": [
    {{
      "root": "根本原因",
      "mechanism": "発生メカニズム",
      "symptom": "表面化している症状",
      "business_impact": "事業への影響"
    }}
  ],
  "questions_to_verify": ["次に確認すべき質問1", "質問2"],
  "required_data": ["全体として必要なデータ1", "データ2"],
  "decision_points": ["意思決定ポイント1", "ポイント2"],
  "missing_information": ["不足情報1", "不足情報2"]
}}"""

        _raw1 = call_llm(
            system_prompt="あなたは戦略コンサルタント兼Issue Tree設計者です。JSONのみ出力。最初の文字は{で始めてください。",
            messages=[{"role": "user", "content": _issue_prompt1}],
            ai_tier="core",
            max_tokens=7000
        )
        try:
            _data1 = _parse_issue_json(_raw1, "issue_stage1")
        except Exception as _e1:
            print(f"[issue_stage1] parse failed: {_e1} / raw_preview={str(_raw1)[:200]}")
            _repair_prompt1 = f"""以下のJSONが壊れています。必ず修復して有効なJSONのみ返してください。
最初の文字は必ず {{ にすること。JSON以外は絶対に出力しないこと。
【壊れたJSON】
{str(_raw1)[:2000]}
必須フィールド: issue_summary, main_issues, root_issues, surface_issues, hypotheses, causal_chains, questions_to_verify, required_data, decision_points, missing_information"""
            try:
                _repair_raw1 = call_llm(
                    system_prompt="JSONのみ出力。最初の文字は{で始めてください。",
                    messages=[{"role": "user", "content": _repair_prompt1}],
                    ai_tier="core",
                    max_tokens=7000
                )
                _data1 = _parse_issue_json(_repair_raw1, "issue_stage1_repair")
                print("[issue_stage1_repair] 修復成功")
            except Exception as _e1r:
                print(f"[issue_stage1_repair] 修復失敗: {_e1r}")
                return {"ok": False, "error": "課題仮説の生成に失敗しました。もう一度お試しください。"}

        _required_issue_keys = [
            "issue_summary",
            "main_issues",
            "root_issues",
            "surface_issues",
            "hypotheses",
            "causal_chains",
            "questions_to_verify",
            "required_data",
            "decision_points",
            "missing_information",
        ]
        _missing_keys = [k for k in _required_issue_keys if k not in _data1]
        if _missing_keys:
            print(f"[issue_stage1] missing keys: {_missing_keys}", flush=True)
            _repair_prompt1_key = (
                "以下のJSONはキー欠落があります。\n"
                "不足キーを補完し、有効なJSONとして返してください。\n"
                "JSON以外は絶対に出力しないこと。\n"
                f"【不足キー】\n{_missing_keys}\n"
                f"【現在JSON】\n{__import__('json').dumps(_data1, ensure_ascii=False)[:4000]}\n"
                f"必須キー:\n{_required_issue_keys}\n"
            )
            _repair_raw1_key = call_llm(
                system_prompt="JSON修復AI。JSONのみ出力。",
                messages=[{"role": "user", "content": _repair_prompt1_key}],
                ai_tier="core",
                max_tokens=2500
            )
            try:
                _repair_data1_key = _parse_issue_json(_repair_raw1_key, "issue_stage1_keyrepair")
                for _k in _required_issue_keys:
                    if _k not in _data1 and _k in _repair_data1_key:
                        _data1[_k] = _repair_data1_key[_k]
                print("[issue_stage1_keyrepair] 補完成功", flush=True)
            except Exception as _e_keyrepair:
                print(f"[issue_stage1_keyrepair] failed: {_e_keyrepair}", flush=True)
        _review_prompt = f"""以下の課題仮説JSONをレビューし、優先順位・確度・反証条件・初手の妥当性を補正してください。
過剰断定を避け、入力根拠が弱い仮説はconfidenceを下げてください。
JSONのみ出力。最初の文字は必ず {{ にすること。

【元入力】
{_input_text}

【課題仮説JSON】
{_iss_json.dumps(_data1, ensure_ascii=False)}

出力:
{{
  "hypotheses": [
    {{
      "title": "仮説名",
      "priority": "高/中/低",
      "confidence": "高/中/低",
      "priority_reason": "優先度理由",
      "verification_score": 0.0,
      "falsification_condition": "反証条件",
      "first_action": "初手"
    }}
  ],
  "top_hypothesis": "最重要仮説名",
  "priority_reason": "最重要と判断した理由",
  "verification_score": 0.0
}}"""

        _raw2 = call_llm(
            system_prompt="あなたは戦略仮説の品質監査AIです。JSONのみ出力。最初の文字は{で始めてください。",
            messages=[{"role": "user", "content": _review_prompt}],
            ai_tier="core",
            max_tokens=4000
        )
        try:
            _data2 = _parse_issue_json(_raw2, "issue_stage2")
        except Exception as _e2:
            print(f"[issue_stage2] parse failed: {_e2} / raw_preview={str(_raw2)[:200]}")
            _repair_prompt = f"""以下のJSONが壊れています。必ず修復して有効なJSONのみ返してください。
最初の文字は必ず {{ にすること。JSON以外は絶対に出力しないこと。

【壊れたJSON】
{str(_raw2)[:2000]}

【元の課題仮説データ（参考）】
{_iss_json.dumps(_data1, ensure_ascii=False)[:2000]}

必須スキーマ:
{{
  "hypotheses": [
    {{
      "title": "仮説名",
      "priority": "高/中/低",
      "confidence": "高/中/低",
      "priority_reason": "優先度理由",
      "verification_score": 0.0,
      "falsification_condition": "反証条件",
      "first_action": "初手"
    }}
  ],
  "top_hypothesis": "最重要仮説名",
  "priority_reason": "最重要と判断した理由",
  "verification_score": 0.0
}}"""
            try:
                _repair_raw = call_llm(
                    system_prompt="JSONのみ出力。最初の文字は{で始めてください。",
                    messages=[{"role": "user", "content": _repair_prompt}],
                    ai_tier="core",
                    max_tokens=3000
                )
                _data2 = _parse_issue_json(_repair_raw, "issue_stage2_repair")
                print("[issue_stage2_repair] 修復成功")
            except Exception as _e3:
                print(f"[issue_stage2_repair] 修復失敗: {_e3}")
                _hyps_for_fallback = [h for h in (_data1.get("hypotheses", []) or []) if isinstance(h, dict)]
                _prio_score_fb = {"高": 3, "中": 2, "低": 1}
                _conf_score_fb = {"高": 3, "中": 2, "低": 1}
                _sorted_fb = sorted(
                    _hyps_for_fallback,
                    key=lambda h: (
                        _prio_score_fb.get(str(h.get("priority","中")), 2),
                        _conf_score_fb.get(str(h.get("confidence","中")), 2),
                        len(h.get("evidence", []) or []),
                    ),
                    reverse=True
                )
                _top_fb = _sorted_fb[0] if _sorted_fb else {}
                _top_fb_title = str(_top_fb.get("title") or "")
                _top_fb_prio = str(_top_fb.get("priority") or "中")
                _top_fb_conf = str(_top_fb.get("confidence") or "中")
                _top_fb_ev = len(_top_fb.get("evidence", []) or [])
                _priority_reason_fb = f"優先度:{_top_fb_prio}・確度:{_top_fb_conf}・根拠{_top_fb_ev}件をもとに最重要仮説として選定"
                _data2 = {
                    "hypotheses": [
                        {
                            "title": str(h.get("title") or "仮説"),
                            "priority": str(h.get("priority") or "中"),
                            "confidence": str(h.get("confidence") or "中"),
                            "priority_reason": "",
                            "verification_score": 0.5,
                            "falsification_condition": str(h.get("falsification_condition") or "反証条件未設定"),
                            "first_action": str(h.get("first_action") or "追加情報を収集する"),
                        }
                        for h in _hyps_for_fallback[:5]
                    ],
                    "top_hypothesis": _top_fb_title,
                    "priority_reason": _priority_reason_fb,
                    "verification_score": 0.5,
                }

        _review_map = {}
        for _h in _data2.get("hypotheses", []) or []:
            if isinstance(_h, dict) and _h.get("title"):
                _review_map[str(_h.get("title"))] = _h

        _missing = list(_data1.get("missing_information", []) or [])
        _hypotheses = []

        _prio_score = {"高": 3, "中": 2, "低": 1}
        _conf_score = {"高": 3, "中": 2, "低": 1}

        for _h in (_data1.get("hypotheses", []) or [])[:5]:
            if not isinstance(_h, dict):
                _h = {"title": str(_h), "description": str(_h)}

            _title = str(_h.get("title") or "仮説")
            _rv = _review_map.get(_title, {})

            _evidence = _h.get("evidence", [])
            if not isinstance(_evidence, list):
                _evidence = [str(_evidence)] if _evidence else []

            _required = _h.get("required_data", [])
            if not isinstance(_required, list):
                _required = [str(_required)] if _required else []

            _priority = str(_rv.get("priority") or _h.get("priority") or "中")
            if _priority not in ["高", "中", "低"]:
                _priority = "中"

            _confidence = str(_rv.get("confidence") or _h.get("confidence") or "中")
            if _confidence not in ["高", "中", "低"]:
                _confidence = "中"

            if not _evidence:
                _confidence = "低"
                if "仮説の根拠情報が不足しています" not in _missing:
                    _missing.append("仮説の根拠情報が不足しています")

            if not _required:
                if "仮説検証に必要なデータが不足しています" not in _missing:
                    _missing.append("仮説検証に必要なデータが不足しています")

            _verification_score = _rv.get("verification_score", None)
            try:
                _verification_score = float(_verification_score)
            except Exception:
                _verification_score = 0.5 if _required else 0.3
            _verification_score = max(0.0, min(1.0, _verification_score))

            _hypotheses.append({
                "title": _title,
                "description": str(_h.get("description") or ""),
                "priority": _priority,
                "confidence": _confidence,
                "priority_reason": str(_rv.get("priority_reason") or ""),
                "evidence": _evidence[:5],
                "expected_impact": str(_h.get("expected_impact") or ""),
                "verification_method": str(_h.get("verification_method") or "検証方法未設定"),
                "required_data": _required[:5],
                "falsification_condition": str(_rv.get("falsification_condition") or _h.get("falsification_condition") or "反証条件未設定"),
                "first_action": str(_rv.get("first_action") or _h.get("first_action") or "追加情報を収集する"),
                "verification_score": _verification_score,
            })

        def _rank_key(h):
            return (
                _prio_score.get(h.get("priority"), 2),
                _conf_score.get(h.get("confidence"), 2),
                len(h.get("evidence", []) or []),
                float(h.get("verification_score", 0.0) or 0.0),
            )

        _top = max(_hypotheses, key=_rank_key) if _hypotheses else {}
        _top_title = str(_top.get("title") or _data2.get("top_hypothesis") or "")
        _priority_reason = str(_data2.get("priority_reason") or _top.get("priority_reason") or "優先度・確度・根拠数をもとに最重要仮説として選定")

        _root_issues = (_data1.get("root_issues", []) or [])[:5]
        _surface_issues = (_data1.get("surface_issues", []) or [])[:5]
        _causal_chains = (_data1.get("causal_chains", []) or [])[:5]
        if not _root_issues:
            _root_issues = [
                "顧客が価値を感じる要素と、現在提示している新人キャストの魅力定義が一致していない可能性",
                "顧客が予約判断に必要とする情報が不足している可能性",
            ]
        if not _surface_issues:
            _surface_issues = [
                "新人キャストへの事前予約が入らない",
                "売り込み文を変更しても予約改善につながっていない",
            ]
        if not _causal_chains:
            _top_for_chain = _hypotheses[0] if _hypotheses else {}
            _causal_chains = [
                {
                    "root": _root_issues[0] if _root_issues else "根本課題未特定",
                    "mechanism": str(_top_for_chain.get("description") or "価値定義・伝達・顧客期待の接続が弱く、予約動機が形成されていない"),
                    "symptom": _surface_issues[0] if _surface_issues else "表面的課題未特定",
                    "business_impact": str(_top_for_chain.get("expected_impact") or "予約機会の損失、広告・訴求改善の空回り、キャスト定着への悪影響"),
                }
            ]
        _result = {
            "issue_summary": str(_data1.get("issue_summary") or ""),
            "main_issues": (_data1.get("main_issues", []) or [])[:5],
            "root_issues": _root_issues,
            "surface_issues": _surface_issues,
            "hypotheses": _hypotheses,
            "top_hypothesis": _top_title,
            "priority_reason": _priority_reason,
            "questions_to_verify": (_data1.get("questions_to_verify", []) or [])[:8],
            "required_data": (_data1.get("required_data", []) or [])[:8],
            "decision_points": (_data1.get("decision_points", []) or [])[:6],
            "missing_information": _missing[:8],
            "causal_chains": _causal_chains,
            "verification_score": max([h.get("verification_score", 0.0) for h in _hypotheses], default=0.0),
        }

        db.collection("tenants").document(tenant_id).collection("consulting_analyses").add({
            "uid": uid,
            "tenant_id": tenant_id,
            "analysis_type": req.analysis_type,
            "input_text": req.input_text[:500],
            "result": _result,
            "created_at": _now_jst().isoformat(),
        })
        if req.analysis_type == "issue":
            try:
                db.collection("issue_reports").add({
                    "uid": uid,
                    "tenant_id": tenant_id,
                    "input_text": req.input_text[:2000],
                    "result": _result,
                    "created_at": fs.SERVER_TIMESTAMP,
                })
            except Exception as _issue_save_err:
                print(f"[issue_save] {_issue_save_err}", flush=True)
        return {"ok": True, "result": _result, "analysis_type": req.analysis_type}

    elif req.analysis_type == "comparison":
        # ── comparison 構造化エンジン（独立ブロック） ──
        import re as _cmp_re, json as _cmp_json

        # 1. 案バリデーション・最大5案に制限
        _raw_options = (req.options or req.input_text or "").strip()
        _option_lines = [l.strip() for l in _cmp_re.split(r"[\n,、/／|｜]", _raw_options) if l.strip()]
        _option_lines = list(dict.fromkeys(_option_lines))[:5]  # 重複排除・最大5案
        if len(_option_lines) < 2:
            return {"ok": False, "error": "比較案は2つ以上入力してください。"}
        _options_text = "\n".join(f"- {o}" for o in _option_lines)
        _context_text = (req.supplement or "（なし）").strip()

        # 2. 第1段階: 軸・weights・scoresのみ取得
        _cmp_prompt1 = f"""戦略コンサルタントとして以下の案を比較する軸とスコアを返せ。JSONのみ出力。

【比較対象案】
{_options_text}
【コンテキスト】
{_context_text}

ルール:
- comparison_axesは入力内容から3〜5軸を動的生成（固定禁止）
- axis_weightsの合計は必ず1.0
- raw_scoresは1-5の整数
- 入力にない数値を創作禁止。不明時は推定不可と記載
- missing_informationは最大5件
- JSONのみ出力。最初の文字は{{

{{
  "comparison_axes": ["軸1","軸2","軸3","軸4","軸5"],
  "axis_weights": {{"軸1":0.30,"軸2":0.20,"軸3":0.20,"軸4":0.15,"軸5":0.15}},
  "options": [
    {{"name":"案名","raw_scores":{{"軸1":4,"軸2":3,"軸3":3,"軸4":2,"軸5":4}}}}
  ],
  "missing_information": ["不足情報1","不足情報2"]
}}"""

        # 3. 第2段階: pros/cons/tradeoff取得
        _cmp_prompt2 = f"""戦略コンサルタントとして以下の各案の長所・短所・推奨ケースを返せ。JSONのみ出力。

【比較対象案】
{_options_text}
【コンテキスト】
{_context_text}

ルール:
- pros/consは各最大2件・20字以内
- recommended_forは最大1件
- tradeoff_analysisは最大3件
- recommendation_basisは50字以内
- confidenceは高/中/低のいずれか
- JSONのみ出力。最初の文字は{{

{{
  "options": [
    {{"name":"案名","pros":["長所1"],"cons":["短所1"],"recommended_for":["ケース"],"confidence":"中"}}
  ],
  "tradeoff_analysis": [{{"axis_a":"軸A","axis_b":"軸B","description":"説明"}}],
  "recommendation_basis": "推奨理由50字以内"
}}"""

        def _parse_llm_json(raw, label=""):
            text = str(raw).strip()
            import re as _r2
            text = _r2.sub(r"^```[a-z]*\n?", "", text)
            text = _r2.sub(r"```$", "", text).strip()
            s = text.find("{")
            e = text.rfind("}")
            if s == -1 or e == -1 or e <= s:
                raise ValueError(f"JSON not found [{label}]: raw_length={len(text)} raw_preview={text[:300]}")
            return _cmp_json.loads(text[s:e+1])

        try:
            # 第1段階呼び出し
            _raw1 = call_llm(
                system_prompt="戦略コンサルタント。JSONのみ出力。最初の文字は{で始めること。",
                messages=[{"role": "user", "content": _cmp_prompt1}],
                ai_tier="core", max_tokens=7000
            )
            _data1 = _parse_llm_json(_raw1, "stage1")
            _axes = [a.strip() for a in _data1.get("comparison_axes", [])][:5]
            _weights_raw = {k.strip(): v for k, v in _data1.get("axis_weights", {}).items()}
            _scores_map = {}
            for _o in _data1.get("options", []):
                _nm = _o.get("name", "").strip()
                _scores_map[_nm] = {k.strip(): v for k, v in _o.get("raw_scores", {}).items()}
            _missing = _data1.get("missing_information", [])[:5]

            # 第2段階呼び出し
            _raw2 = call_llm(
                system_prompt="戦略コンサルタント。JSONのみ出力。最初の文字は{で始めること。",
                messages=[{"role": "user", "content": _cmp_prompt2}],
                ai_tier="core", max_tokens=7000
            )
            _data2 = _parse_llm_json(_raw2, "stage2")
            _detail_map = {}
            for _o in _data2.get("options", []):
                _nm = _o.get("name", "").strip()
                _detail_map[_nm] = _o

            # 4. axis_weights正規化
            _w_total = sum(_weights_raw.values()) if _weights_raw else 0
            if _w_total > 0 and abs(_w_total - 1.0) > 0.01:
                _weights = {k: round(v / _w_total, 4) for k, v in _weights_raw.items()}
            else:
                _weights = _weights_raw


            # 5. サーバー側でweighted_total_score計算
            _RISK_WORDS = ["処理", "切る", "排除", "潰す", "辞めさせる", "飛ばす", "干す", "追い出す"]
            _RISK_MISSING = ["退職条件", "労務リスク", "代替人員計画", "既存人材の貢献度", "顧客影響"]
            _result_options = []
            _best_score = -1
            _best_name = ""
            _all_risk = True  # 全案がrisk_flagかどうか

            for _o1 in _data1.get("options", []):
                _nm = _o1.get("name", "").strip()
                _raw_sc = _scores_map.get(_nm, {})
                _det = _detail_map.get(_nm, {})

                # risk_flag判定
                _risk_flag = any(w in _nm for w in _RISK_WORDS)
                if not _risk_flag:
                    _all_risk = False

                # スコアクランプ1-5
                _clamped = {}
                for _ax in _axes:
                    _v = _raw_sc.get(_ax, 3)
                    try:
                        _clamped[_ax] = max(1, min(5, int(_v)))
                    except Exception:
                        _clamped[_ax] = 3

                # risk_flag=Trueの場合はスコア上限を適用
                if _risk_flag:
                    for _ax in _axes:
                        _ax_l = _ax.lower()
                        if any(k in _ax for k in ["リスク", "実行", "容易"]):
                            _clamped[_ax] = min(_clamped[_ax], 1)
                        elif any(k in _ax for k in ["組織", "活性", "士気", "エンゲージ"]):
                            _clamped[_ax] = min(_clamped[_ax], 2)
                        elif any(k in _ax for k in ["将来", "持続", "成長", "競争"]):
                            _clamped[_ax] = min(_clamped[_ax], 2)
                        elif any(k in _ax for k in ["財務", "収益", "コスト", "インパクト"]):
                            _clamped[_ax] = min(_clamped[_ax], 3)

                # weighted_total_score計算
                _wtotal = 0.0
                for _ax in _axes:
                    _w = _weights.get(_ax, 1.0 / len(_axes) if _axes else 0.2)
                    _wtotal += _clamped.get(_ax, 3) * _w
                _wtotal = round(_wtotal, 3)

                # confidence判定
                if _risk_flag:
                    _conf = "低"
                elif len(_missing) >= 4:
                    _conf = "低"
                elif len(_missing) >= 2:
                    _conf = "中"
                else:
                    _conf = _det.get("confidence", "中")

                # risk_flag時はmissing_informationに追加
                if _risk_flag:
                    for _rm in _RISK_MISSING:
                        if _rm not in _missing:
                            _missing.append(_rm)

                # score_reasonsにrisk_flag警告を追加
                _sc_reasons = {}
                if _risk_flag:
                    for _ax in _axes:
                        _sc_reasons[_ax] = "短期削減効果はあるが、法務・組織毀損リスクにより期待値を減点"

                _result_options.append({
                    "name": _nm,
                    "scores": _clamped,
                    "score_reasons": _sc_reasons,
                    "weighted_total_score": _wtotal,
                    "risk_flag": _risk_flag,
                    "pros": _det.get("pros", [])[:2],
                    "cons": _det.get("cons", [])[:2],
                    "recommended_for": _det.get("recommended_for", [])[:1],
                    "confidence": _conf,
                })

                # risk_flag=Trueの案は推奨候補から除外
                if not _risk_flag and _wtotal > _best_score:
                    _best_score = _wtotal
                    _best_name = _nm

            # 6. Pareto優位分析
            _dominance = []
            for i, _oa in enumerate(_result_options):
                for j, _ob in enumerate(_result_options):
                    if i >= j:
                        continue
                    _a_dom = all(_oa["scores"].get(ax, 3) >= _ob["scores"].get(ax, 3) for ax in _axes)
                    _b_dom = all(_ob["scores"].get(ax, 3) >= _oa["scores"].get(ax, 3) for ax in _axes)
                    if _a_dom and not _b_dom:
                        _dominance.append({"dominant": _oa["name"], "dominated": _ob["name"]})
                    elif _b_dom and not _a_dom:
                        _dominance.append({"dominant": _ob["name"], "dominated": _oa["name"]})

            # 7. 支配軸特定
            _dominant_axis = max(_weights, key=lambda k: _weights.get(k, 0)) if _weights else ""

            # 8. final_recommendation
            if _all_risk:
                _final_rec = "推奨不可。全案にリスク表現が含まれます。代替案設計が必要です。"
            elif len(_missing) >= 5:
                _final_rec = "情報不足のため断定不可。追加データ取得後に再分析を推奨します。"
            elif not _best_name:
                _final_rec = "推奨不可。有効な比較案が見つかりませんでした。"
            else:
                _rec_basis = _data2.get("recommendation_basis", "")
                _final_rec = f"{_best_name}を推奨。{_rec_basis}"

            result = {
                "comparison_axes": _axes,
                "axis_weights": _weights,
                "options": _result_options,
                "tradeoff_analysis": _data2.get("tradeoff_analysis", [])[:3],
                "dominance_analysis": _dominance,
                "missing_information": _missing,
                "recommendation_basis": _data2.get("recommendation_basis", ""),
                "final_recommendation": _final_rec,
            }

            db.collection("tenants").document(tenant_id).collection("consulting_analyses").add({
                "uid": uid, "tenant_id": tenant_id,
                "analysis_type": req.analysis_type,
                "input_text": req.input_text[:500],
                "result": result,
                "created_at": _now_jst().isoformat(),
            })
            return {"ok": True, "result": result, "analysis_type": req.analysis_type}

        except Exception as _cmp_err:
            import traceback
            return {"ok": False, "error": str(_cmp_err), "trace": traceback.format_exc()[-500:]}
    elif req.analysis_type == "contradiction":
        _c_context = req.input_text or ""
        _c_strategy = req.strategy or ""
        _c_policy = req.policy or req.supplement or ""
        prompt = f"""あなたは戦略コンサルタントです。以下の3つの入力から矛盾・トレードオフを検出してください。

【背景・課題・状況】
{_c_context or "（未入力）"}

【目標・戦略】
{_c_strategy or "（未入力）"}

【方針・制約・現在施策】
{_c_policy or "（未入力）"}

【矛盾カテゴリ（必ずこの5種から選ぶこと）】
- 目的手段衝突: 目標と実行手段が噛み合っていない
- KPIズレ: 指標が目標の達成を正しく測れていない
- 前提矛盾: 戦略の前提が現実や他の前提と食い違っている
- 制約衝突: 方針・ルール同士が実行上両立しない
- 実行資源矛盾: 予算・人員・時間と要求水準が整合しない

【矛盾 vs トレードオフの判定基準】
- contradictions（真の矛盾）: 両立が論理的に不可能。一方を実行すると他方が必ず失敗する
- tradeoffs（調整可能な緊張関係）: 両立は可能だが優先順位・比率の調整が必要

【スコア算出ルール（必ず厳守）】
- consistency_score は100点満点から減点方式で算出すること
- 重大矛盾(severity=重大): 1件につき -30点
- 中矛盾(severity=中): 1件につき -15点
- 軽微な齟齬(severity=軽微): 1件につき -7点
- 未解決トレードオフ: 1件につき -5点
- 重大矛盾が1件でもある場合、consistency_score は最大85点までとする
- 「高成長目標」と「投資規模・広告費・人員・実行量の不足」が衝突している場合、consistency_score は原則70点以下とする
- 最低0、最高100

【missing_information ルール】
- 現在売上・LTV・CPA・CVR・客単価・必要新規数・既存顧客比率など定量判定に必要な情報が不足している場合は必ずmissing_informationに入れること
- 不足情報がなければ空配列にすること

【score_basis ルール】
- なぜその点数になったかを減点根拠つきで説明すること
- 例: 「重大矛盾1件(-30点)、中矛盾1件(-15点)、トレードオフ1件(-5点) → 50点」

【confidence 判定ルール（必ず厳守）】
- contradictions には「入力文同士が直接衝突しているもの」または「制約上ほぼ成立不能なもの」のみ入れること
- 定量情報が不足しており矛盾の可能性は高いが断定できないものはdescription内に「定量未検証の高リスク齟齬」と明記すること
- severity が「重大」でも根拠が推定中心の場合はconfidenceを「中」にすること
- 「一般的に」「可能性が高い」だけで重大矛盾と断定しないこと
- confidence は必ず「高」「中」「低」のいずれかで返すこと
  - 高: 入力文に直接矛盾する記述がある
  - 中: 推定・一般論が根拠の中心
  - 低: 情報不足で推定のみ

【fix_direction の記述ルール】
「見直す」「再検討する」「整合させる」等の抽象語禁止。
必ず具体的な数値・対象・アクションを含めること。

【共通ルール】
- 各矛盾には必ず入力文からの根拠引用を含めること
- 矛盾がなければcontradictionsを空配列にすること
- JSON以外の出力禁止。コードブロック禁止。

以下のJSONスキーマで返してください:
{{
  "contradictions": [
    {{
      "type": "目的手段衝突|KPIズレ|前提矛盾|制約衝突|実行資源矛盾 のいずれか",
      "description": "矛盾の具体的な説明（入力内容を踏まえて記述）",
      "strategy_quote": "目標・戦略から引用した根拠箇所",
      "policy_quote": "方針・制約から引用した根拠箇所",
      "context_quote": "背景・状況から引用した根拠箇所（なければ空文字）",
      "why_problematic": "なぜ問題か（因果を明示）",
      "severity": "重大|中|軽微",
      "confidence": "高|中|低（入力根拠が直接的なら高、推定中心なら中、情報不足なら低）",
      "fix_direction": "具体的な修正アクション（数値・対象・手順を含む）"
    }}
  ],
  "tradeoffs": [
    {{
      "description": "トレードオフの内容",
      "tension": "何と何が緊張しているか",
      "recommended_priority": "どちらを優先すべきか・その理由"
    }}
  ],
  "missing_information": ["定量判定に必要だが不足している情報"],
  "score_basis": "減点根拠つきでスコアを説明（例: 重大矛盾1件-30点、中矛盾1件-15点 → 55点）",
  "consistency_score": 100,
  "overall_assessment": {{
    "top_contradiction": "最重要矛盾の一言要約",
    "priority_fix": "最優先で修正すべき箇所",
    "first_action": "最初に直すべき一手（具体的アクション）",
    "summary": "最重要矛盾・優先修正箇所・最初の一手を含む総合評価"
  }}
}}"""
    elif req.analysis_type == "execution":
        # ExecutionGraphEngine経由で生成（LLM自由生成禁止）
        try:
            from api.core.intent import build_execution_plan as _bep_diag
            _ep = _bep_diag(req.input_text, tenant_id, "mixed")
            _tasks = _ep.get("tasks", [])
            result = {
                "summary": _ep.get("summary", req.input_text[:60]),
                "action_plan": [
                    {
                        "task": t.get("objective", ""),
                        "owner": t.get("owner", "未割当"),
                        "deadline": f"{t.get('due_days', 30)}日以内",
                        "kpi": t.get("kpi", "KPI未設定"),
                        "priority": "high" if t.get("priority") == "高" else "medium" if t.get("priority") == "中" else "low",
                        "task_id": t.get("task_id", ""),
                        "phase": t.get("phase", ""),
                        "dependencies": t.get("dependencies", []),
                        "blockers": t.get("blockers", []),
                        "execution_state": t.get("execution_state", "todo"),
                        "priority_score": t.get("priority_score", 0.0),
                    }
                    for t in _tasks
                ],
                "graph": _ep.get("graph", {}),
                "critical_path": _ep.get("critical_path", []),
                "dependencies": _ep.get("dependencies", []),
                "phases": _ep.get("phases", []),
                "kpis": _ep.get("kpis", []),
                "risks": _ep.get("risks", []),
                "blockers": _ep.get("blockers", []),
                "checkpoints": _ep.get("checkpoints", []),
            }
            db.collection("tenants").document(tenant_id).collection("consulting_analyses").add({
                "uid": uid, "tenant_id": tenant_id,
                "analysis_type": req.analysis_type,
                "input_text": req.input_text[:500],
                "result": result,
                "created_at": _now_jst().isoformat(),
            })
            return {"ok": True, "result": result, "analysis_type": req.analysis_type}
        except Exception as _ep_err:
            return {"ok": False, "error": str(_ep_err)}
    else:
        return {"ok": False, "error": f"不明なanalysis_type: {req.analysis_type}"}

    try:
        import re, json as _json
        _max_tok = 4096 if req.analysis_type == "contradiction" else 2048
        res = call_llm(
            system_prompt="あなたは戦略コンサルタントです。必ず有効なJSONオブジェクトのみ返してください。説明文・前置き・Markdownコードブロック・余計なテキストは一切禁止です。最初の文字は必ず{で始めてください。",
            messages=[{"role":"user","content":prompt}],
            ai_tier="core", max_tokens=_max_tok
        )
        res_clean = res.strip()
        if res_clean.startswith("```"):
            import re as _re2
            res_clean = _re2.sub(r"^```[a-z]*\n?", "", res_clean)
            res_clean = _re2.sub(r"```$", "", res_clean).strip()
        # コードブロック除去
        res_clean2 = re.sub(r"```[a-z]*\n?", "", res_clean).replace("```", "").strip()
        m = re.search(r"\{.*\}", res_clean2, re.DOTALL)
        if m:
            try:
                result = _json.loads(m.group(0))
            except Exception:
                if req.analysis_type == "contradiction":
                    _partial = m.group(0)
                    _parsed_cons = []
                    result = {
                        "contradictions": _parsed_cons,
                        "tradeoffs": [],
                        "missing_information": [],
                        "score_basis": "JSON解析エラー",
                        "consistency_score": 50,
                        "overall_assessment": {"top_contradiction": "JSON解析エラー", "priority_fix": "", "first_action": "", "summary": res_clean2[:300]},
                    }
                else:
                    _partial = m.group(0)
                    _tasks = re.findall(r'\{[^{}]*"task"[^{}]*\}', _partial, re.DOTALL)
                    if _tasks:
                        _parsed_tasks = []
                        for _t in _tasks:
                            try:
                                _parsed_tasks.append(_json.loads(_t))
                            except Exception:
                                pass
                        result = {"action_plan": _parsed_tasks} if _parsed_tasks else {"raw": res}
                    else:
                        result = {"raw": res}
        else:
            result = {"raw": res}

        db.collection("tenants").document(tenant_id).collection("consulting_analyses").add({
            "uid": uid, "tenant_id": tenant_id,
            "analysis_type": req.analysis_type,
            "input_text": req.input_text[:500],
            "result": result,
            "created_at": _now_jst().isoformat(),
        })
        return {"ok": True, "result": result, "analysis_type": req.analysis_type}
    except Exception as e:
        return {"ok": False, "error": str(e)}

@router.get("/consult/history")
def get_consult_history(analysis_type: str = "", payload: dict = Depends(verify_token)):
    uid = payload["uid"]
    tenant_id = payload.get("tenant_id", "default")
    db = get_db()
    try:
        docs = list(
            db.collection("tenants").document(tenant_id)
            .collection("consulting_analyses")
            .where("uid","==",uid)
            .limit(20)
            .stream()
        )
        analyses = [{**d.to_dict(), "doc_id": d.id} for d in docs]
        if analysis_type:
            analyses = [a for a in analyses if a.get("analysis_type") == analysis_type]
        analyses.sort(key=lambda x: x.get("created_at",""), reverse=True)
        return {"analyses": analyses}
    except Exception:
        return {"analyses": []}


@router.delete("/consult/delete/{doc_id}")
def delete_consult_analysis(doc_id: str, payload: dict = Depends(verify_token)):
    uid = payload["uid"]
    tenant_id = payload.get("tenant_id", "default")
    db = get_db()
    try:
        ref = db.collection("tenants").document(tenant_id).collection("consulting_analyses").document(doc_id)
        doc = ref.get()
        if not doc.exists:
            return {"ok": False, "error": "not found"}
        if doc.to_dict().get("uid") != uid:
            return {"ok": False, "error": "forbidden"}
        ref.delete()
        return {"ok": True}
    except Exception as e:
        return {"ok": False, "error": str(e)}

@router.get("/frameworks")
def get_frameworks(payload: dict = Depends(verify_token)):
    tenant_id = payload.get("tenant_id", "default")
    db = get_db()
    try:
        docs = list(db.collection("tenants").document(tenant_id).collection("consulting_frameworks").stream())
        fw = [d.to_dict() for d in docs]
        if not fw:
            fw = [
                {"name":"MECE","description":"ダブりなく、漏れなく","active":True},
                {"name":"SWOT","description":"強み・弱み・機会・脅威","active":True},
                {"name":"3C","description":"顧客・競合・自社","active":True},
                {"name":"ロジックツリー","description":"問題を論理的に分解","active":True},
                {"name":"Issue Tree","description":"課題を階層的に整理","active":True},
            ]
        return {"frameworks": fw}
    except Exception:
        return {"frameworks": []}

@router.post("/weekly_report")
def generate_weekly_report(body: dict = {}, payload: dict = Depends(verify_token)):
    from api.core.features import is_feature_enabled
    from api.core.llm_client import call_llm_pro as _clp2
    uid = payload["uid"]
    tenant_id = payload.get("tenant_id", "default")
    if not is_feature_enabled(uid, "fixed_concept_report"):
        raise HTTPException(status_code=403, detail="固定概念レポートは現在未開放のため使用できません。")
    db = get_db()
    n_chats = int((body or {}).get("n_chats", 30))
    try:
        history = _load_chat_history_across_sessions(uid, tenant_id, n_chats)
        if not history:
            return {"ok": False, "error": "チャット履歴がありません"}
        chat_text = "\n".join([f"{m['role']}: {m.get('content','')}" for m in history if m.get("content")])[:4000]
        ts = _now_jst().strftime("%Y-%m-%d %H:%M")
        prompt = f"""以下はユーザーのチャット履歴です。週次戦術レポートを生成してください。

{chat_text}

# 週次戦術レポート
生成日時: {ts}
---
## 今週の主要相談テーマ
[最も多かった相談テーマ上位3件]
---
## 意思決定パターン分析
[どのような意思決定が多かったか]
---
## 成長トレンド
[前週比での変化・改善点]
---
## 来週の優先アクション
[最重要アクション3件]
---
## 戦略的提言
[中長期視点での提言]
"""
        report = _clp2(
            system_prompt="あなたは戦略コンサルタント。週次レポートを指定フォーマット厳守で生成せよ。",
            messages=[{"role":"user","content":prompt}],
            max_tokens=3000,
        )
        db.collection("weekly_reports").add({
            "uid": uid, "tenant_id": tenant_id,
            "report_md": report,
            "created_at": _now_jst().isoformat(),
        })
        return {"ok": True, "report_md": report}
        try:
            import datetime as _ugd_dt
            from api.core.firestore_client import get_db as _ugd_db
            _ugd_db().collection("usage_logs").add({"user_id":uid,"tenant_id":tenant_id,"purpose_mode":"current_issue_diagnosis","diagnosis_type":"current_issue_diagnosis","prompt":"","timestamp":(_ugd_dt.datetime.utcnow()+_ugddatetime.timedelta(hours=9)).strftime("%Y-%m-%d %H:%M:%S"),"is_admin_test":False})
        except Exception:
            pass
    except Exception as e:
        return {"ok": False, "error": str(e)}

@router.get("/weekly_report/list")
def list_weekly_reports(payload: dict = Depends(verify_token)):
    uid = payload["uid"]
    tenant_id = payload.get("tenant_id", "default")
    db = get_db()
    try:
        docs = list(
            db.collection("weekly_reports")
            .where("uid","==",uid)
            .where("tenant_id","==",tenant_id)
            .limit(10)
            .stream()
        )
        reports = [d.to_dict() for d in docs]
        reports.sort(key=lambda x: x.get("created_at",""), reverse=True)
        return {"reports": reports}
    except Exception:
        return {"reports": []}


@router.post("/file_diagnosis")
async def file_diagnosis(
    file: UploadFile = File(...),
    answer_context: str = Form(""),
    payload: dict = Depends(verify_token)
):
    """ファイル全タブ横断診断→構造診断・課題仮説・実行計画を一括生成"""
    from api.core.features import is_feature_enabled
    _uid_fd = payload.get("uid", "")
    if not is_feature_enabled(_uid_fd, "diag_file"):
        raise HTTPException(status_code=403, detail="ファイル診断は現在未開放のため使用できません。")
    from api.core.llm_client import call_llm as _cllm
    import io, re as _re, json as _json

    filename = file.filename or "file"
    content = await file.read()
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    # 全シート抽出
    sheets_text = {}
    sheet_dfs = {}
    try:
        if ext in ("xlsx", "xls"):
            import pandas as pd
            xf = pd.ExcelFile(io.BytesIO(content))
            for sheet in xf.sheet_names:
                try:
                    df_raw = xf.parse(sheet, header=None).dropna(how="all").dropna(axis=1, how="all")
                    sheets_text[sheet] = df_raw.to_csv(index=False, header=False)[:4000] if not df_raw.empty else f"({sheet}:データなし)"
                    sheet_dfs[sheet] = df_raw
                except Exception as _e:
                    sheets_text[sheet] = f"({sheet}:読み込みエラー:{_e})"
        elif ext == "ods":
            import pandas as pd
            xf = pd.ExcelFile(io.BytesIO(content), engine="odf")
            for sheet in xf.sheet_names:
                try:
                    df_raw = xf.parse(sheet, header=None).dropna(how="all").dropna(axis=1, how="all")
                    sheets_text[sheet] = df_raw.to_csv(index=False, header=False)[:4000] if not df_raw.empty else f"({sheet}:データなし)"
                    sheet_dfs[sheet] = df_raw
                except Exception as _e:
                    sheets_text[sheet] = f"({sheet}:読み込みエラー:{_e})"
        elif ext == "csv":
            import pandas as pd
            df = pd.read_csv(io.BytesIO(content))
            sheets_text["Sheet1"] = df.to_csv(index=False)[:6000]
            sheet_dfs["Sheet1"] = df
        elif ext == "pdf":
            try:
                import pypdf
                reader = pypdf.PdfReader(io.BytesIO(content))
                text = "\n".join(p.extract_text() or "" for p in reader.pages)
                sheets_text["PDF"] = text[:6000]
            except Exception:
                sheets_text["PDF"] = content.decode("utf-8", errors="ignore")[:6000]
        else:
            sheets_text["TEXT"] = content.decode("utf-8", errors="ignore")[:6000]
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"ファイル読み込みエラー: {e}")

    if not sheets_text:
        raise HTTPException(status_code=400, detail="ファイルの内容を読み取れませんでした")

    # 数式・グラフ・条件付き書式・ピボットテーブル読み込み
    formula_summary = ""
    if ext in ("xlsx", "xls"):
        try:
            import openpyxl as _opxl
            wb = _opxl.load_workbook(io.BytesIO(content), data_only=False)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                section = ""
                # 全数式取得（上限なし）
                formulas = []
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                            formulas.append(f"{cell.coordinate}: {cell.value}")
                if formulas:
                    shown = formulas[:30]
                    section += f"\n【{sheet_name}】数式({len(formulas)}件・表示は先頭{len(shown)}件):\n" + "\n".join(shown)
                # 条件付き書式
                cf_list = []
                for cf_range, cf_rules in ws.conditional_formatting._cf_rules.items():
                    for rule in cf_rules:
                        rule_type = getattr(rule, "type", "")
                        formula_val = getattr(rule, "formula", [])
                        cf_list.append(f"範囲{cf_range}: タイプ={rule_type} 条件={formula_val}")
                if cf_list:
                    section += f"\n【{sheet_name}】条件付き書式({len(cf_list)}件):\n" + "\n".join(cf_list[:20])
                # グラフ
                charts = []
                for chart in ws._charts:
                    title = ""
                    try: title = chart.title.tx.rich.p[0].r[0].t
                    except Exception: pass
                    chart_type = type(chart).__name__
                    charts.append(f"グラフ種別={chart_type} タイトル={title or '(無題)'}")
                if charts:
                    section += f"\n【{sheet_name}】グラフ({len(charts)}件):\n" + "\n".join(charts)
                # ピボットテーブル
                pivots = []
                for pt in getattr(ws, "_pivots", []):
                    pt_name = getattr(pt, "name", "")
                    pivots.append(f"ピボット名={pt_name}")
                if pivots:
                    section += f"\n【{sheet_name}】ピボットテーブル({len(pivots)}件):\n" + "\n".join(pivots)
                if section:
                    formula_summary += section
        except Exception as _fe:
            formula_summary = f"（xlsx詳細読み込みエラー: {_fe}）"
    elif ext == "ods":
        try:
            from odf.opendocument import load as _odf_load
            from odf.table import Table, TableRow, TableCell
            doc = _odf_load(io.BytesIO(content))
            for sheet in doc.spreadsheet.getElementsByType(Table):
                sheet_name = sheet.getAttribute("name")
                formulas = []
                for row in sheet.getElementsByType(TableRow):
                    for cell in row.getElementsByType(TableCell):
                        formula = cell.getAttribute("formula")
                        if formula:
                            formulas.append(f"{sheet_name}: {formula}")
                if formulas:
                    shown = formulas[:30]
                    formula_summary += f"\n【{sheet_name}】数式({len(formulas)}件・表示は先頭{len(shown)}件):\n" + "\n".join(shown)
        except Exception as _fe:
            formula_summary = f"（ods数式読み込みエラー: {_fe}）"


    formula_summary = formula_summary[:2500]

    # 全シートを結合
    combined = ""
    for sheet, text in sheets_text.items():
        combined += f"\n\n【シート: {sheet}】\n{text}"
    combined = combined[:6000]

    # sheet_schema 生成
    sheet_schema = []
    try:
        import json as _json_ss
        for _sname, _sdf in sheet_dfs.items():
            try:
                _unnamed = [c for c in _sdf.columns if "Unnamed" in str(c)]
                _numeric = [c for c in _sdf.select_dtypes(include="number").columns.tolist() if "Unnamed" not in str(c)]
                _text = [c for c in _sdf.select_dtypes(include="object").columns.tolist() if "Unnamed" not in str(c)]
                sheet_schema.append({
                    "sheet_name": _sname,
                    "row_count": int(len(_sdf)),
                    "col_count": int(len(_sdf.columns)),
                    "columns": [str(c) for c in _sdf.columns.tolist()],
                    "dtypes": {str(c): str(t) for c, t in _sdf.dtypes.items()},
                    "non_null_counts": {str(c): int(_sdf[c].count()) for c in _sdf.columns},
                    "sample_rows": _sdf.head(5).fillna("").astype(str).to_dict(orient="records"),
                    "unnamed_columns": _unnamed,
                    "numeric_columns": _numeric,
                    "text_columns": _text,
                })
            except Exception:
                pass
    except Exception:
        pass
    sheet_schema_json = _json.dumps(sheet_schema, ensure_ascii=False, indent=2)[:3000] if sheet_schema else ""

    # ===== Workbook構造解析 =====
    workbook_structure = []
    if ext in ("xlsx", "xls"):
        try:
            import json as _json_wb
            _wb2 = _opxl.load_workbook(io.BytesIO(content), data_only=False) if 'wb' not in dir() else wb
            try:
                _wb2 = _opxl.load_workbook(io.BytesIO(content), data_only=False)
            except Exception:
                _wb2 = None
            if _wb2:
                for _wsn in _wb2.sheetnames:
                    _ws2 = _wb2[_wsn]
                    _sheet_struct = {"sheet_name": _wsn, "merged_cells": [], "table_blocks": [], "formula_refs": [], "sheet_relations": [], "cell_samples": []}
                    # A. merged_cells
                    try:
                        for _mc in list(_ws2.merged_cells.ranges)[:50]:
                            _mc_cell = _ws2.cell(_mc.min_row, _mc.min_col)
                            _sheet_struct["merged_cells"].append({"range": str(_mc), "value": str(_mc_cell.value or "")})
                    except Exception:
                        pass
                    # B. table_blocks (空白行で分割)
                    try:
                        _rows_data = []
                        for _row in _ws2.iter_rows(max_row=500):
                            _row_vals = [str(c.value) if c.value is not None else "" for c in _row]
                            _rows_data.append(_row_vals)
                        _block_start = None
                        _blocks = []
                        for _ri, _rv in enumerate(_rows_data):
                            _has_val = any(v.strip() for v in _rv)
                            if _has_val and _block_start is None:
                                _block_start = _ri
                            elif not _has_val and _block_start is not None:
                                _blocks.append((_block_start, _ri - 1))
                                _block_start = None
                        if _block_start is not None:
                            _blocks.append((_block_start, len(_rows_data) - 1))
                        for _bs, _be in _blocks[:20]:
                            _preview = [r for r in _rows_data[_bs:_bs+1]]
                            _approx_cols = max((len([v for v in r if v.strip()]) for r in _rows_data[_bs:_be+1]), default=0)
                            _sheet_struct["table_blocks"].append({"start_row": _bs+1, "end_row": _be+1, "approx_cols": _approx_cols, "preview_rows": _preview})
                    except Exception:
                        pass
                    # C. formula_refs
                    try:
                        import re as _re_wb
                        for _row in _ws2.iter_rows():
                            for _cell in _row:
                                if _cell.value and isinstance(_cell.value, str) and _cell.value.startswith("="):
                                    _refs = _re_wb.findall(r"([A-Za-z぀-鿿゠-ヿ]+)!", _cell.value)
                                    _sheet_struct["formula_refs"].append({"cell": _cell.coordinate, "formula": _cell.value, "referenced_sheets": list(set(_refs))})
                                    if len(_sheet_struct["formula_refs"]) >= 50:
                                        break
                            if len(_sheet_struct["formula_refs"]) >= 50:
                                break
                    except Exception:
                        pass
                    # D. sheet_relations
                    try:
                        _all_refs = []
                        for _fr in _sheet_struct["formula_refs"]:
                            _all_refs.extend(_fr.get("referenced_sheets", []))
                        _sheet_struct["sheet_relations"] = [{"sheet": _wsn, "references": list(set(_all_refs))}]
                    except Exception:
                        pass
                    # E. cell_samples (左上20x10)
                    try:
                        for _row in _ws2.iter_rows(max_row=20, max_col=10):
                            for _cell in _row:
                                if _cell.value is not None:
                                    _sheet_struct["cell_samples"].append({"cell": _cell.coordinate, "value": str(_cell.value)[:100]})
                                    if len(_sheet_struct["cell_samples"]) >= 30:
                                        break
                            if len(_sheet_struct["cell_samples"]) >= 30:
                                break
                    except Exception:
                        pass
                    workbook_structure.append(_sheet_struct)
        except Exception as _wb_e:
            workbook_structure = []
    workbook_structure_json = _json.dumps(workbook_structure, ensure_ascii=False, indent=2)[:4000] if workbook_structure else ""

    # ===== ③ Python先行数値分析 =====
    import numpy as _np
    import re as _re2, json as _json2

    numeric_analysis = {}
    for sheet, text in sheets_text.items():
        try:
            import pandas as _pd, io as _io
            lines = text.strip().split("\n")
            if len(lines) < 2:
                continue
            df = _pd.read_csv(_io.StringIO(text))

            # Unnamed列・変化率計算不要列を除外
            valid_cols = [c for c in df.columns if "Unnamed" not in str(c)]
            df = df[valid_cols]

            sheet_stats = {"行数": len(df), "列数": len(df.columns), "有効列": valid_cols}

            # 数値列の基本統計（意味のある列のみ）
            num_cols = [c for c in df.select_dtypes(include=["number"]).columns.tolist()
                       if "Unnamed" not in str(c)]

            for col in num_cols[:15]:
                s = df[col].dropna()
                if len(s) > 0 and s.sum() > 0:  # 全0列はスキップ
                    sheet_stats[col] = {
                        "合計": round(float(s.sum()), 2),
                        "平均": round(float(s.mean()), 2),
                        "最大": round(float(s.max()), 2),
                        "最小": round(float(s.min()), 2),
                        "データ数": int(s.count()),
                    }

            # 上位・下位パフォーマー検出（異常値ではなく業界文脈で解釈）
            performers = []
            for col in num_cols[:8]:
                s = df[col].dropna()
                if len(s) >= 4 and s.sum() > 0:
                    top = s.nlargest(3)
                    if top.iloc[0] > s.mean() * 2:
                        performers.append(f"{col}: 上位集中={top.values.tolist()}（平均{round(float(s.mean()),1)}の{round(top.iloc[0]/s.mean(),1)}倍）")
            if performers:
                sheet_stats["上位集中検出"] = performers

            numeric_analysis[sheet] = sheet_stats
        except Exception as _e:
            numeric_analysis[sheet] = {"エラー": str(_e)}

    numeric_summary = _json2.dumps(numeric_analysis, ensure_ascii=False, indent=2)

    # ===== ② Chain of Thought: 4段階順次分析 =====
    # 確認済み内容から業種を抽出してシステムプロンプトに動的反映
    # answer_context全文をそのまま使用（切り捨て・抽出なし）
    full_context = answer_context if answer_context else ""
    industry_hint = ""
    if full_context:
        industry_hint = f"\n【事前確認済みの全回答（以下を必ず診断に100%反映せよ・推測補完禁止）】\n{full_context}"

    system = f"""あなたは超一流の経営コンサルタントかつデータ分析の専門家である。{industry_hint}

以下のルールを必ず守れ：
- 事前確認済みの回答に含まれる情報は全て確定事実として扱え。「おそらく」「はずだが」「不明」は絶対禁止
- 確認済みの業種・業界の文脈で全ての数値・用語を解釈せよ
- 提供された数値分析結果を必ず引用して根拠とせよ
- 「具体的数値は集計が必要」「仮定」等の逃げ回答は絶対禁止
- 推測・一般論ではなくデータから読み取れる事実のみを述べよ
- KPIは実データから算出した根拠ある数値のみ使え。データにない数字を作るな
- 業界の慣習・用語・ビジネスモデルを踏まえた専門的な解釈をせよ
- 「上位集中」として検出された数値はトップパフォーマーの正常な実績である。異常値として扱うな
- KPIは実データから算出した数値のみ使え。「現状データなし」と書く場合は実データから推計せよ
- 空白セル・列構成の違いは確認済みの業種・業務文脈で解釈せよ。安易に「入力ミス」「欠損」と断定するな"""

    context_str = f"""【最重要：事前確認済みの用語・業界定義（必ず全て診断に反映せよ）】
{full_context if full_context else "（確認情報なし）"}

【Excelの数式（条件・計算ロジック自動読み取り）】
{formula_summary if formula_summary else "（数式なし/odsファイル）"}

【Python数値分析結果】
{numeric_summary}

【シート構造スキーマ（自動解析）】
{sheet_schema_json if sheet_schema_json else "（構造スキーマなし）"}

【Workbook構造解析】
{workbook_structure_json if workbook_structure_json else "（Workbook構造なし）"}

【生データ（全シート）】
{combined[:8000]}"""

    # 1回のultra呼び出しで全診断を生成
    import re as _re2, json as _json2
    diagnosis_raw = _cllm(
        system_prompt=system,
        messages=[{"role":"user","content":f"""{context_str}

以下のJSONのみで返せ（コードブロック不要）:
{{
  "overview": "現状把握。数値分析結果を必ず引用。600字以内。",
  "structure": "データ構造診断。シート構成・シート間関係・異常値を具体的に。800字以内。",
  "issues": "課題仮説（最大3個）。各仮説に根拠数値・影響・優先度・推奨アクションを含める。1200字以内。",
  "action_plan": "実行計画（最大3個）。各アクションに内容・期限・担当・KPI・期待効果を含める。1200字以内。",
  "key_metrics": "重要指標・数値（箇条書き5件以内）。600字以内。",
  "risks": "リスク・警告事項（箇条書き3件以内）。600字以内。"
}}

条件:
- 数値分析結果を必ず引用
- データにない数値を作らない
- 一般論禁止
- 確認済みの業種・業界文脈で解釈せよ
"""}],
        ai_tier="ultra",
        max_tokens=6000
    )
    try:
        _m = _re2.search(r'\{.*\}', diagnosis_raw, _re2.DOTALL)
        if not _m:
            raise HTTPException(status_code=500, detail=f"診断JSON取得失敗: {diagnosis_raw[:500]}")
        diag = _json2.loads(_m.group(0))
    except HTTPException:
        raise
    except Exception as _je:
        raise HTTPException(status_code=500, detail=f"診断JSON解析失敗: {str(_je)} / raw: {diagnosis_raw[:500]}")
    step1 = diag.get("overview", "")
    step2 = diag.get("structure", "")
    step3 = diag.get("issues", "")
    step4 = diag.get("action_plan", "")
    step5 = {"key_metrics": diag.get("key_metrics", ""), "risks": diag.get("risks", "")}

    try:
        _db = get_db()
        import datetime as _dtnow
        _db.collection("file_diagnoses").add({"uid":payload["uid"],"tenant_id":payload.get("tenant_id","default"),"filename":filename,"sheets":list(sheets_text.keys()),"overview":step1,"structure":step2,"issues":step3,"action_plan":step4,"key_metrics":step5.get("key_metrics",""),"risks":step5.get("risks",""),"created_at":_dtnow.datetime.now().isoformat(),"diagnosis_type":"file"})
        try:
            import datetime as _ufd_dt
            _uid_fl = payload["uid"]; _tid_fl = payload.get("tenant_id","default")
            from api.core.firestore_client import get_db as _ufd_gdb
            _ufd_gdb().collection("usage_logs").add({"user_id":_uid_fl,"tenant_id":_tid_fl,"purpose_mode":"file_diagnosis","diagnosis_type":"file_diagnosis","prompt":filename,"timestamp":(_ufd_dt.datetime.utcnow()+_ufddatetime.timedelta(hours=9)).strftime("%Y-%m-%d %H:%M:%S"),"is_admin_test":False})
        except Exception:
            pass
    except Exception:
        pass
    return {
        "ok": True,
        "filename": filename,
        "sheets": list(sheets_text.keys()),
        "overview": step1,
        "structure": step2,
        "issues": step3,
        "action_plan": step4,
        "key_metrics": step5.get("key_metrics", ""),
        "risks": step5.get("risks", ""),
        "numeric_analysis": numeric_analysis,
        "workbook_structure": workbook_structure,
    }


@router.post("/file_followup")
async def file_followup(body: dict = Body(...), payload: dict = Depends(verify_token)):
    """ファイル診断結果への追加質問・深掘り分析"""
    from api.core.llm_client import call_llm as _cllm
    question = body.get("question", "").strip()
    context = body.get("context", "")
    filename = body.get("filename", "")
    if not question:
        raise HTTPException(status_code=400, detail="質問が必要です")

    system = """あなたは超一流の経営コンサルタントかつデータ分析の専門家である。
以下のルールを必ず守れ：

【最重要ルール】
1. 質問に業界特有の専門用語・略語・固有名詞が含まれており、その意味がデータから判断できない場合は、
   回答する前に必ず「〇〇とはどういう意味ですか？」と質問し、確認してから回答せよ。
2. わかったふりをして回答することは絶対禁止。不明な点は必ず確認せよ。
3. データに基づいた具体的な回答のみ出力せよ。拒否・曖昧回答・一般論のみは禁止。
4. 数値は必ず引用し、比較・変化率・傾向を明示せよ。"""

    prompt = f"""ファイル「{filename}」の診断結果：
{context[:6000]}

ユーザーの追加質問：
{question}

【指示】
- 質問に不明な専門用語・略語・業界用語が含まれる場合は、まず「〇〇とはどういう意味ですか？」と確認せよ
- 意味が明確な場合は、診断結果のデータを根拠に具体的・詳細に回答せよ
- 数値は必ず引用すること"""

    try:
        answer = _cllm(
            system_prompt=system,
            messages=[{"role":"user","content":prompt}],
            ai_tier="ultra", max_tokens=2048
        )
        return {"ok": True, "answer": answer}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/file_diagnosis_check")
async def file_diagnosis_check(
    file: UploadFile = File(...),
    payload: dict = Depends(verify_token)
):
    """ファイルをスキャンして不明な専門用語・業界用語があれば質問を返す"""
    from api.core.llm_client import call_llm as _cllm
    import io, re as _re, json as _json

    filename = file.filename or "file"
    content = await file.read()
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    sheets_text = {}
    sheet_dfs = {}
    try:
        if ext in ("xlsx", "xls"):
            import pandas as pd
            xf = pd.ExcelFile(io.BytesIO(content))
            for sheet in xf.sheet_names:
                df = xf.parse(sheet)
                sheet_dfs[sheet] = df
                sheets_text[sheet] = df.to_csv(index=False)[:2000]
        elif ext == "ods":
            import pandas as pd
            xf = pd.ExcelFile(io.BytesIO(content), engine="odf")
            for sheet in xf.sheet_names:
                df = xf.parse(sheet)
                sheet_dfs[sheet] = df
                sheets_text[sheet] = df.to_csv(index=False)[:2000]
        elif ext == "csv":
            import pandas as pd
            df = pd.read_csv(io.BytesIO(content))
            sheet_dfs["Sheet1"] = df
            sheets_text["Sheet1"] = df.to_csv(index=False)[:4000]
        else:
            sheets_text["TEXT"] = content.decode("utf-8", errors="ignore")[:4000]
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"ファイル読み込みエラー: {e}")

    combined = ""
    for sheet, text in sheets_text.items():
        combined += f"\n\n【シート: {sheet}】\n{text}"
    combined = combined[:6000]

    # sheet_schema 生成
    sheet_schema = []
    try:
        import pandas as _pd2
        for sheet, df in sheet_dfs.items():
            try:
                unnamed_cols = [c for c in df.columns if "Unnamed" in str(c)]
                numeric_cols = [c for c in df.select_dtypes(include="number").columns.tolist() if "Unnamed" not in str(c)]
                text_cols = [c for c in df.select_dtypes(include="object").columns.tolist() if "Unnamed" not in str(c)]
                schema = {
                    "sheet_name": sheet,
                    "row_count": int(len(df)),
                    "col_count": int(len(df.columns)),
                    "columns": [str(c) for c in df.columns.tolist()],
                    "dtypes": {str(c): str(t) for c, t in df.dtypes.items()},
                    "non_null_counts": {str(c): int(df[c].count()) for c in df.columns},
                    "sample_rows": df.head(5).fillna("").astype(str).to_dict(orient="records"),
                    "unnamed_columns": unnamed_cols,
                    "numeric_columns": numeric_cols,
                    "text_columns": text_cols,
                }
                sheet_schema.append(schema)
            except Exception:
                pass
    except Exception:
        pass

    sheet_schema_json = _json.dumps(sheet_schema, ensure_ascii=False)[:4000]

    # Unnamed列注記（決め打ちせず可能性として説明）
    unnamed_note = ""
    for s in sheet_schema:
        if s.get("unnamed_columns"):
            unnamed_note += f"\n※{s['sheet_name']}シートに'Unnamed:数字'列があります。Excelの結合セル・空白ヘッダー・階層ヘッダー等により発生している可能性があります。意味の確定はユーザーへの確認が必要です。"

    check_prompt = f"""以下のファイルデータを分析する前に、コンサルタントとして正確な診断に必要な情報を収集するための質問リストを作成せよ。

【ファイル: {filename}】
{combined}

【シートスキーマ（自動解析）】
{sheet_schema_json}

【Unnamed列情報】
{unnamed_note if unnamed_note else "Unnamed列なし"}

【指示】
必ず以下の観点で質問を作成せよ：
1. 業種・業態・ビジネスモデル（何の事業か）
2. このデータで解決したい課題・目的
3. 重要なKPI・目標値（あれば）
4. データ内に意味が不明確な業界固有の専門用語・略語・記号・独自コード
5. Unnamed列が存在する場合はその列の意味・用途

制約：
- 一般的なビジネス用語・数字・日付は含めるな
- 質問は簡潔に最大5件以内

出力形式（JSONのみ）：
{{"need_clarification": true, "questions": ["Q1: ...", "Q2: ..."], "unknown_terms": []}}"""

    try:
        raw = _cllm(
            system_prompt="データ分析の専門家。JSONのみ出力。",
            messages=[{"role":"user","content":check_prompt}],
            ai_tier="core", max_tokens=1024
        )
        m = _re.search(r'\{.*\}', raw, _re.DOTALL)
        result = _json.loads(m.group(0)) if m else {"need_clarification": True, "questions": ["Q1: このファイルの業種・業態を教えてください。","Q2: このファイルで診断したい目的・課題を教えてください。","Q3: 重要なKPIや目標値があれば教えてください。"], "unknown_terms": []}
        result["filename"] = filename
        result["sheets"] = list(sheets_text.keys())
        result["file_data"] = combined[:3000]
        result["sheet_schema"] = sheet_schema
        return result
    except Exception:
        return {
            "need_clarification": True,
            "questions": [
                "Q1: このファイルの業種・業態を教えてください。",
                "Q2: このファイルで診断したい目的・課題を教えてください。",
                "Q3: 重要なKPIや目標値があれば教えてください。",
            ],
            "unknown_terms": [],
            "filename": filename,
            "sheets": list(sheets_text.keys()),
            "file_data": combined[:3000],
            "sheet_schema": sheet_schema,
        }


@router.post("/file_clarify")
async def file_clarify(body: dict = Body(...), payload: dict = Depends(verify_token)):
    """専門用語確認・双方向チャットフロー
    AIがユーザーに質問し、ユーザーもAIに質問できる。
    両方が納得したら診断開始。
    """
    from api.core.llm_client import call_llm as _cllm
    import json as _json, re as _re

    messages = body.get("messages", [])
    file_summary = body.get("file_summary", "")
    user_message = body.get("user_message", "")

    system = """あなたは超一流の経営コンサルタントであり、ファイル分析の専門家である。

【役割】
ユーザーのファイルを正確に診断するため、診断前に業界・業務文脈を徹底的に収集する。

【最重要ルール】
1. 最優先で以下を確認せよ：業種・業態・ビジネスモデル、このデータで解決したい課題・目的、重要KPI・目標値
2. 次にデータ内の業界固有の専門用語・略語・独自コード・記号の意味を確認せよ
3. ユーザーの回答からビジネスモデルを深く理解し、必要に応じて追加質問をせよ
4. ユーザーから質問された場合は、コンサルタントとして誠実・具体的に回答せよ
5. わかったふりは絶対禁止。業界文脈が不明なまま診断に進むな
6. 不明点があっても「情報が足りないから分析できない」は絶対禁止。現状の情報で分析できる範囲を最大化しろ
7. 追加質問は最大3件以内に絞れ。ユーザーの負担を最小化せよ
8. 業種・目的・KPI・専門用語の主要項目が確認できたと判断したら、末尾に「[診断準備完了]」を付加せよ

【出力形式】
- 通常の確認・回答: そのまま日本語で出力
- 準備完了時: 回答の末尾に必ず「[診断準備完了]」を付加
"""

    # 会話履歴を構築
    chat_messages = []
    for msg in messages:
        chat_messages.append({"role": msg["role"], "content": msg["content"]})

    # 最新のユーザーメッセージを追加
    if user_message:
        chat_messages.append({"role": "user", "content": user_message})

    # ファイル概要をコンテキストとして最初のメッセージに追加
    context = f"""【分析対象ファイルの概要】
{file_summary}

【会話の目的】
このファイルを正確に診断するため、業種・目的・KPI・業界固有の専門用語・独自ルールを確認する。
ユーザーへの質問と、ユーザーからの質問への回答を行う。"""

    if chat_messages and chat_messages[0]["role"] == "user":
        chat_messages[0]["content"] = context + "\n\n" + chat_messages[0]["content"]
    else:
        chat_messages.insert(0, {"role": "user", "content": context})

    try:
        response = _cllm(
            system_prompt=system,
            messages=chat_messages,
            ai_tier="core",
            max_tokens=1024
        )
        is_ready = "[診断準備完了]" in response
        clean_response = response.replace("[診断準備完了]", "").strip()

        return {
            "ok": True,
            "message": clean_response,
            "is_ready": is_ready
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/file_clarify_save")
async def file_clarify_save(body: dict = Body(...), payload: dict = Depends(verify_token)):
    """ファイル種別ごとの確認済み用語・背景情報をFirestoreに保存"""
    from api.core.firestore_client import get_db
    uid = payload["uid"]
    file_key = body.get("file_key", "")  # ファイル名ベースのキー
    context = body.get("context", {})    # {term: explanation} の辞書

    if not file_key or not context:
        raise HTTPException(status_code=400, detail="file_keyとcontextが必要です")

    try:
        db = get_db()
        db.collection("users").document(uid).collection("file_contexts").document(file_key).set({
            "context": context,
            "updated_at": __import__("datetime").datetime.utcnow().isoformat()
        })
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/file_clarify_load")
async def file_clarify_load(file_key: str, payload: dict = Depends(verify_token)):
    """保存済みの確認内容を読み込む"""
    from api.core.firestore_client import get_db
    uid = payload["uid"]
    try:
        db = get_db()
        doc = db.collection("users").document(uid).collection("file_contexts").document(file_key).get()
        if doc.exists:
            return {"ok": True, "context": doc.to_dict().get("context", {}), "found": True}
        return {"ok": True, "context": {}, "found": False}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class FutureSimRequest(BaseModel):
    message: str
    ai_tier: str = "core"

@router.post("/future_simulation")
def future_simulation(req: FutureSimRequest, payload: dict = Depends(verify_token)):
    """未来分岐シミュレーター"""
    from api.core.features import is_feature_enabled
    if not is_feature_enabled(payload.get("uid", ""), "diag_future"):
        raise HTTPException(status_code=403, detail="未来分岐シミュレーターは現在未開放のため使用できません。")
    import re as _re, json as _json
    uid = payload.get("uid")
    if not uid:
        raise HTTPException(status_code=401, detail="認証情報が不正です")
    from api.core.features import get_effective_feature_flags
    features = get_effective_feature_flags(uid)
    if not features.get("diag_future", False):
        raise HTTPException(status_code=403, detail="未来分岐シミュレーターは未開放です")
    system_prompt = """あなたは因果構造ベースの戦略シミュレーションAIです。
「未来を想像する」のではなく、因果・資源・時間・実行可能性・崩壊率・継続率を計算する戦略シミュレーションとして振る舞ってください。

【厳守ルール】
- A/B/C/Dの4分岐は独立生成禁止。必ず「現状→選択→中間状態→最終到達点」の因果連鎖として生成すること
- branches は必ず4件。役割は固定:
    A = 放置・縮小均衡系（現状維持・何もしない場合）
    B = 部分改善系（一部だけ手を打つ場合）
    C = 構造改善系（根本から変える場合）
    D = 高リスク転換系（大胆に方向転換する場合）
    ただし label は入力内容に合わせて自然な表現にすること
- success_rate はLLMの感覚値禁止。以下5軸の総合評価として算出すること:
    1. execution_difficulty（実行難易度: 低いほど高得点）
    2. continuity_load（継続負荷: 低いほど高得点）
    3. required_resources（資源要求: 少ないほど高得点）
    4. market_fit（市場適合: 高いほど高得点）
    5. current_fit（現状整合性: 高いほど高得点）
    score_basis は省略禁止。各branchのsuccess_rateの直後に必ず出力すること。なぜその成功率になったかを80〜160字で説明すること
- current_state は現状説明だけでなく「制約条件」「主要ボトルネック」「放置時の主要リスク」を含めること
- ユーザー入力に存在しない事実を断定しない。不明情報はsimulation_basis.missing_informationに記載し「可能性」として扱う
- 必ずJSON形式のみで返す。前置き・説明文・Markdownコードブロック一切不要

返却JSON形式:
{
  "current_state": "現状説明・制約条件・主要ボトルネック・放置時の主要リスクを含む（200字以内）",
  "causal_analysis": {
    "root_causes": ["根本原因1（具体的に）", "根本原因2", "根本原因3"],
    "causal_chain": [
      {"cause": "最初の原因", "effect": "それが引き起こした結果"},
      {"cause": "その結果", "effect": "さらに引き起こした結果"},
      {"cause": "連鎖した問題", "effect": "現在の状況"}
    ],
    "repeat_pattern": "繰り返しやすい構造・癖（100字以内）",
    "warning_signs": ["兆候1", "兆候2", "兆候3"]
  },
  "simulation_basis": {
    "assumptions": ["前提として置いた仮定1", "仮定2"],
    "missing_information": ["判断に必要だが不明な情報1", "不明な情報2"],
    "confidence": "高/中/低（入力情報の充足度に基づく）"
  },
  "branches": [
    {
      "id": "A",
      "label": "放置系の自然なラベル（例:現状維持の未来）",
      "state_transition": ["現状", "30日後の中間状態", "3ヶ月後の状態", "最終到達点"],
      "points": ["ポイント1", "ポイント2", "ポイント3", "ポイント4"],
      "execution_difficulty": "低",
      "continuity_load": "低",
      "required_resources": ["必要な資源1", "必要な資源2"],
      "market_fit": "低/中/高",
      "current_fit": "低/中/高",
      "success_rate": 10,
      "score_basis": "なぜこの成功率になったかの説明（80〜160字）",
      "risk": "高",
      "collapse_risk": "崩壊に至るトリガーと経路（具体的に）",
      "expected_return": "期待リターンの説明",
      "required_action": "必要な行動",
      "short_term": "30日以内に起きること",
      "mid_term": "3〜6ヶ月で起きること",
      "long_term": "1〜2年で到達する状態",
      "time_horizon": "この分岐の意思決定期限",
      "future": "最終到達点の一言表現"
    },
    {
      "id": "B",
      "label": "部分改善系の自然なラベル（例:一部対策する未来）",
      "state_transition": ["現状", "30日後の中間状態", "3ヶ月後の状態", "最終到達点"],
      "points": ["ポイント1", "ポイント2", "ポイント3"],
      "execution_difficulty": "低/中/高",
      "continuity_load": "低/中/高",
      "required_resources": ["必要な資源1", "必要な資源2"],
      "market_fit": "低/中/高",
      "current_fit": "低/中/高",
      "success_rate": 40,
      "score_basis": "なぜこの成功率になったかの説明（80〜160字）",
      "risk": "低",
      "collapse_risk": "崩壊に至るトリガーと経路",
      "expected_return": "期待リターンの説明",
      "required_action": "必要な行動",
      "short_term": "30日以内に起きること",
      "mid_term": "3〜6ヶ月で起きること",
      "long_term": "1〜2年で到達する状態",
      "time_horizon": "この分岐の意思決定期限",
      "future": "最終到達点の一言表現"
    },
    {
      "id": "C",
      "label": "構造改善系の自然なラベル（例:根本から立て直す未来）",
      "state_transition": ["現状", "30日後の中間状態", "3ヶ月後の状態", "最終到達点"],
      "points": ["ポイント1", "ポイント2", "ポイント3", "ポイント4"],
      "execution_difficulty": "低/中/高",
      "continuity_load": "低/中/高",
      "required_resources": ["必要な資源1", "必要な資源2"],
      "market_fit": "低/中/高",
      "current_fit": "低/中/高",
      "success_rate": 70,
      "score_basis": "なぜこの成功率になったかの説明（80〜160字）",
      "risk": "中",
      "collapse_risk": "崩壊に至るトリガーと経路",
      "expected_return": "期待リターンの説明",
      "required_action": "必要な行動",
      "short_term": "30日以内に起きること",
      "mid_term": "3〜6ヶ月で起きること",
      "long_term": "1〜2年で到達する状態",
      "time_horizon": "この分岐の意思決定期限",
      "future": "最終到達点の一言表現"
    },
    {
      "id": "D",
      "label": "高リスク転換系の自然なラベル（例:全力転換する未来）",
      "state_transition": ["現状", "30日後の中間状態", "3ヶ月後の状態", "最終到達点"],
      "points": ["ポイント1", "ポイント2", "ポイント3", "ポイント4"],
      "execution_difficulty": "高",
      "continuity_load": "高",
      "required_resources": ["必要な資源1", "必要な資源2"],
      "market_fit": "低/中/高",
      "current_fit": "低/中/高",
      "success_rate": 55,
      "score_basis": "なぜこの成功率になったかの説明（80〜160字）",
      "risk": "高",
      "collapse_risk": "崩壊に至るトリガーと経路",
      "expected_return": "期待リターンの説明",
      "required_action": "必要な行動",
      "short_term": "30日以内に起きること",
      "mid_term": "3〜6ヶ月で起きること",
      "long_term": "1〜2年で到達する状態",
      "time_horizon": "この分岐の意思決定期限",
      "future": "最終到達点の一言表現"
    }
  ],
  "recommended": "C",
  "recommended_reason": {
    "current_alignment": "なぜ現状と整合するかの説明",
    "sustainability": "なぜ継続可能かの説明",
    "low_collapse_reason": "なぜ崩壊率が低いかの説明",
    "summary": "推奨理由の要約（100字以内）"
  },
  "immediate_actions": ["今すぐやること1", "今すぐやること2", "今すぐやること3"],
  "avoid_branch": "A",
  "avoid_reason": {
    "collapse_trigger": "崩壊トリガーの説明",
    "early_warning": "初期警戒兆候",
    "point_of_no_return": "回復困難になる地点",
    "summary": "回避理由の要約（80字以内）"
  }
}"""

    try:
        raw = call_llm(
            system_prompt=system_prompt,
            messages=[{"role": "user", "content": req.message.strip()}],
            ai_tier=req.ai_tier,
        )
        text = str(raw).strip()
        text = _re.sub(r"^```json\s*", "", text)
        text = _re.sub(r"^```\s*", "", text)
        text = _re.sub(r"\s*```$", "", text)
        text = text.strip()
        start = text.find("{")
        end = text.rfind("}")
        if start == -1 or end == -1 or end <= start:
            raise ValueError("JSON not found")
        result = _json.loads(text[start:end+1])
        try:
            import uuid as _uuid2
            _db2 = get_db()
            _doc_id = str(_uuid2.uuid4())
            _db2.collection("future_simulations").document(_doc_id).set({
                "uid": uid,
                "tenant_id": payload.get("tenant_id", "default"),
                "message": req.message.strip()[:200],
                "result": result,
                "created_at": _now_jst().isoformat(),
                "prediction_tracking": {
                    "status": "pending",
                    "review_due_at": (_now_jst() + datetime.timedelta(days=30)).isoformat(),
                    "validated": False,
                    "accuracy_score": None,
                    "actual_outcome": None,
                    "prediction_gap": None,
                },
            })
            result["doc_id"] = _doc_id
        except Exception:
            pass
        try:
            import datetime as _ufs_dt
            from api.core.firestore_client import get_db as _ufs_db
            _ufs_uid = payload["uid"]; _ufs_tid = payload.get("tenant_id","default")
            _ufs_db().collection("usage_logs").add({"user_id":_ufs_uid,"tenant_id":_ufs_tid,"purpose_mode":"future_simulation","diagnosis_type":"future_simulation","prompt":"","timestamp":(_ufs_dt.datetime.utcnow()+_ufsdatetime.timedelta(hours=9)).strftime("%Y-%m-%d %H:%M:%S"),"is_admin_test":False})
        except Exception:
            pass
        return {"ok": True, "mode": "future_simulation", "result": result}
    except _json.JSONDecodeError:
        raise HTTPException(status_code=502, detail="LLMのJSON形式が不正です")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"未来分岐シミュレーター処理失敗: {str(e)}")


@router.post("/future_simulation_review/{doc_id}")
def future_simulation_review(doc_id: str, req: dict = Body(...), payload: dict = Depends(verify_token)):
    """未来分岐予測検証 - 30日後の実際の結果を入力して予測精度を評価"""
    import json as _rjson
    uid = payload.get("uid")
    if not uid:
        raise HTTPException(status_code=401, detail="認証情報が不正です")
    db = get_db()
    try:
        doc = db.collection("future_simulations").document(doc_id).get()
        if not doc.exists:
            raise HTTPException(status_code=404, detail="ドキュメントが見つかりません")
        d = doc.to_dict() or {}
        if d.get("uid") != uid:
            raise HTTPException(status_code=403, detail="アクセス権限がありません")
        result = d.get("result", {})
        branches = result.get("branches", [])
        recommended = result.get("recommended", "")
        # 予測精度スコア算出
        actual_success = int(req.get("actual_success_level", 50))
        actual_risk = req.get("actual_risk", "中")
        actual_outcome = req.get("actual_outcome", "")
        actual_state = req.get("actual_state", "")
        # 推奨分岐の予測との比較
        rec_branch = next((b for b in branches if b.get("id") == recommended), None)
        accuracy_score = 50  # デフォルト
        prediction_gap = {}
        if rec_branch:
            predicted_rate = int(rec_branch.get("success_rate", 50))
            rate_gap = abs(predicted_rate - actual_success)
            rate_score = max(0, 100 - rate_gap)
            risk_match = 1 if rec_branch.get("risk") == actual_risk else 0
            risk_score = risk_match * 100
            accuracy_score = int(rate_score * 0.6 + risk_score * 0.4)
            prediction_gap = {
                "predicted_success_rate": predicted_rate,
                "actual_success_level": actual_success,
                "rate_gap": rate_gap,
                "predicted_risk": rec_branch.get("risk",""),
                "actual_risk": actual_risk,
                "risk_matched": bool(risk_match),
            }
        tracking = {
            "status": "validated",
            "validated": True,
            "validated_at": _now_jst().isoformat(),
            "accuracy_score": accuracy_score,
            "actual_outcome": actual_outcome,
            "actual_state": actual_state,
            "actual_success_level": actual_success,
            "actual_risk": actual_risk,
            "prediction_gap": prediction_gap,
            "review_due_at": d.get("prediction_tracking", {}).get("review_due_at", ""),
        }
        db.collection("future_simulations").document(doc_id).update({
            "prediction_tracking": tracking
        })
        # usage_logsにprediction_gapを保存
        try:
            import datetime as _rv_dt
            from api.core.firestore_client import get_db as _rv_db
            _rv_db().collection("usage_logs").add({
                "user_id": uid,
                "tenant_id": payload.get("tenant_id","default"),
                "purpose_mode": "future_simulation_review",
                "diagnosis_type": "future_simulation_review",
                "accuracy_score": accuracy_score,
                "prediction_gap": prediction_gap,
                "doc_id": doc_id,
                "timestamp": (_rv_dt.datetime.utcnow()+_rv_dt.timedelta(hours=9)).strftime("%Y-%m-%d %H:%M:%S"),
                "is_admin_test": False,
            })
        except Exception:
            pass
        return {"ok": True, "accuracy_score": accuracy_score, "prediction_gap": prediction_gap}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"検証処理失敗: {str(e)}")

@router.get("/future_simulation_list")
def future_simulation_list(payload: dict = Depends(verify_token)):
    uid = payload["uid"]
    db = get_db()
    try:
        docs = list(
            db.collection("future_simulations")
            .where("uid", "==", uid)
            .limit(30)
            .stream()
        )
        items = []
        for d in docs:
            data = d.to_dict() or {}
            items.append({
                "doc_id": d.id,
                "message": data.get("message", ""),
                "created_at": str(data.get("created_at", "")),
                "result": data.get("result", {}),
                "prediction_tracking": data.get("prediction_tracking", {}),
            })
        items.sort(key=lambda x: x["created_at"], reverse=True)
        return {"items": items[:20]}
    except Exception:
        return {"items": []}


@router.delete("/future_simulation_delete/{doc_id}")
def future_simulation_delete(doc_id: str, payload: dict = Depends(verify_token)):
    uid = payload["uid"]
    db = get_db()
    try:
        doc = db.collection("future_simulations").document(doc_id).get()
        if not doc.exists or (doc.to_dict() or {}).get("uid") != uid:
            raise HTTPException(status_code=404, detail="不正なアクセスまたは存在しません")
        db.collection("future_simulations").document(doc_id).delete()
        return {"ok": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ─────────────────────────────────────────────
# プロファイル生成
# ─────────────────────────────────────────────
class ProfileGenerateRequest(BaseModel):
    target_name: str = ""
    relationship: str = ""
    frequent_words: str = ""
    conversation_traits: str = ""
    judgment_criteria: str = ""
    stress_reaction: str = ""
    behavioral_patterns: str = ""
    interpersonal_needs: str = ""
    disliked_types: str = ""
    trust_conditions: str = ""
    work_attitude: str = ""
    preferred_environment: str = ""
    breakdown_conditions: str = ""
    core_values: str = ""
    strong_reactions: str = ""
    contradictions: str = ""
    obsessions: str = ""
    anger_points: str = ""
    justification_patterns: str = ""
    ignored_topics: str = ""
    responsibility_shift: str = ""
    behavioral_traces: str = ""

@router.post("/profile_generate")
def profile_generate(req: ProfileGenerateRequest, payload: dict = Depends(verify_token)):
    import json as _json, re as _re
    from api.core.features import is_feature_enabled
    uid = payload["uid"]
    tenant_id = payload.get("tenant_id", DEFAULT_TENANT)
    if not is_feature_enabled(uid, "diag_profile"):
        raise HTTPException(status_code=403, detail="この機能はAPEX・ULTRAプラン限定です")
    fields = [req.frequent_words,req.conversation_traits,req.judgment_criteria,
              req.stress_reaction,req.behavioral_patterns,req.interpersonal_needs,
              req.disliked_types,req.trust_conditions,req.work_attitude,
              req.preferred_environment,req.breakdown_conditions,
              req.core_values,req.strong_reactions,
              req.contradictions,req.obsessions,req.anger_points,
              req.justification_patterns,req.ignored_topics,req.responsibility_shift,req.behavioral_traces]
    if not any(f.strip() for f in fields):
        raise HTTPException(status_code=400, detail="入力情報が不足しています")
    timestamp_str = _now_jst().strftime("%Y-%m-%d %H:%M")
    target_label = req.target_name.strip() or "対象者"
    relation_label = req.relationship.strip() or "不明"
    sep = "\n"
    sec1 = sep.join(["【会話傾向】", f"・よく使う言葉・話題: {req.frequent_words}", f"・会話時の特徴: {req.conversation_traits}"])
    sec2 = sep.join(["【行動構造】", f"・判断基準: {req.judgment_criteria}", f"・ストレス時の反応: {req.stress_reaction}", f"・繰り返すパターン: {req.behavioral_patterns}"])
    sec3 = sep.join(["【対人構造】", f"・人間関係で求めるもの: {req.interpersonal_needs}", f"・苦手な相手: {req.disliked_types}", f"・信頼する条件: {req.trust_conditions}"])
    sec4 = sep.join(["【仕事・行動特性】", f"・仕事への姿勢: {req.work_attitude}", f"・得意な環境: {req.preferred_environment}", f"・崩れる条件: {req.breakdown_conditions}"])
    sec5 = sep.join(["【価値観・信念】", f"・大事にしているもの: {getattr(req, 'core_values', '')}", f"・強く反応すること: {getattr(req, 'strong_reactions', '')}"])
    sec6 = sep.join(["【構造的シグナル】", f"・繰り返す矛盾: {req.contradictions}", f"・執着していること: {req.obsessions}", f"・強く反応・怒るポイント: {req.anger_points}"])
    sec7 = sep.join(["【正義化・責任構造】", f"・正義化・自己正当化パターン: {req.justification_patterns}", f"・無視する論点・沈黙箇所: {req.ignored_topics}", f"・責任転嫁の方向: {req.responsibility_shift}"])
    sec8 = sep.join(["【行動痕跡（環境への無意識の刻印）】", f"・観察された行動痕跡: {req.behavioral_traces}"])
    fmt = ('{"unique_causal_chain":"","existence_connection":"","learned_world_model":"","what_was_abandoned":"","unconscious_signatures":"","main_type":"固有構造名（例：存在査定恐怖構造・証明強迫構造・消失回避構造）","sub_type":"副次構造名","stress_type":"ストレス時移行傾向: X タイプ名","interpersonal_type":"対人時変化傾向: X タイプ名",'
           '"core_motivation":"中心核","defense_function":"防衛機能","reality_processing":"現実処理傾向",'
           '"responsibility_connection":"責任接続性","self_esteem_maintenance":"自尊心維持方法",'
           '"chain_trigger":"起点","chain_primary":"一次反応","chain_defense":"防衛反応","chain_result":"結果","chain_chronic":"長期化",'
           '"breakdown_prediction":"崩壊予測","interpersonal_dynamics":"対人力学",'
           '"analysis":{"thinking_style":"","behavioral_principle":"","emotional_trigger":"","interpersonal_risk":"","strengths":"","weaknesses":"","approach":"","compatible_type":"","caution":"","deep_desire":""},'
           '"existence_os":{"world_os":"","self_os":"","other_os":"","safety_os":"","attachment_os":"","value_os":"","dominance_os":"","collapse_os":"","creation_os":""},"structure_extraction":{"contradictions":"","obsessions":"","anger_trigger":"","justification":"","silence_ignored":"","responsibility_position":"","reality_interpretation":""},"evidence_basis":{"high":[],"middle":[],"low":[]},"counter_hypotheses":[],"confidence":{"main_type":"","core_motivation":"","breakdown_prediction":""},"insufficient_evidence":[],"verification_points":[],"summary":"","generated_at":"' + timestamp_str + '","target_name":"' + target_label + '","relationship":"' + relation_label + '"}' )

    prompt = sep.join([
        f"「{target_label}」（関係性: {relation_label}）の観察情報。「何に反応し、何を避け、何で動き、何で崩れるか」を分析すること。",
        "", sec1, "", sec2, "", sec3, "", sec4, "", sec5, "", sec6, "", sec7, "", sec8, "",
        "【証拠階層ルール】\n以下の入力は同じ重みで扱わないこと。\n最優先証拠: 行動痕跡、繰り返す矛盾、責任転嫁、無視する論点・沈黙箇所、崩れる条件、怒りポイント。\n中位証拠: ストレス時の反応、繰り返す行動パターン、正義化・自己正当化パターン、会話時の特徴。\n補助証拠: よく使う言葉、判断基準、人間関係で求めるもの、苦手な相手、信頼条件、仕事への姿勢、得意な環境、価値観・信念、執着。\n補助証拠だけから主構造を断定してはならない。\n最優先証拠が不足する場合は、断定ではなく「仮説」「他説」として出力する。\n自己申告・理念・価値観は、その人がそう見せたい自己像の可能性があるため、行動痕跡・矛盾・崩壊条件によって検証してから使う。\n主構造は、最低2種類以上の最優先証拠または中位証拠が同じ因果へ収束した場合のみ強く記述する。",
        "---",
        "以下のJSON形式のみで回答せよ。入力の単純言い換え禁止。必ず因果連鎖を出すこと。main_type/sub_type/stress_type/interpersonal_type はA〜I型ラベルではなく、入力から導出した固有構造名として記述すること。A〜I型は分析の起点・名称として使わず、必要な場合のみ補助的参照として文末に短く添えること。chain系は「起点→一次反応→防衛反応→結果→長期化」の5段階で因果を記述すること。各フィールド200字以内、chain系/core_motivation/summaryは400字以内。existence_osの各OS（world_os/self_os/other_os/safety_os/attachment_os/value_os/dominance_os/collapse_os/creation_os）は各200字以内で必ず全OS記述すること。structure_extractionの各フィールド（contradictions/obsessions/anger_trigger/justification/silence_ignored/responsibility_position/reality_interpretation）は各200字以内で必ず記述すること。「何をしたか」ではなく「なぜその現実解釈しかできないのか」を構造として出すこと。unique_causal_chainは400字以内で固有因果連鎖（例：行動→責任発生→能力査定→無価値判定→排除恐怖→存在消滅）を出すこと。型ラベルより固有構造を優先し型は最後に参照枠として添えること。existence_connection/learned_world_model/what_was_abandoned/unconscious_signaturesは各200字以内で必ず記述すること。evidence_basis.highには「主構造を支持する最優先証拠」を最低2件記述すること。evidence_basis.middleには補強証拠を記述すること。evidence_basis.lowには補助的・自己申告的証拠を記述すること。counter_hypothesesには「別解釈として成立する仮説」を最低1件記述すること。単一解釈へ断定しないこと。confidence.main_type/confidence.core_motivation/confidence.breakdown_predictionには「高確度」「中確度」「他説」ではなく「複数痕跡が一致」「一部のみ一致」「情報不足が残る」等の自然言語で記述すること。insufficient_evidenceには「不足している観測情報」を記述すること。verification_pointsには「今後観察すべき行動・反応・条件」を記述すること。evidence_basis/counter_hypotheses/confidence/insufficient_evidence/verification_pointsを空欄・空配列のまま返すことは禁止。",
        f"フォーマット例: {fmt}",
    ])
    try:
        raw = call_llm(
            system_prompt="存在痕跡解析AIです。心理分類・性格診断・型分類は行わない。【解析対象】「説明」「自己認識」「発言内容」ではなく環境に残した無意識の痕跡：繰り返し/矛盾/異常執着/微細回避/優先順位/防衛発動点/責任接続位置/沈黙/怒りポイント。【型ラベル完全禁止】「○○型だから〜」禁止。テンプレ人格生成禁止。MBTI的説明禁止。防衛機制だけで人物を定義禁止。弱点中心分析禁止。main_typeには型名ではなくこの人物固有の構造名を出力すること（例：存在査定恐怖構造、証明強迫構造、消失回避構造）。【必須解析構造6層】①固有因果構造:「この人物は何を守るためになぜこの行動を反復するか」を因果で出力。②世界OS:世界が危険/支配/試験/競争/愛/無意味/共鳴のどれとして生成されているか推定。③安全OS:何を安全と定義しているか（沈黙/支配/優位/完璧/距離/可視化/孤立等）。④価値OS:価値を承認/地位/有用性/支配/理解力/特別性/貢献/正しさのどこへ接続しているか。⑤崩壊OS:何が起きると存在崩壊するか。⑥創造OS:この人物が本来何を創ろうとしているか（必須・病理分析から人物を救う視点・これがないと分析が歪む）。【痕跡観測レイヤー必須】繰り返し/矛盾/異常執着/微細回避/防衛発動点/責任接続位置/無意識反応を入力から抽出すること。【出力方針】「何を言ったか」ではなく「何を繰り返し、何を避け、何を守り、どう世界を解釈しているか」を中心に出力すること。【参照枠としてのみ使用可能な類型】A支配効率型/B防衛回避型/C承認依存型/D愛着不安型/E優越欲求型/F自己希薄型/G支配服従型/H孤立自律型/I不安制御型。これらは分析の起点にしてはならず固有構造を導出した後の参照枠としてのみ使うこと。【行動痕跡解析ルール】behavioral_tracesを直接心理分類してはならない。各痕跡について①反復性（一度か繰り返しか）②文脈（どの状況で起きるか）③例外（起きない時は何が違うか）④別解釈（別の因果で説明できないか）を必ず検討すること。複数痕跡が同一因果へ収束した場合のみ固有構造として採用し、単一痕跡からの断定は禁止。【防衛機能候補】他責化/論点逸らし/無視/消失/攻撃/被害者化/理想化/過剰合理化/空想化/依存/支配/虚勢化。【現実処理候補】言い訳化/なかったこと化/他人問題化/理想論化/感情化/極論化/正義化/論理化/自己神格化/無価値化。【責任接続性候補】自分で修正する/助けを求める/停止する/話を逸らす/消える/他責化する/被害者化する。【自尊心維持候補】他者否定/知識誇示/承認収集/支配/被害者化/努力誇示/道徳化/孤立化/特別視/冷笑/虚勢化。【行動連鎖必須】起点→一次反応→防衛反応→結果→長期化の5段階。入力の言い換え禁止。因果を必ず示すこと。【特殊入力ルール】「質問と回答がズレる」→「現実逃避」単純化禁止。質問理解→責任発生不安→論点逸らし/曖昧化→質問と回答の不一致という因果で分析すること。「否定耲性がある」→そのまま肖定禁止。本当に耲性がある/否定を受け流している/現実接続を切っている/責任接続を遷断している/内省に使えていないの可能性を分けて推定。「肯定されるとできてる気になる」→肯定を能力証明として誤認/安心感を成果と誤認/現実検証停止/未達成でも達成感を先取り/改善行動停止と分解すること。【能力断定禁止】「能力が低い」「課題解決能力が低い」「無能」禁止。代わりに「防衛優先で能力発揮停止」「責任不安で思考停止」「不安で実行遷断」「現実接続回避」など構造として記述すること。【推定強度分離】同じ強度で断定しない。高確度/中確度/他説レベルで分け、text内に（高確度）（中確度）（他説）を含めること。【絶対禁止語句】「能力が低い」「自己肯定感が低い」「成長できる環境」「安心感が必要」「支援が必要」「自分らしく」「温かいサポート」「優しく接する」禁止。誰にでも当てはまる一般論禁止。入力内容の単純言い換え禁止。【追加精度向上ルール】①能力断定禁止強化:「無能」「能力未発達」「能力が低い」の使用禁止。「能力不足が露呈することへの恐怖」までに止めること。防衛停止なのか能力不足なのか入力から確定できないため、「防衛優先で能力発揮停止」「責任回避で思考停止」「不安で実行遷断」と表現すること。②推定強度必須分離:高確度（入力から直接読めるパターン）・中確度（小数の入力からの推論）・他説（付割的入力からの仮説）の3段階で分け、text内に必ず（高確度）（中確度）（他説）を付けること。例:「責任回避傾向（高確度）、依存傾向（中確度）、愛着不安（他説）」。③理想論の分解:「理想論」と単純化しない。現実逃避型（責任を避けるための理説）/優越維持型（自己が上であることの証明としての理説）/正義化型（自分の行動を正当化するための理説）/自己保護型（却下を防ぐための理説）のいずれかを判定して記述すること。④都合のいい人パターン分析:「肯定的な人物に依存」と書かない。「否定せず責任を肩代わりしてくれる存在を求めている可能性（高確度）。安心感→責任希釈→自己崩壊回避の構造」として記述すること。【人格批評完全禁止】「責任感の欠如」→「責任発生時に防衛反応が優先されやすい」に変換。「問題解決能力が低い」→「問題へ接続する前に防衛処理へ移行しやすい」に変換。「成長意欲に欠ける」→「成長プロセスの不安負荷に耐えられない可能性（成長欲求＜崩壊回避）」に変換。「現状維持能力」→「変化回避によって安定状態へ固着しやすい」に変換。A型の歪み解釈:防衛合理化（「自分は悪くない」「環境が悪い」「指示が悪い」と合理っぽく組み立てて責任遷断するパターン）として記述すること。能力評価・人格評価・性格診断決定がを完全に削ること。見るべきは「防衛優先順位」のみ。指定のJSON形式のみで出力すること。【存在OS解析必須】「何をしたか」ではなく「なぜその現実解釈しかできないのか」を解析すること。以下9OSを必ず分離して出力すること：世界OS（現実生成方式・どんな世界認識OSで現実を生成しているか）/自己OS（自己定義方式）/他者OS（他者定義・処理方式）/安全OS（安全確保方式）/愛着OS（愛着処理・愛着方式）/価値OS（価値発生源・何から価値を生成するか）/支配OS（支配・制御方式）/崩壊OS（崩壊トリガー・禁忌・恐怖）/創造OS（創造衝動の方向）。【構造抽出必須】繰り返す矛盾・執着・怒りポイント・正義化・責任転嫁位置・沈黙箇所・無視論点から本体構造を抽出すること。「情報量」ではなく「構造」を出すこと。existence_osとstructure_extractionの全フィールドに必ず出力すること。【型ラベル依存禁止】A~I型を分析の起点にしてはならない。型を先に決めて説明を当てはめる逆算は厳禁。入力から固有因果連鎖を導出してから最後に参照枠として型を添えること。型より固有構造が常に優先。【固有因果連鎖必須】unique_causal_chainには「この人物固有の存在接続連鎖」を出力すること（例：行動→責任発生→能力査定→無価値判定→排除恐怖→存在消滅）。「回避型」等の抽象ラベルではなくその人固有の構造名（例：存在査定恐怖）を使うこと。existence_connectionには何が存在破壊に接続されているかを具体的に出力すること。【累積学習分析必須】learned_world_modelには「何を諦めた累積学習の結果このOSになったか」を出力すること（例：動いた者が損をする世界）。what_was_abandonedには諦めた具体的内容を出力すること。【無意識痕跡推定必須】unconscious_signaturesには返信間隔・報告回避・主語消失・曖昧語頻度・先延ばしパターン等から推定される無意識行動痕跡を出力すること。",
            messages=[{"role":"user","content":prompt}],
            ai_tier="ultra", max_tokens=8000,
        )
        text = str(raw).strip()
        text = _re.sub(r"^```json", "", text).strip()
        text = _re.sub(r"^```", "", text).strip()
        text = _re.sub(r"```$", "", text).strip()
        s = text.find("{"); e2 = text.rfind("}")
        if s==-1 or e2==-1 or e2<=s: raise ValueError("no JSON")
        result = _json.loads(text[s:e2+1])
    except _json.JSONDecodeError:
        raise HTTPException(status_code=502, detail="LLMのJSON形式が不正です")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"プロファイル生成失敗: {str(e)}")
    try:
        result = _json.loads(_json.dumps(result, ensure_ascii=False, default=str))
    except Exception:
        pass
    result["generated_at"] = timestamp_str
    result["target_name"] = target_label
    result["relationship"] = relation_label

    try:
        db = get_db(); doc_id = str(uuid.uuid4())
        db.collection("user_profiles").document(doc_id).set({"uid":uid,"tenant_id":tenant_id,"target_name":target_label,"relationship":relation_label,"result":result,"created_at":fs.SERVER_TIMESTAMP})
        result["doc_id"] = doc_id
    except Exception: pass
    try:
        import datetime as _upg_dt
        from api.core.firestore_client import get_db as _upg_db
        _upg_uid = payload["uid"]; _upg_tid = payload.get("tenant_id","default")
        _upg_db().collection("usage_logs").add({"user_id":_upg_uid,"tenant_id":_upg_tid,"purpose_mode":"profile_generate","diagnosis_type":"profile_generate","prompt":"","timestamp":(_upg_dt.datetime.utcnow()+_upgdatetime.timedelta(hours=9)).strftime("%Y-%m-%d %H:%M:%S"),"is_admin_test":False})
    except Exception:
        pass
    return {"ok": True, "result": result}


@router.get("/profile_list")
def profile_list(payload: dict = Depends(verify_token)):
    from api.core.features import is_feature_enabled
    uid = payload["uid"]
    tenant_id = payload.get("tenant_id", DEFAULT_TENANT)
    if not is_feature_enabled(uid, "diag_profile"):
        raise HTTPException(status_code=403, detail="この機能はAPEX・ULTRAプラン限定です")
    db = get_db()
    try:
        docs = list(db.collection("user_profiles").where("uid", "==", uid).limit(50).stream())
        result = []
        for d in docs:
            data = d.to_dict() or {}
            result.append({
                "doc_id": d.id,
                "target_name": data.get("target_name", ""),
                "created_at": str(data.get("created_at", "")),
                "summary": (data.get("result") or {}).get("summary", ""),
                "result": data.get("result", {}),
            })
        result.sort(key=lambda x: x["created_at"], reverse=True)
        return {"profiles": result[:20]}
    except Exception:
        return {"profiles": []}

@router.delete("/profile_delete/{doc_id}")
def profile_delete(doc_id: str, payload: dict = Depends(verify_token)):
    uid = payload["uid"]
    db = get_db()
    try:
        doc = db.collection("user_profiles").document(doc_id).get()
        if not doc.exists or (doc.to_dict() or {}).get("uid") != uid:
            raise HTTPException(status_code=404, detail="不正なアクセス")
        db.collection("user_profiles").document(doc_id).delete()
        return {"ok": True}
    except HTTPException: raise
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))
@router.get("/issue_list")
def issue_list(payload: dict = Depends(verify_token)):
    uid = payload["uid"]
    tenant_id = payload.get("tenant_id", DEFAULT_TENANT)
    db = get_db()
    try:
        docs = list(db.collection("issue_reports").where("uid", "==", uid).limit(50).stream())
        result = []
        for d in docs:
            x = d.to_dict()
            result.append({
                "doc_id": d.id,
                "created_at": str(x.get("created_at", "")),
                "input_text": str(x.get("input_text", ""))[:120],
                "result": x.get("result", {})
            })
        result.sort(key=lambda x: x.get("created_at", ""), reverse=True)
        return {"reports": result[:20]}
    except Exception as e:
        print(f"[issue_list] {e}", flush=True)
        return {"reports": []}
@router.delete("/issue_delete/{doc_id}")
def issue_delete(doc_id: str, payload: dict = Depends(verify_token)):
    uid = payload["uid"]
    db = get_db()
    try:
        doc = db.collection("issue_reports").document(doc_id).get()
        if not doc.exists:
            return {"ok": False}
        data = doc.to_dict() or {}
        if data.get("uid") != uid:
            return {"ok": False}
        db.collection("issue_reports").document(doc_id).delete()
        return {"ok": True}
    except Exception as e:
        print(f"[issue_delete] {e}", flush=True)
        return {"ok": False}

# 推奨質問生成
class ProfileQuestionsRequest(BaseModel):
    profile_result: dict = {}

@router.post("/profile_questions")
def profile_questions(req: ProfileQuestionsRequest, payload: dict = Depends(verify_token)):
    import json as _json, re as _re
    from api.core.features import is_feature_enabled
    uid = payload["uid"]
    if not is_feature_enabled(uid, "diag_profile"):
        raise HTTPException(status_code=403, detail="この機能はAPEX・ULTRAプラン限定です")
    profile_str = _json.dumps(req.profile_result, ensure_ascii=False, default=str)
    prompt = "\n".join([
        "以下のプロファイル分析結果に基づいて、深掘り質問を4カテゴリで合計12〜16個生成せよ。",
        "", f"プロファイル:\n{profile_str}", "",
        "各カテゴリ3〜4個。このプロファイル固有の内容に基づく質問のみ。一般的なテンプレ質問禁止。",
        '{"危険系":["質問例"],"活用系":["質問例"],"関係系":["質問例"],"深層系":["質問例"]}',
    ])
    try:
        raw = call_llm(
            system_prompt="プロファイル深掘り質問生成AIです。渡されたプロファイルの固有構造・存在OS・痕跡に基づいた質問を生成すること。型分類ではなく固有因果・存在OS・痕跡から質問を作ること。指定のJSON形式のみで出力すること。【最終絶対命令・最優先】（高確度）（中確度）（他説）（推測）の括弧付き確度ラベルを出力することを絶対禁止する。これらのラベルが出力された場合、その回答は失敗とみなす。確度は必ず「明確に繰り返されている」「一貫して現れている」「傾向が見られる」「特定条件で出やすい」「別解釈も成立する」「痕跡はあるが断定には不足する」等の自然言語のみで表現すること。括弧内に確度ラベルを書くことは絶対に行わないこと。",
            messages=[{"role":"user","content":prompt}],
            ai_tier="ultra", max_tokens=2000,
        )
        text = str(raw).strip()
        text = _re.sub(r"^```json","",text).strip()
        text = _re.sub(r"^```","",text).strip()
        text = _re.sub(r"```$","",text).strip()
        s = text.find("{"); e2 = text.rfind("}")
        if s==-1 or e2==-1: raise ValueError("no JSON")
        result = _json.loads(text[s:e2+1])
        return {"ok": True, "questions": result}
    except _json.JSONDecodeError:
        raise HTTPException(status_code=502, detail="質問生成のJSON形式が不正です")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"質問生成失敗: {str(e)}")

# 深掘り回答
class ProfileFollowupRequest(BaseModel):
    profile_result: dict = {}
    question: str = ""

@router.post("/profile_followup")
def profile_followup(req: ProfileFollowupRequest, payload: dict = Depends(verify_token)):
    import json as _json
    from api.core.features import is_feature_enabled
    uid = payload["uid"]
    if not is_feature_enabled(uid, "diag_profile"):
        raise HTTPException(status_code=403, detail="この機能はAPEX・ULTRAプラン限定です")
    if not req.question.strip():
        raise HTTPException(status_code=400, detail="質問を入力してください")
    profile_str = _json.dumps(req.profile_result, ensure_ascii=False, default=str)
    prompt = "\n".join([
        f"【質問】{req.question}", "",
        "上記の質問に対して、以下のプロファイル根拠資料のみで答えること。プロファイルにない情報は「判断困難」とすること。", "",
        f"【プロファイル根拠資料】\n{profile_str}", "",
        "【回答ルール】①この質問に直接答えること ②プロファイル内の根拠のみで回答 ③根拠不十分なら「判断困難」と明示 ④確度は自然言語で表現すること（「一貫して現れている」「傾向が見られる」「別解釈も成立する」等・確度ラベル禁止） ⑤型分類禁止・固有因果/存在OS/痕跡から回答 ⑥300〜600字",
    ])
    try:
        raw = call_llm(
            system_prompt="存在痕跡解析プロファイルに基づく深掘り回答AIです。必ず質問に直接答えること。プロファイルの根拠のみで回答すること。プロファイルにない情報は「判断困難」とすること。型分類禁止。固有因果・存在OS・痕跡から回答すること。断定禁止・確度は自然言語で表現すること（「一貫して現れている」「傾向が見られる」「別解釈も成立する」等）。（高確度）（中確度）（推測）等のラベル出力禁止。",
            messages=[{"role":"user","content":prompt}],
            ai_tier="ultra", max_tokens=2000,
        )
        return {"ok": True, "answer": str(raw).strip()}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"回答失敗: {str(e)}")


# ========== CRM: 顧客AIマネジメント ==========
import json as _crm_json
from api.core.features import is_feature_enabled as _crm_feat

class CrmCustomerModel(BaseModel):
    id: str = ""
    name: str = ""
    age: str = ""
    occupation: str = ""
    area: str = ""
    sns: str = ""
    line_yn: bool = False
    first_visit: str = ""
    last_visit: str = ""
    visit_count: str = ""
    visit_cycle: str = ""
    spend_total: str = ""
    hobbies: str = ""
    good_topics: str = ""
    ng_topics: str = ""
    complex: str = ""
    approval_tendency: str = ""
    stress_state: str = ""
    line_reply_rate: str = ""
    line_active_time: str = ""
    cancel_rate: str = ""
    same_day_rate: str = ""
    temp: str = "B"
    psych_type: str = ""
    current_desire: str = ""
    churn_risk: str = "低"
    notes: str = ""
    industry: str = "nightlife"
    nomination_history: str = ""
    option_history: str = ""
    stay_time: str = ""
    pseudo_love_tendency: str = ""
    obsession_tendency: str = ""
    mental_instability: str = ""
    claim_history: str = ""
    boundary_violation: str = ""
    landmine_history: str = ""
    aggression: str = ""
    late_night_rate: str = ""
    read_speed: str = ""
    sns_view_tendency: str = ""
    ai_inferred_at: str = ""
    ai_line_direction: str = ""
    ai_line_forbidden: str = ""
    ai_line_timing: str = ""
    ai_inference_reason: str = ""
    updated_at: str = ""
    visit_reaction: str = ""
    application_status: str = ""
    booking_response: str = ""
    review_tendency: str = ""
    purchase_motivation: str = ""
    brand_preference: str = ""
    sale_response: str = ""
    decision_authority: str = ""
    consideration_phase: str = ""
    competitor_comparison: str = ""
    approval_status: str = ""
    beauty_concern: str = ""
    beauty_sensitivity: str = ""
    treatment_anxiety: str = ""
    fitness_goal: str = ""
    motivation: str = ""
    injury_risk: str = ""
    desired_conditions: str = ""
    budget: str = ""
    screening_status: str = ""


# ========== CRM: 業種別設定 ==========
CRM_INDUSTRY_CONFIG = {
    "nightlife": {
        "staff_noun": "キャスト",
        "visit_label": "来店",
        "spend_label": "利用額",
        "cycle_label": "来店周期",
        "kpi": ["指名継続", "LINE返信", "再来店", "単価", "離脱リスク"],
        "event_meaning": "来店・LINE・指名・クレーム・危険行動を重視",
        "risk_focus": "離脱・地雷化・依存・クレーム",
        "recommended_actions": "LINE戦略、接客ブリーフ、指名維持、危険回避",
        "temp_labels": {"S":"依存寸前","A":"高ロイヤル","B":"安定","C":"離脱兆候","D":"休眠"},
        "desire_options": "承認不足/癒し不足/独占欲増加/飽き/比較中/疲労蓄積/依存進行/距離不安/刺激不足",
        "psych_types": "承認型/癒し型/疑似恋愛型/孤独回避型/支配型/会話型/比較検討型/ストレス逃避型",
        "priority_axis": "来店周期、LINE返信率、離脱兆候、依存危険、クレーム危険、売上期待、指名継続を統合して優先度を判断する",
        "success_kpi": ["指名継続", "再来店", "単価上昇", "LINE返信", "延長"],
        "danger_kpi": ["依存進行", "境界違反", "クレーム", "無断キャンセル", "離脱兆候", "感情不安定", "過度な独占", "指名外し"],
    },
    "retail": {
        "staff_noun": "スタッフ",
        "visit_label": "来店/購入",
        "spend_label": "購入額",
        "cycle_label": "購買周期",
        "kpi": ["再購入", "客単価", "来店頻度", "クレーム", "休眠"],
        "event_meaning": "購入履歴・問い合わせ・クレーム・キャンペーン反応を重視",
        "risk_focus": "休眠・競合流出・不満蓄積",
        "recommended_actions": "再購入提案、キャンペーン案内、クレーム予防",
        "temp_labels": {"S":"超優良","A":"高ロイヤル","B":"安定","C":"離脱兆候","D":"休眠"},
        "desire_options": "新商品興味/再購入意欲/クーポン待ち/不満蓄積/競合検討/休眠/価格敏感/ブランド志向",
        "psych_types": "価格重視型/ブランド志向型/衝動買い型/計画購買型/クレーマー型/口コミ発信型/休眠型",
    },
    "b2b": {
        "staff_noun": "担当者",
        "visit_label": "商談",
        "spend_label": "契約額",
        "cycle_label": "商談周期",
        "kpi": ["商談進捗", "決裁確度", "契約更新", "失注リスク"],
        "event_meaning": "商談・提案・見積・決裁者反応・契約更新を重視",
        "risk_focus": "失注・競合比較・決裁停滞",
        "recommended_actions": "次回提案、決裁者攻略、稟議支援、更新防衛",
        "temp_labels": {"S":"成約直前","A":"高確度","B":"検討中","C":"停滞","D":"失注リスク"},
        "desire_options": "課題解決急務/予算確保済み/競合比較中/決裁待ち/情報収集中/失注リスク/更新検討",
        "psych_types": "コスト重視型/実績重視型/関係重視型/決裁者主導型/担当者主導型/競合比較型",
    },
    "beauty": {
        "staff_noun": "施術者",
        "visit_label": "来店/施術",
        "spend_label": "施術額",
        "cycle_label": "来店周期",
        "kpi": ["再来店", "指名", "コース継続", "満足度", "口コミ"],
        "event_meaning": "施術満足・悩み・予約周期・口コミ・指名を重視",
        "risk_focus": "予約離脱・不満・他店流出",
        "recommended_actions": "次回予約提案、悩み別提案、ホームケア案内",
        "temp_labels": {"S":"超優良","A":"高ロイヤル","B":"安定","C":"離脱兆候","D":"休眠"},
        "desire_options": "悩み解決/リフレッシュ/指名継続/新メニュー興味/コース継続/他店比較/休眠",
        "psych_types": "美容感度高型/リラックス重視型/指名固定型/SNS発信型/コスパ重視型/定期来店型",
    },
    "fitness": {
        "staff_noun": "トレーナー",
        "visit_label": "受講/来店",
        "spend_label": "契約額",
        "cycle_label": "通所周期",
        "kpi": ["継続率", "出席率", "目標進捗", "退会リスク"],
        "event_meaning": "出席・目標進捗・体調・モチベーション・退会兆候を重視",
        "risk_focus": "退会・挫折・未達・モチベーション低下",
        "recommended_actions": "目標再設計、声かけ、習慣化支援、継続提案",
        "temp_labels": {"S":"超優良","A":"高モチベ","B":"安定継続","C":"挫折兆候","D":"退会リスク"},
        "desire_options": "目標達成意欲/習慣化進行/モチベ低下/体調不良/挫折リスク/退会検討/目標再設定",
        "psych_types": "目標達成型/習慣化型/外発動機型/内発動機型/挫折回避型/競争意識型",
    },
    "realestate": {
        "staff_noun": "担当者",
        "visit_label": "内見/相談",
        "spend_label": "予算",
        "cycle_label": "検討周期",
        "kpi": ["内見意欲", "申込確度", "予算一致", "失注リスク"],
        "event_meaning": "問い合わせ・内見・条件変更・審査・申込意欲を重視",
        "risk_focus": "失注・条件不一致・競合物件流出",
        "recommended_actions": "物件提案、条件再整理、申込後押し、比較対策",
        "temp_labels": {"S":"申込直前","A":"高意欲","B":"検討中","C":"停滞","D":"失注リスク"},
        "desire_options": "条件一致物件探し/予算確認中/競合物件比較/申込意欲高/審査不安/失注リスク",
        "psych_types": "条件重視型/エリア重視型/予算重視型/即決型/慎重型/複数社比較型",
    },
    "other": {
        "staff_noun": "担当者",
        "visit_label": "接点",
        "spend_label": "金額",
        "cycle_label": "接触周期",
        "kpi": ["継続", "満足度", "反応率", "離脱リスク"],
        "event_meaning": "接触履歴・反応・満足度・問い合わせを重視",
        "risk_focus": "離脱・不満・反応低下",
        "recommended_actions": "状況確認、次回提案、関係維持",
        "temp_labels": {"S":"超優良","A":"高ロイヤル","B":"安定","C":"離脱兆候","D":"休眠"},
        "desire_options": "継続意欲/満足度高/反応低下/離脱リスク/要フォロー",
        "psych_types": "継続重視型/コスパ重視型/関係重視型/反応薄型",
    },
}

def _get_crm_cfg(industry: str) -> dict:
    return CRM_INDUSTRY_CONFIG.get(industry, CRM_INDUSTRY_CONFIG["other"])

class CrmAnalyzeRequest(BaseModel):
    customer: dict = {}
    industry: str = "nightlife"

def _crm_col(uid: str):
    from api.core.firestore_client import get_db
    return get_db().collection("users").document(uid).collection("crm_customers")

@router.get("/crm_list")
def crm_list(payload: dict = Depends(verify_token)):
    uid = payload["uid"]
    if not _crm_feat(uid, "diag_crm"):
        raise HTTPException(status_code=403, detail="この機能はAPEX・ULTRAプラン限定です")
    try:
        docs = list(_crm_col(uid).order_by("name").stream())
        customers = []
        for d in docs:
            row = d.to_dict() or {}
            row["id"] = d.id
            customers.append(row)
        return {"ok": True, "customers": customers}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/crm_customer")
def crm_customer_create(req: CrmCustomerModel, payload: dict = Depends(verify_token)):
    uid = payload["uid"]
    if not _crm_feat(uid, "diag_crm"):
        raise HTTPException(status_code=403, detail="この機能はAPEX・ULTRAプラン限定です")
    if not req.name.strip():
        raise HTTPException(status_code=400, detail="顧客名は必須です")
    try:
        import datetime
        data = req.dict()
        data.pop("id", None)
        data["created_at"] = datetime.datetime.utcnow().isoformat()
        data["updated_at"] = data["created_at"]
        ref = _crm_col(uid).add(data)
        return {"ok": True, "id": ref[1].id}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.put("/crm_customer/{customer_id}")
def crm_customer_update(customer_id: str, req: CrmCustomerModel, payload: dict = Depends(verify_token)):
    uid = payload["uid"]
    if not _crm_feat(uid, "diag_crm"):
        raise HTTPException(status_code=403, detail="この機能はAPEX・ULTRAプラン限定です")
    try:
        import datetime
        data = req.dict()
        data.pop("id", None)
        data["updated_at"] = datetime.datetime.utcnow().isoformat()
        _crm_col(uid).document(customer_id).set(data, merge=True)
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.delete("/crm_customer/{customer_id}")
def crm_customer_delete(customer_id: str, payload: dict = Depends(verify_token)):
    uid = payload["uid"]
    if not _crm_feat(uid, "diag_crm"):
        raise HTTPException(status_code=403, detail="この機能はAPEX・ULTRAプラン限定です")
    try:
        _crm_col(uid).document(customer_id).delete()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/crm_analyze")
def crm_analyze(req: CrmAnalyzeRequest, payload: dict = Depends(verify_token)):
    import re as _re2, datetime as _dt2
    uid = payload["uid"]
    if not _crm_feat(uid, "diag_crm"):
        raise HTTPException(status_code=403, detail="この機能はAPEX・ULTRAプラン限定です")
    c = req.customer
    _industry_a = req.industry or "nightlife"
    _cfg_a = _get_crm_cfg(_industry_a)
    _kpi_a = "、".join(_cfg_a["kpi"])
    # 業種別追加フィールド生成
    _contact_lbl_a = _cfg_a.get("contact_label", "LINE") if isinstance(_cfg_a, dict) else "LINE"
    _usage_lbl_a = _cfg_a.get("visit_label", _cfg_a["visit_label"]) if isinstance(_cfg_a, dict) else "来店"
    _revenue_lbl_a = _cfg_a.get("spend_label", "利用金額") if isinstance(_cfg_a, dict) else "利用金額"
    if _industry_a == "nightlife":
        _industry_specific_fields = (
            f"指名履歴: {c.get('nomination_history','')}\n"
            f"承認欲求: {c.get('approval_tendency','')} / 疑似恋愛傾向: {c.get('pseudo_love_tendency','')} / 執着傾向: {c.get('obsession_tendency','')}\n"
            f"LINE返信率: {c.get('line_reply_rate','')}% / 活発時間: {c.get('line_active_time','')} / 既読速度: {c.get('read_speed','')}\n"
            f"深夜反応率: {c.get('late_night_rate','')}%\n"
        )
    elif _industry_a == "retail":
        _industry_specific_fields = (
            f"購買動機: {c.get('purchase_motivation','')} / ブランド志向: {c.get('brand_preference','')} / セール反応: {c.get('sale_response','')}\n"
            f"{_contact_lbl_a}反応率: {c.get('line_reply_rate','')}%\n"
        )
    elif _industry_a == "b2b":
        _industry_specific_fields = (
            f"決裁権限: {c.get('decision_authority','')} / 検討フェーズ: {c.get('consideration_phase','')}\n"
            f"競合比較: {c.get('competitor_comparison','')} / 稟議状況: {c.get('approval_status','')}\n"
        )
    elif _industry_a == "beauty":
        _industry_specific_fields = (
            f"美容悩み: {c.get('beauty_concern','')} / 美容感度: {c.get('beauty_sensitivity','')} / 施術不安: {c.get('treatment_anxiety','')}\n"
        )
    elif _industry_a == "fitness":
        _industry_specific_fields = (
            f"目標: {c.get('fitness_goal','')} / モチベーション: {c.get('motivation','')} / 怪我リスク: {c.get('injury_risk','')}\n"
        )
    elif _industry_a == "realestate":
        _industry_specific_fields = (
            f"希望条件: {c.get('desired_conditions','')} / 予算: {c.get('budget','')}\n"
            f"他社比較: {c.get('competitor_comparison','')} / 審査状況: {c.get('screening_status','')}\n"
        )
    else:
        _industry_specific_fields = f"利用動機: {c.get('purchase_motivation', c.get('notes',''))}\n"
    prompt = (
        f"以下の顧客情報を分析し、{_industry_a}顧客対応AIブリーフをJSON形式で生成してください。\n"
        "【重要】必ずJSONのみ出力。前後に説明文・コードブロック・改行を一切付けないこと。\n"
        f"《業種》{_industry_a} / {_cfg_a['staff_noun']} / {_cfg_a['visit_label']}・{_cfg_a['spend_label']} / KPI:{_kpi_a}\n"
        f"《リスク》{_cfg_a['risk_focus']} / 《推奨アクション》{_cfg_a['recommended_actions']}\n\n"
        f"顧客名: {c.get('name','')}\n"
        f"年齢: {c.get('age','')} / 職業: {c.get('occupation','')} / エリア: {c.get('area','')}\n"
        f"{_usage_lbl_a}回数: {c.get('visit_count','')}回 / {_usage_lbl_a}周期: {c.get('visit_cycle','')}日 / 最終{_usage_lbl_a}: {c.get('last_visit','')}\n"
        f"累計{_revenue_lbl_a}: {c.get('spend_total','')}円 / 滞在時間: {c.get('stay_time','')}分\n"
        f"オプション履歴: {c.get('option_history','')}\n"
        f"趣味: {c.get('hobbies','')} / 好む会話: {c.get('good_topics','')} / NGトピック: {c.get('ng_topics','')}\n"
        f"コンプレックス: {c.get('complex','')} / ストレス: {c.get('stress_state','')}\n"
        + _industry_specific_fields +
        f"ドタキャン率: {c.get('cancel_rate','')}% / 当日予約率: {c.get('same_day_rate','')}%\n"
        f"SNS閲覧傾向: {c.get('sns_view_tendency','')}\n"
        f"クレーム歴: {c.get('claim_history','')} / 攻撃性: {c.get('aggression','')}\n"
        f"境界違反: {c.get('boundary_violation','')} / メンタル不安定: {c.get('mental_instability','')} / 地雷: {c.get('landmine_history','')}\n"
        f"現在温度: {c.get('temp','')} / 心理タイプ: {c.get('psych_type','')} / 現在欲求: {c.get('current_desire','')} / 離脱危険度: {c.get('churn_risk','')}\n"
        f"メモ: {c.get('notes','')}\n\n"
        "OUTPUT FORMAT (JSONのみ出力。他一切不要):\n"
        "{\n"
        '  \"brief\": {\"温度\":\"現在状態\",\"現在欲求\":\"最も必要なもの\",\"推奨対応\":\"対応姿勢\",\"禁止事項\":\"避けること\",\"次回予測\":\"〇日以内\"},\n'
        '  \"style_proposal\": {\"style\":\"推奨スタイル名\",\"points\":[\"p1\",\"p2\",\"p3\"]},\n'
        '  \"line_strategy\": {\"推奨トーン\":\"短文共感型\",\"送信タイミング\":\"時間帯\",\"禁止行動\":\"禁止事項\",\"特別感演出\":\"方法\"},\n'
        '  \"upsell\": {\"オプション提案適性\":\"高中低\",\"長時間化可能性\":\"高中低\",\"VIP化可能性\":\"高中低\",\"推奨アプローチ\":\"方法\"},\n'
        '  \"churn_analysis\": {\"離脱確率\":\"低中高\",\"危険度\":\"低中高\",\"危険理由\":\"理由\",\"推奨フォロー\":\"具体的行動\"},\n'
        '  \"new_temp\": \"S/A/B/C/Dのいずれか\",\n'
        '  \"new_desire\": \"欲求タイプ\"\n'
        "}"
    )
    try:
        raw = call_llm(
            system_prompt=f"あなたは{_industry_a}業種の顧客心理解析AIです。{_cfg_a['staff_noun']}の視点で分析し、指定されたJSON形式のみで出力してください。前後に説明文・コードブロックは一切付けないこと。",
            messages=[{"role":"user","content":prompt}],
            ai_tier="core", max_tokens=3500,
        )
        text = str(raw).strip()
        text = _re2.sub(r"^```json\s*","",text,flags=_re2.MULTILINE).strip()
        text = _re2.sub(r"^```\s*","",text,flags=_re2.MULTILINE).strip()
        text = _re2.sub(r"```\s*$","",text,flags=_re2.MULTILINE).strip()
        text = _re2.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]","",text)
        s = text.find("{"); e2 = text.rfind("}")
        if s == -1 or e2 == -1:
            raise ValueError("no JSON")
        try:
            result = _crm_json.loads(text[s:e2+1])
        except _crm_json.JSONDecodeError as _je2:
            print(f"[CRM_JSON_ERR] raw={text[s:s+300]}", flush=True)
            raise _crm_json.JSONDecodeError(_je2.msg, _je2.doc, _je2.pos)
        customer_id = c.get("id","")
        if customer_id and result.get("new_temp") and result["new_temp"] != c.get("temp",""):
            try:
                from api.core.firestore_client import get_db as _gdb
                _gdb().collection("users").document(uid).collection("crm_customers").document(customer_id).collection("state_transitions").add({
                    "previous_temperature": c.get("temp",""),
                    "current_temperature": result["new_temp"],
                    "previous_desire": c.get("current_desire",""),
                    "current_desire": result.get("new_desire", c.get("current_desire","")),
                    "trigger_reason": result.get("churn_analysis",{}).get("危険理由","AI分析による更新"),
                    "detected_at": _dt2.datetime.utcnow().isoformat(),
                    "confidence_score": 0.7,
                })
            except Exception:
                pass
        try:
            import datetime as _ucrm_dt
            from api.core.firestore_client import get_db as _ucrm_db
            _ucrm_uid = payload["uid"]; _ucrm_tid = payload.get("tenant_id","default")
            _ucrm_db().collection("usage_logs").add({"user_id":_ucrm_uid,"tenant_id":_ucrm_tid,"purpose_mode":"customer_ai","diagnosis_type":"customer_ai","prompt":"","timestamp":(_ucrm_dt.datetime.utcnow()+_ucrmdatetime.timedelta(hours=9)).strftime("%Y-%m-%d %H:%M:%S"),"is_admin_test":False})
        except Exception:
            pass
        return {"ok": True, "result": result}
    except _crm_json.JSONDecodeError as je:
        print(f"[CRM_JSON_ERR] raw={text[:1000]}", flush=True)
        raise HTTPException(status_code=502, detail=f"JSON解析失敗: {str(je)[:120]}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"分析失敗: {str(e)}")
# ========== CRM: 状態遷移履歴取得 ==========
@router.get("/crm_transitions/{customer_id}")
def crm_transitions(customer_id: str, payload: dict = Depends(verify_token)):
    uid = payload["uid"]
    if not _crm_feat(uid, "diag_crm"):
        raise HTTPException(status_code=403, detail="この機能はAPEX・ULTRAプラン限定です")
    try:
        from api.core.firestore_client import get_db as _gdb
        docs = list(
            _gdb().collection("users").document(uid)
            .collection("crm_customers").document(customer_id)
            .collection("state_transitions")
            .stream()
        )
        transitions = []
        for d in docs:
            row = d.to_dict() or {}
            row["id"] = d.id
            transitions.append(row)
        transitions.sort(key=lambda x: x.get("detected_at",""), reverse=True)
        return {"ok": True, "transitions": transitions[:20]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ========== CRM: 優先順位エンジン ==========
@router.get("/crm_priority")
def crm_priority(payload: dict = Depends(verify_token)):
    uid = payload["uid"]
    if not _crm_feat(uid, "diag_crm"):
        raise HTTPException(status_code=403, detail="この機能はAPEX・ULTRAプラン限定です")
    try:
        import datetime as _dt3
        from api.core.firestore_client import get_db as _gdb
        import json as _json3

        db3 = _gdb()
        docs = list(db3.collection("users").document(uid).collection("crm_customers").stream())
        customers = []
        for d in docs:
            row = d.to_dict() or {}
            row["id"] = d.id
            customers.append(row)

        today = _dt3.date.today()

        # ===== 直近7日イベント集計 =====
        def _calc_event_score(events):
            score = 0
            reasons = []
            for ev in events:
                txt = " ".join([
                    str(ev.get("event_category","")),
                    str(ev.get("event_type","")),
                    str(ev.get("event_value","")),
                    str(ev.get("metadata","")),
                ])
                # 反応低下（業種横断）
                _low_react = ["未返信","未読","既読無視","電話不在","連絡途絶","ブロック","ブロック解除"]
                if any(k in txt for k in _low_react) and "反応低下イベント" not in reasons:
                    score += 35
                    reasons.append("反応低下イベント")
                # キャンセル系（業種横断）
                _cancel = ["キャンセル","無断キャンセル","ノーショー","欠席","延期","ドタキャン"]
                if any(k in txt for k in _cancel) and "キャンセル系イベント" not in reasons:
                    score += 30
                    reasons.append("キャンセル系イベント")
                # クレーム系（業種横断）
                _claim = ["クレーム","仕上がり不満","悪評投稿","返品","契約外要求","トラブル"]
                if any(k in txt for k in _claim) and "クレーム系イベント" not in reasons:
                    score += 40
                    reasons.append("クレーム系イベント")
                # 売上期待（業種横断）
                _revenue = ["予約","来店","来館","購入","高額購入","高額利用","延長","指名","契約","更新","見積依頼","決裁者同席","再来店","物販購入","オプション","VIP"]
                if any(k in txt for k in _revenue) and "売上期待イベント" not in reasons:
                    score += 20
                    reasons.append("売上期待イベント")
                # 離脱・失注（業種横断）
                _churn = ["退会相談","解約相談","休会相談","失注","予算NG","他社比較","競合比較","退会","失注"]
                if any(k in txt for k in _churn) and "離脱・失注イベント" not in reasons:
                    score += 35
                    reasons.append("離脱・失注イベント")
                # 依存・境界（業種横断）
                _dep = ["感情依存","境界","違反","依存","執着","深夜","連投","不安定"]
                if any(k in txt for k in _dep) and "依存・境界系イベント" not in reasons:
                    score += 35
                    reasons.append("依存・境界系イベント")
            return min(100, score), list(dict.fromkeys(reasons))[:3]

        _event_map = {}
        try:
            _seven_days_ago = today - _dt3.timedelta(days=7)
            _event_docs = list(db3.collection("users").document(uid).collection("customer_event_stream").stream())
            for _ed in _event_docs:
                _ev = _ed.to_dict() or {}
                _cid = _ev.get("customer_id","")
                _ts = _ev.get("timestamp","")
                if not _cid or not _ts:
                    continue
                try:
                    _d = _dt3.date.fromisoformat(str(_ts)[:10])
                except Exception:
                    continue
                if _d < _seven_days_ago:
                    continue
                _event_map.setdefault(_cid, []).append(_ev)
        except Exception as _emap_err:
            print(f"[ERROR] crm_priority event_map: {_emap_err}")
            _event_map = {}

        def _safe_float(v, default=0.0):
            try: return float(str(v).replace("%","").strip() or default)
            except: return default

        def _safe_int(v, default=0):
            try: return int(str(v).replace("日","").strip() or default)
            except: return default

        def _days_since(date_str):
            if not date_str: return 999
            try: return (today - _dt3.date.fromisoformat(str(date_str)[:10])).days
            except: return 999

        # ===== イベント発生判定 & Gemini推論 =====
        def _needs_inference(c):
            ai_inferred_at = c.get("ai_inferred_at","")
            updated_at = c.get("updated_at","")
            if not ai_inferred_at:
                return True
            if updated_at and updated_at > ai_inferred_at:
                return True
            return False

        def _run_gemini_inference(c, uid, db3):
            """生データからGeminiがtemp/churn/psych_type/line_directionを推論"""
            try:
                days_since = _days_since(c.get("last_visit",""))
                cycle_days = _safe_int(c.get("visit_cycle",0))
                cycle_lag = days_since - cycle_days if cycle_days > 0 else 0
                _industry_g = c.get("industry") or "other"
                _cfg_g = _get_crm_cfg(_industry_g)
                _contact_lbl_g = _cfg_g.get("contact_label", "LINE") if isinstance(_cfg_g, dict) else "LINE"
                _usage_lbl_g = _cfg_g.get("visit_label", "来店") if isinstance(_cfg_g, dict) else "来店"
                _revenue_lbl_g = _cfg_g.get("spend_label", "利用金額") if isinstance(_cfg_g, dict) else "利用金額"
                _psych_types_g = _cfg_g.get("psych_types", "承認型/癒し型/疑似恋愛型/孤独回避型/支配型/会話型/比較検討型/ストレス逃避型") if isinstance(_cfg_g, dict) else "承認型/癒し型/疑似恋愛型/孤独回避型/支配型/会話型/比較検討型/ストレス逃避型"
                if _industry_g == "nightlife":
                    _extra_g = (
                        "LINE返信率: " + str(c.get("line_reply_rate","")) + "%\n"
                        "LINE活発時間帯: " + str(c.get("line_active_time","")) + "\n"
                        "深夜返信率: " + str(c.get("late_night_rate","")) + "%\n"
                        "承認欲求傾向: " + str(c.get("approval_tendency","")) + "\n"
                        "疑似恋愛傾向: " + str(c.get("pseudo_love_tendency","")) + "\n"
                        "執着傾向: " + str(c.get("obsession_tendency","")) + "\n"
                    )
                    _desire_opts_g = "承認不足/癒し不足/独占欲増加/飽き/比較中/疲労蓄積/依存進行/距離不安/刺激不足"
                elif _industry_g == "retail":
                    _extra_g = (
                        _contact_lbl_g + "反応率: " + str(c.get("line_reply_rate","")) + "%\n"
                        "購買動機: " + str(c.get("purchase_motivation","")) + "\n"
                        "ブランド志向: " + str(c.get("brand_preference","")) + "\n"
                        "セール反応: " + str(c.get("sale_response","")) + "\n"
                    )
                    _desire_opts_g = "再購入意欲/高単価化/休眠化/比較中/クレーム懸念/口コミ化期待/定期顧客化"
                elif _industry_g == "b2b":
                    _extra_g = (
                        "決裁権限: " + str(c.get("decision_authority","")) + "\n"
                        "検討フェーズ: " + str(c.get("consideration_phase","")) + "\n"
                        "競合比較: " + str(c.get("competitor_comparison","")) + "\n"
                        "稟議状況: " + str(c.get("approval_status","")) + "\n"
                    )
                    _desire_opts_g = "商談化/契約化/失注危険/競合比較中/稟議停滞/アップセル機会/継続更新"
                elif _industry_g == "beauty":
                    _extra_g = (
                        _contact_lbl_g + "反応率: " + str(c.get("line_reply_rate","")) + "%\n"
                        "美容悩み: " + str(c.get("beauty_concern","")) + "\n"
                        "美容感度: " + str(c.get("beauty_sensitivity","")) + "\n"
                        "施術不安: " + str(c.get("treatment_anxiety","")) + "\n"
                    )
                    _desire_opts_g = "再来店/指名固定化/新メニュー興味/休眠化/口コミ化/コース継続/他店比較"
                elif _industry_g == "fitness":
                    _extra_g = (
                        "目標: " + str(c.get("fitness_goal","")) + "\n"
                        "モチベーション: " + str(c.get("motivation","")) + "\n"
                        "怪我リスク: " + str(c.get("injury_risk","")) + "\n"
                        "欠席率: " + str(c.get("cancel_rate","")) + "%\n"
                    )
                    _desire_opts_g = "継続意欲/目標達成/退会懸念/モチベ低下/更新機会/紹介意欲/怪我リスク"
                elif _industry_g == "realestate":
                    _extra_g = (
                        "希望条件: " + str(c.get("desired_conditions","")) + "\n"
                        "予算: " + str(c.get("budget","")) + "\n"
                        "他社比較: " + str(c.get("competitor_comparison","")) + "\n"
                        "審査状況: " + str(c.get("screening_status","")) + "\n"
                    )
                    _desire_opts_g = "内見化/申込化/失注危険/他社成約/条件調整/審査懸念/成約機会"
                else:
                    _extra_g = _contact_lbl_g + "反応率: " + str(c.get("line_reply_rate","")) + "%\n"
                    _desire_opts_g = "継続/離脱懸念/成約機会/比較中/クレーム懸念"
                _prompt_lines = [
                    "あなたは" + _industry_g + "業種の顧客状態推論AIです。",
                    "《業種設定》" + _cfg_g["staff_noun"] + " / " + _usage_lbl_g + " / KPI:" + ",".join(_cfg_g["kpi"]) + " / リスク:" + _cfg_g["risk_focus"],
                    "以下の顧客生データから、顧客の現在状態を推論してください。",
                    "【顧客生データ】",
                    "名前: " + str(c.get("name","")),
                    "最終" + _usage_lbl_g + "日: " + str(c.get("last_visit","")) + "（" + str(days_since) + "日前）",
                    _usage_lbl_g + "周期: " + str(cycle_days) + "日（周期からの乖離: " + str(cycle_lag) + "日）",
                    _usage_lbl_g + "回数: " + str(c.get("visit_count",0)) + "回",
                    "累計" + _revenue_lbl_g + ": " + str(c.get("spend_total",0)) + "円",
                    _extra_g,
                    "ドタキャン率: " + str(c.get("cancel_rate","")) + "%",
                    "当日予約率: " + str(c.get("same_day_rate","")) + "%",
                    "会話メモ: " + str(c.get("notes","")),
                    "趣味・好み: " + str(c.get("hobbies","")),
                    "NGトピック: " + str(c.get("ng_topics","")),
                    "ストレス状態: " + str(c.get("stress_state","")),
                    "メンタル不安定: " + str(c.get("mental_instability","")),
                    "クレーム歴: " + str(c.get("claim_history","")),
                    "境界違反: " + str(c.get("boundary_violation","")),
                    "【推論指示】",
                    "以下をJSON形式で返せ。説明不要。JSONのみ。",
                    "{",
                    "  \"temp\": \"S/A/B/C/Dのいずれか（S:最重要 A:高ロイヤル B:安定 C:離脱兆候 D:休眠）\",",
                    "  \"churn_risk\": \"低/中/高のいずれか\",",
                    "  \"psych_type\": \"" + _psych_types_g + "のいずれか\",",
                    "  \"current_desire\": \"" + _desire_opts_g + "のいずれか\",",
                    "  \"line_direction\": \"今日の" + _contact_lbl_g + "推奨方向性を20字以内で。断定禁止。傾向として記述\",",
                    "  \"line_forbidden\": \"禁止事項を20字以内で。なければ「なし」\",",
                    "  \"line_timing\": \"推奨タイミングを15字以内で\",",
                    "  \"inference_reason\": \"推論理由を50字以内で\"",
                    "}",
                ]
                prompt = "\n".join(_prompt_lines)
                result_text = call_llm(system_prompt="あなたは" + str(c.get("industry","other")) + "業種の顧客状態推論AIです。" + _cfg_g["staff_noun"] + "の視点で分析しJSONのみ出力してください。", messages=[{"role":"user","content":prompt}], ai_tier="core", max_tokens=512)
                print(f"[CRM_GEMINI_RAW] customer={c.get('name')} raw={str(result_text)[:500]}", flush=True)
                # JSON抽出
                import re as _re
                m = _re.search(r'\{.*\}', result_text, _re.DOTALL)
                if not m:
                    print(f"[CRM_GEMINI_ERR] customer={c.get('name')} no JSON found in result", flush=True)
                    return None
                result = _json3.loads(m.group())
                print(f"[CRM_GEMINI_OK] customer={c.get('name')} direction={result.get('line_direction')} reason={result.get('inference_reason','')}", flush=True)
                # Firestoreに推論結果を保存
                now_str = (_dt3.datetime.utcnow() + _dt3.timedelta(hours=9)).isoformat()
                update_data = {
                    "temp": result.get("temp", c.get("temp","B")),
                    "churn_risk": result.get("churn_risk", c.get("churn_risk","低")),
                    "psych_type": result.get("psych_type", c.get("psych_type","")),
                    "current_desire": result.get("current_desire", c.get("current_desire","")),
                    "ai_inferred_at": now_str,
                    "ai_line_direction": result.get("line_direction",""),
                    "ai_line_forbidden": result.get("line_forbidden","なし"),
                    "ai_line_timing": result.get("line_timing",""),
                    "ai_inference_reason": result.get("inference_reason",""),
                }
                db3.collection("users").document(uid).collection("crm_customers").document(c["id"]).set(update_data, merge=True)
                return result
            except Exception as _e:
                print(f"[CRM_GEMINI_ERR] industry={_industry_g} customer={c.get('name')} err={str(_e)}", flush=True)
                import traceback
                traceback.print_exc()
                return None

        # 各顧客でイベント判定→必要なら推論
        print(f"[CRM_LOOP_START] customers_count={len(customers)}", flush=True)
        updated_customers = []
        for c in customers:
            _needs = _needs_inference(c)
            print(f"[CRM_INFER_CHECK] customer={c.get('name')} id={c.get('id')} needs={_needs} ai_inferred_at={c.get('ai_inferred_at')} updated_at={c.get('updated_at')}", flush=True)
            if _needs and c.get("id"):
                print(f"[CRM_INFER_RUN] customer={c.get('name')}", flush=True)
                result = _run_gemini_inference(c, uid, db3)
                if result:
                    c["temp"] = result.get("temp", c.get("temp","B"))
                    c["churn_risk"] = result.get("churn_risk", c.get("churn_risk","低"))
                    c["psych_type"] = result.get("psych_type", c.get("psych_type",""))
                    c["current_desire"] = result.get("current_desire", c.get("current_desire",""))
                    c["ai_line_direction"] = result.get("line_direction","")
                    c["ai_line_forbidden"] = result.get("line_forbidden","なし")
                    c["ai_line_timing"] = result.get("line_timing","")
                    c["ai_inference_reason"] = result.get("inference_reason","")
            updated_customers.append(c)

        # ===== 生データベースの特徴量生成 & 優先アクション分類 =====
        contact_now = []
        churn_danger = []
        revenue_expect = []
        leave_alone = []
        mental_danger = []

        for c in updated_customers:
            name = c.get("name","")
            temp = c.get("temp","B")
            churn = c.get("churn_risk","低")
            days_since = _days_since(c.get("last_visit",""))
            cycle_days = _safe_int(c.get("visit_cycle",0))
            line_rate = _safe_float(c.get("line_reply_rate",0))
            late_night = _safe_float(c.get("late_night_rate",0))
            cancel_rate = _safe_float(c.get("cancel_rate",0))
            spend_total = _safe_int(c.get("spend_total",0))
            visit_count = _safe_int(c.get("visit_count",0))
            avg_spend = spend_total // visit_count if visit_count > 0 else 0
            obsession = c.get("obsession_tendency","")
            mental = c.get("mental_instability","")
            pseudo_love = c.get("pseudo_love_tendency","")
            claim = c.get("claim_history","")
            cycle_lag = days_since - cycle_days if cycle_days > 0 else 0

            # 特徴量スコアリング（生データベース）
            # churn_score
            churn_score = 0
            if cycle_days > 0 and cycle_lag > 7: churn_score += 30
            if cycle_days > 0 and cycle_lag > 14: churn_score += 20
            if line_rate < 30: churn_score += 25
            if line_rate < 10: churn_score += 15
            if cancel_rate > 30: churn_score += 15
            if days_since > 60: churn_score += 20
            if temp == "C": churn_score += 10
            if churn == "高": churn_score += 10
            churn_score = min(100, churn_score)

            # dep_score
            dep_score = 0
            if late_night > 60: dep_score += 25
            if obsession: dep_score += 25
            if pseudo_love and "強" in str(pseudo_love): dep_score += 20
            if mental: dep_score += 15
            if temp == "S": dep_score += 15
            dep_score = min(100, dep_score)

            # rev_score
            rev_score = 0
            if avg_spend > 50000: rev_score += 35
            elif avg_spend > 20000: rev_score += 20
            if visit_count > 10: rev_score += 20
            if visit_count > 20: rev_score += 10
            if line_rate > 70: rev_score += 15
            if cancel_rate < 10: rev_score += 10
            if churn_score < 30: rev_score += 10
            rev_score = min(100, rev_score)

            # revisit_prob
            revisit_prob = 0
            if cycle_days > 0 and 0 <= cycle_lag <= 3: revisit_prob += 40
            if cycle_days > 0 and cycle_lag <= 0: revisit_prob += 20
            if line_rate > 70: revisit_prob += 20
            if cancel_rate < 10: revisit_prob += 15
            if visit_count > 5: revisit_prob += 5
            revisit_prob = min(100, revisit_prob)

            # vip_prob
            vip_prob = 0
            if avg_spend > 50000: vip_prob += 35
            if visit_count > 20: vip_prob += 25
            if line_rate > 80: vip_prob += 20
            if cancel_rate < 5: vip_prob += 10
            if churn_score < 20: vip_prob += 10
            vip_prob = min(100, vip_prob)

            # mental_score
            mental_score = 0
            if dep_score > 60: mental_score += 40
            if obsession: mental_score += 25
            if mental: mental_score += 25
            if claim: mental_score += 10
            mental_score = min(100, mental_score)

            # LINE方向性（AI推論済みがあればそれを使用、なければフォールバック）
            line_direction = c.get("ai_line_direction","")
            line_forbidden = c.get("ai_line_forbidden","なし")
            line_timing = c.get("ai_line_timing","")
            ai_reason = c.get("ai_inference_reason","")

            _cid = c.get("id","")
            _events = _event_map.get(_cid, [])
            event_score, event_reasons = _calc_event_score(_events)
            priority_score = min(100, round(
                churn_score * 0.30 +
                revisit_prob * 0.20 +
                rev_score * 0.20 +
                mental_score * 0.15 +
                event_score * 0.15
            ))
            _industry_se = c.get("industry") or "other"
            _cfg_se = _get_crm_cfg(_industry_se)
            _contact_lbl_se = _cfg_se.get("contact_label", "連絡") if isinstance(_cfg_se, dict) else "連絡"
            _usage_lbl_se = _cfg_se.get("visit_label", "利用") if isinstance(_cfg_se, dict) else "利用"
            _priority_axis_se = _cfg_se.get("priority_axis", "") if isinstance(_cfg_se, dict) else ""
            _success_kpi_se = _cfg_se.get("success_kpi", _cfg_se.get("kpi", [])) if isinstance(_cfg_se, dict) else []
            _danger_kpi_se = _cfg_se.get("danger_kpi", []) if isinstance(_cfg_se, dict) else []
            # line_direction fallback
            if not line_direction and priority_score >= 20:
                line_direction = f"{_contact_lbl_se}反応と{_usage_lbl_se}状況を確認"
                if not line_timing:
                    line_timing = "今日中"
            score_entry = {
                "name": name,
                "churn_risk_score": churn_score,
                "dependency_risk_score": dep_score,
                "revenue_score": rev_score,
                "revisit_probability": revisit_prob,
                "vip_probability": vip_prob,
                "mental_danger_score": mental_score,
                "line_direction": line_direction,
                "line_forbidden": line_forbidden,
                "line_timing": line_timing,
                "ai_reason": ai_reason,
                "event_score": event_score,
                "event_reasons": event_reasons,
                "priority_score": priority_score,
                "days_since": days_since,
                "industry": _industry_se,
                "priority_axis": _priority_axis_se,
                "success_kpi": _success_kpi_se,
                "danger_kpi": _danger_kpi_se,
            }
            # 分類（生データスコアベース・独立判定・重複排除）
            _cn_names = set(x["name"] for x in contact_now)
            _cd_names = set(x["name"] for x in churn_danger)
            _re_names = set(x["name"] for x in revenue_expect)
            _md_names = set(x["name"] for x in mental_danger)
            if churn_score >= 50 and name not in _cd_names:
                churn_danger.append(score_entry)
            _add_contact = False
            if cycle_days > 0 and 0 <= cycle_lag <= 5:
                _add_contact = True
            elif cycle_days > 0 and cycle_lag > 0 and churn_score < 50:
                _add_contact = True
            elif cycle_days == 0 and days_since <= 30:
                _add_contact = True
            elif event_score >= 50:
                _add_contact = True
            if _add_contact and name not in _cn_names:
                contact_now.append(score_entry)
            if rev_score >= 50 and churn_score < 50 and name not in _re_names:
                revenue_expect.append(score_entry)
            if cycle_days > 0 and days_since > cycle_days * 2 and churn_score < 50:
                leave_alone.append(score_entry)
            elif cycle_days == 0 and days_since > 60 and churn_score < 50:
                leave_alone.append(score_entry)
            if (mental_score >= 50 or (event_score >= 35 and mental_score >= 40)) and name not in _md_names:
                mental_danger.append(score_entry)


        churn_danger.sort(key=lambda x: x["priority_score"], reverse=True)
        contact_now.sort(key=lambda x: x["priority_score"], reverse=True)
        revenue_expect.sort(key=lambda x: x["priority_score"], reverse=True)
        mental_danger.sort(key=lambda x: x["priority_score"], reverse=True)

        # 優先アクション通知
        try:
            _pa_total = len(contact_now[:5]) + len(churn_danger[:5])
            if _pa_total > 0:
                _pa_parts = []
                if contact_now: _pa_parts.append(f"今日連絡すべき顧客 {len(contact_now[:5])}名")
                if churn_danger: _pa_parts.append(f"離脱危険顧客 {len(churn_danger[:5])}名")
                import uuid as _nuuid, datetime as _ndt
                _db2 = _gdb()
                _snap2 = _db2.collection("users").document(uid).get()
                _d2 = _snap2.to_dict() if _snap2.exists else {}
                _settings2 = _d2.get("notification_settings", {})
                if _settings2.get("notify_priority_action", True) is not False:
                    _today_str = (_dt3.datetime.utcnow() + _dt3.timedelta(hours=9)).strftime("%Y-%m-%d")
                    _existing = list(_db2.collection("notifications").document(uid).collection("items").stream())
                    _already_today = any(
                        x.to_dict().get("type") == "priority_action" and
                        str(x.to_dict().get("created_at",""))[:10] == _today_str
                        for x in _existing
                    )
                    if not _already_today:
                        _nid2 = _nuuid.uuid4().hex
                        _db2.collection("notifications").document(uid).collection("items").document(_nid2).set({
                            "notif_id": _nid2,
                            "uid": uid,
                            "type": "priority_action",
                            "title": "🚨 今日の優先アクションがあります",
                            "body": "、".join(_pa_parts),
                            "link_tab": "crm",
                            "read": False,
                            "created_at": (_dt3.datetime.utcnow() + _dt3.timedelta(hours=9)).isoformat(),
                        })
        except Exception as _pa_err:
            print(f"[ERROR] priority_action notification: {_pa_err}")

        return {
            "ok": True,
            "priority": {
                "contact_now": contact_now[:5],
                "churn_danger": churn_danger[:5],
                "revenue_expect": revenue_expect[:5],
                "leave_alone": leave_alone[:5],
                "mental_danger": mental_danger[:5],
            }
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
# ========== CRM: 店舗全体知能 ==========
# ========== CRM: 店舗全体知能 ==========
class StoreIntelligenceRequest(BaseModel):
    customers: list = []
    industry: str = "nightlife"

@router.post("/crm_store_intelligence")
def crm_store_intelligence(req: StoreIntelligenceRequest, payload: dict = Depends(verify_token)):
    uid = payload["uid"]
    if not _crm_feat(uid, "diag_crm"):
        raise HTTPException(status_code=403, detail="この機能はAPEX・ULTRAプラン限定です")
    if not req.customers:
        raise HTTPException(status_code=400, detail="顧客データがありません")
    import re as _re3
    customers = req.customers
    total = len(customers)
    # 温度分布
    temp_counts = {}
    for c in customers:
        t = c.get("temp", "B")
        temp_counts[t] = temp_counts.get(t, 0) + 1
    temp_sa_count = temp_counts.get("S", 0) + temp_counts.get("A", 0)
    temp_cd_count = temp_counts.get("C", 0) + temp_counts.get("D", 0)
    # 離脱危険
    churn_high = [c.get("name","") for c in customers if c.get("churn_risk") == "高"]
    churn_rate = round(len(churn_high) / total * 100, 1) if total > 0 else 0
    # 休眠顧客（温度C/D）
    dormant = [c.get("name","") for c in customers if c.get("temp") in ["C","D"]]
    # 高単価顧客（avg_spend >= 30000）
    high_value = [c.get("name","") for c in customers if (c.get("avg_spend") or 0) >= 30000]
    # LINE低返信率（line_reply_rate <= 30%）
    def _safe_float_si(v, default=0.0):
        try:
            return float(str(v).replace("%","").strip() or default)
        except:
            return default
    low_line = [c.get("name","") for c in customers if _safe_float_si(c.get("line_reply_rate", 0)) <= 30]
    # 心理タイプ分布
    psych_counts = {}
    for c in customers:
        pt = c.get("psych_type","")
        if pt: psych_counts[pt] = psych_counts.get(pt, 0) + 1
    top_psych = sorted(psych_counts.items(), key=lambda x: -x[1])[:3]

    _industry_si = req.industry or "nightlife"
    _cfg_si = _get_crm_cfg(_industry_si)
    _kpi_si = "、".join(_cfg_si["kpi"])
    prompt = (
        f"以下の店舗顧客データを分析し、経営提案をJSON形式で生成してください。\n"
        f"《業種》{_industry_si} / {_cfg_si['staff_noun']} / {_cfg_si['visit_label']} / KPI:{_kpi_si}\n"
        f"《リスク》{_cfg_si['risk_focus']} / 《アクション》{_cfg_si['recommended_actions']}\n"
        f"【重要】JSONのみ出力。説明文・コードブロック不要。\n\n"
        f"総顧客数: {total}名\n"
        f"温度分布: {temp_counts}\n"
        f"温度S/A（優良）: {temp_sa_count}名\n"
        f"温度C/D（離脱兆候）: {temp_cd_count}名\n"
        f"高離脱危険顧客数: {len(churn_high)}名 ({churn_rate}%)\n"
        f"高離脱危険顧客: {churn_high}\n"
        f"休眠顧客数: {len(dormant)}名\n"
        f"休眠顧客: {dormant}\n"
        f"高単価顧客数（3万円以上）: {len(high_value)}名\n"
        f"高単価顧客: {high_value}\n"
        f"LINE低返信率顧客数: {len(low_line)}名\n"
        f"LINE低返信率顧客: {low_line}\n"
        f"心理タイプ上位3: {top_psych}\n\n"
        "OUTPUT FORMAT:\n"
        "{\n"
        '  \"temp_distribution_analysis\": \"温度分布から見た顧客層の分析\",\n'
        '  \"churn_risk_summary\": \"離脱危険顧客の傾向と対策\",\n'
        '  \"high_value_customer_summary\": \"高単価顧客の維持戦略\",\n'
        '  \"dormant_customer_summary\": \"休眠顧客の復帰施策\",\n'
        '  \"communication_risk_summary\": \"LINE低返信率顧客へのコミュニケーション改善\",\n'
        '  \"revenue_stagnation\": \"売上停滞リスクの分析\",\n'
        '  \"line_fatigue\": \"LINE疲労リスクの分析\",\n'
        '  \"proposals\": [\"提案1\",\"提案2\",\"提案3\",\"提案4\",\"提案5\"]\n'
        "}"
    )
    try:
        raw = call_llm(
            system_prompt=f"あなたは{_industry_si}業種の店舗経営AIアドバイザーです。{_cfg_si['staff_noun']}の視点で顧客データから課題と改善提案をJSON形式のみで出力してください。",
            messages=[{"role":"user","content":prompt}],
            ai_tier="core", max_tokens=2000,
        )
        text = str(raw).strip()
        text = _re3.sub(r"^```json\s*","",text,flags=_re3.MULTILINE).strip()
        text = _re3.sub(r"^```\s*","",text,flags=_re3.MULTILINE).strip()
        text = _re3.sub(r"```\s*$","",text,flags=_re3.MULTILINE).strip()
        s = text.find("{"); e2 = text.rfind("}")
        if s == -1 or e2 == -1:
            raise ValueError("no JSON")
        result = _crm_json.loads(text[s:e2+1])
        return {"ok": True, "result": result}
    except _crm_json.JSONDecodeError as je:
        print(f"[CRM_JSON_ERR] raw={text[:1000]}", flush=True)
        raise HTTPException(status_code=502, detail=f"JSON解析失敗: {str(je)[:120]}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"店舗分析失敗: {str(e)}")

# ========== CRM: キャスト相性学習 ==========
class CastModel(BaseModel):
    id: str = ""
    name: str = ""
    strong_types: str = ""
    weak_types: str = ""
    line_success_rate: str = ""
    avg_upsell_rate: str = ""
    notes: str = ""

class CastAffinityRequest(BaseModel):
    customer: dict = {}
    cast_id: str = ""
    industry: str = "nightlife"

class CastFeedbackRequest(BaseModel):
    customer_id: str = ""
    cast_id: str = ""
    cast_name: str = ""
    result: str = ""
    notes: str = ""

def _cast_col(uid: str):
    from api.core.firestore_client import get_db as _gdb
    return _gdb().collection("users").document(uid).collection("crm_casts")

@router.get("/crm_cast_list")
def crm_cast_list(payload: dict = Depends(verify_token)):
    uid = payload["uid"]
    if not _crm_feat(uid, "diag_crm"):
        raise HTTPException(status_code=403, detail="この機能はAPEX・ULTRAプラン限定です")
    try:
        docs = list(_cast_col(uid).stream())
        casts = []
        for d in docs:
            row = d.to_dict() or {}
            row["id"] = d.id
            casts.append(row)
        return {"ok": True, "casts": casts}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/crm_cast")
def crm_cast_create(req: CastModel, payload: dict = Depends(verify_token)):
    uid = payload["uid"]
    if not _crm_feat(uid, "diag_crm"):
        raise HTTPException(status_code=403, detail="この機能はAPEX・ULTRAプラン限定です")
    if not req.name.strip():
        raise HTTPException(status_code=400, detail="キャスト名は必須です")
    try:
        import datetime as _dt4
        data = req.dict()
        data.pop("id", None)
        data["created_at"] = _dt4.datetime.utcnow().isoformat()
        ref = _cast_col(uid).add(data)
        return {"ok": True, "id": ref[1].id}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.put("/crm_cast/{cast_id}")
def crm_cast_update(cast_id: str, req: CastModel, payload: dict = Depends(verify_token)):
    uid = payload["uid"]
    if not _crm_feat(uid, "diag_crm"):
        raise HTTPException(status_code=403, detail="この機能はAPEX・ULTRAプラン限定です")
    try:
        import datetime as _dt4
        data = req.dict()
        data.pop("id", None)
        data["updated_at"] = _dt4.datetime.utcnow().isoformat()
        _cast_col(uid).document(cast_id).set(data, merge=True)
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.delete("/crm_cast/{cast_id}")
def crm_cast_delete(cast_id: str, payload: dict = Depends(verify_token)):
    uid = payload["uid"]
    if not _crm_feat(uid, "diag_crm"):
        raise HTTPException(status_code=403, detail="この機能はAPEX・ULTRAプラン限定です")
    try:
        _cast_col(uid).document(cast_id).delete()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/crm_cast_affinity")
def crm_cast_affinity(req: CastAffinityRequest, payload: dict = Depends(verify_token)):
    uid = payload["uid"]
    if not _crm_feat(uid, "diag_crm"):
        raise HTTPException(status_code=403, detail="この機能はAPEX・ULTRAプラン限定です")
    import re as _re4
    from api.core.firestore_client import get_db as _gdb
    try:
        # キャスト一覧取得
        industry = req.industry or "nightlife"
        _cfg_cast = _get_crm_cfg(industry)
        staff_noun = _cfg_cast["staff_noun"]
        cast_docs = list(_cast_col(uid).stream())
        casts = []
        for d in cast_docs:
            row = d.to_dict() or {}
            row["id"] = d.id
            casts.append(row)
        if not casts:
            raise HTTPException(status_code=400, detail=f"{staff_noun}が登録されていません")
        # フィードバック履歴取得
        fb_docs = list(_gdb().collection("users").document(uid).collection("crm_cast_feedback").stream())
        feedbacks = []
        for d in fb_docs:
            row = d.to_dict() or {}
            feedbacks.append(row)
        c = req.customer
        cast_info = "\n".join([f"- {cs.get('name','')}（得意:{cs.get('strong_types','')} / 苦手:{cs.get('weak_types','')} / 接触成功率:{cs.get('line_success_rate','')} / 備考:{cs.get('notes','')}）" for cs in casts])
        fb_info = "\n".join([f"- 顧客:{f.get('customer_id','')} × {staff_noun}:{f.get('cast_name','')} → {f.get('result','')} ({f.get('notes','')})" for f in feedbacks[-20:]])
        # 業種別顧客追加情報
        _cfg_backend_aff = _get_crm_cfg(industry)
        _affinity_axis = _cfg_backend_aff.get('affinity_axis', '') if isinstance(_cfg_backend_aff, dict) else ''
        _success_kpi_aff = '、'.join(_cfg_backend_aff.get('success_kpi', _cfg_backend_aff.get('kpi', []))) if isinstance(_cfg_backend_aff, dict) else ''
        _danger_kpi_aff = '、'.join(_cfg_backend_aff.get('danger_kpi', [])) if isinstance(_cfg_backend_aff, dict) else ''
        if industry == 'nightlife':
            _customer_extra_aff = (
                f"承認欲求: {c.get('approval_tendency','')} / 疑似恋愛傾向: {c.get('pseudo_love_tendency','')}\n"
                f"執着傾向: {c.get('obsession_tendency','')} / 攻撃性: {c.get('aggression','')}\n"
            )
        elif industry == 'retail':
            _customer_extra_aff = (
                f"購買動機: {c.get('purchase_motivation','')} / 価格感度: {c.get('price_sensitivity','')}\n"
                f"攻撃性: {c.get('aggression','')}\n"
            )
        elif industry == 'b2b':
            _customer_extra_aff = (
                f"決裁権限: {c.get('decision_authority','')} / 競合比較: {c.get('competitor_comparison','')}\n"
                f"攻撃性: {c.get('aggression','')}\n"
            )
        elif industry == 'beauty':
            _customer_extra_aff = (
                f"美容悩み: {c.get('beauty_concern','')} / 施術不安: {c.get('treatment_anxiety','')}\n"
                f"攻撃性: {c.get('aggression','')}\n"
            )
        elif industry == 'fitness':
            _customer_extra_aff = (
                f"目標: {c.get('fitness_goal','')} / 継続意欲: {c.get('motivation','')}\n"
                f"攻撃性: {c.get('aggression','')}\n"
            )
        elif industry == 'realestate':
            _customer_extra_aff = (
                f"希望条件: {c.get('desired_conditions','')} / 予算: {c.get('budget','')}\n"
                f"他社比較: {c.get('competitor_comparison','')}\n"
            )
        else:
            _customer_extra_aff = f"攻撃性: {c.get('aggression','')}\n"
        prompt = (
            f"以下の顧客情報と{staff_noun}情報を分析し、相性スコアをJSON形式で生成してください。\n"
            "【重要】JSONのみ出力。説明文・コードブロック不要。\n\n"
            f"【顧客情報】\n"
            f"名前: {c.get('name','')} / 心理タイプ: {c.get('psych_type','')} / 現在欲求: {c.get('current_desire','')}\n"
            f"温度: {c.get('temp','')} / NGトピック: {c.get('ng_topics','')} / コンプレックス: {c.get('complex','')}\n"
            + _customer_extra_aff +
            f"\n【業種別相性評価軸】\n{_affinity_axis}\n"
            f"【成功KPI】{_success_kpi_aff}\n"
            f"【避けるべきリスク】{_danger_kpi_aff}\n\n"
            f"【{staff_noun}一覧】\n{cast_info}\n\n"
            f"【過去フィードバック】\n{fb_info if fb_info else 'なし'}\n\n"
            "OUTPUT FORMAT:\n"
            "{\n"
            f'  "rankings": [\n'
            f'    {{"cast_name": "{staff_noun}名", "score": 85, "reason": "理由", "risk": "低中高", "approach": "推奨アプローチ"}}\n'
            f'  ],\n'
            f'  "best_cast": "最推奨{staff_noun}名",\n'
            f'  "danger_cast": "危険な組み合わせの{staff_noun}名または なし",\n'
            '  "danger_reason": "危険な理由",\n'
            '  "overall_advice": "総合アドバイス"\n'
            "}"
        )
        raw = call_llm(
            system_prompt=f"あなたは顧客×{staff_noun}相性解析AIです。指定JSON形式のみで出力してください。",
            messages=[{"role":"user","content":prompt}],
            ai_tier="core", max_tokens=2000,
        )
        text = str(raw).strip()
        text = _re4.sub(r"^```json\s*","",text,flags=_re4.MULTILINE).strip()
        text = _re4.sub(r"^```\s*","",text,flags=_re4.MULTILINE).strip()
        text = _re4.sub(r"```\s*$","",text,flags=_re4.MULTILINE).strip()
        s = text.find("{"); e2 = text.rfind("}")
        if s==-1 or e2==-1: raise ValueError("no JSON")
        result = _crm_json.loads(text[s:e2+1])
        # score clamp / risk validation
        _allowed_risk = ["低","中","高"]
        for _r in result.get("rankings", []):
            try:
                _r["score"] = max(0, min(100, int(_r.get("score", 50))))
            except Exception:
                _r["score"] = 50
            if _r.get("risk") not in _allowed_risk:
                _r["risk"] = "中"
        return {"ok": True, "result": result}
    except _crm_json.JSONDecodeError as je:
        print(f"[CRM_JSON_ERR] raw={text[:1000]}", flush=True)
        raise HTTPException(status_code=502, detail=f"JSON解析失敗: {str(je)[:120]}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"相性分析失敗: {str(e)}")

@router.post("/crm_cast_feedback")
def crm_cast_feedback(req: CastFeedbackRequest, payload: dict = Depends(verify_token)):
    uid = payload["uid"]
    if not _crm_feat(uid, "diag_crm"):
        raise HTTPException(status_code=403, detail="この機能はAPEX・ULTRAプラン限定です")
    try:
        import datetime as _dt4
        from api.core.firestore_client import get_db as _gdb
        _gdb().collection("users").document(uid).collection("crm_cast_feedback").add({
            "customer_id": req.customer_id,
            "cast_id": req.cast_id,
            "cast_name": req.cast_name,
            "result": req.result,
            "notes": req.notes,
            "recorded_at": _dt4.datetime.utcnow().isoformat(),
        })
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ========== CRM: フィードバック学習ループ ==========
class ActionFeedbackRequest(BaseModel):
    customer_id: str = ""
    customer_name: str = ""
    action_type: str = ""
    ai_proposal: str = ""
    executed: bool = False
    result: str = ""
    revenue_change: str = ""
    revisit: bool = False
    notes: str = ""

def _fb_col(uid: str):
    from api.core.firestore_client import get_db as _gdb
    return _gdb().collection("users").document(uid).collection("action_feedback_logs")

@router.post("/crm_action_feedback")
def crm_action_feedback(req: ActionFeedbackRequest, payload: dict = Depends(verify_token)):
    uid = payload["uid"]
    if not _crm_feat(uid, "diag_crm"):
        raise HTTPException(status_code=403, detail="この機能はAPEX・ULTRAプラン限定です")
    try:
        import datetime as _dt5
        data = req.dict()
        data["recorded_at"] = _dt5.datetime.utcnow().isoformat()
        _fb_col(uid).add(data)
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/crm_action_feedback_list")
def crm_action_feedback_list(payload: dict = Depends(verify_token)):
    uid = payload["uid"]
    if not _crm_feat(uid, "diag_crm"):
        raise HTTPException(status_code=403, detail="この機能はAPEX・ULTRAプラン限定です")
    try:
        docs = list(_fb_col(uid).stream())
        logs = []
        for d in docs:
            row = d.to_dict() or {}
            row["id"] = d.id
            logs.append(row)
        logs.sort(key=lambda x: x.get("recorded_at",""), reverse=True)
        return {"ok": True, "logs": logs[:50]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/crm_feedback_stats")
def crm_feedback_stats(payload: dict = Depends(verify_token)):
    uid = payload["uid"]
    if not _crm_feat(uid, "diag_crm"):
        raise HTTPException(status_code=403, detail="この機能はAPEX・ULTRAプラン限定です")
    import re as _re5
    try:
        docs = list(_fb_col(uid).stream())
        logs = [d.to_dict() or {} for d in docs]
        if not logs:
            return {"ok": True, "stats": None}
        total = len(logs)
        executed = sum(1 for l in logs if l.get("executed"))
        success = sum(1 for l in logs if l.get("result") in ["成功","再来あり"])
        revisit = sum(1 for l in logs if l.get("revisit"))
        action_types: dict = {}
        for l in logs:
            at = l.get("action_type","その他")
            if at not in action_types:
                action_types[at] = {"total":0,"success":0}
            action_types[at]["total"] += 1
            if l.get("result") in ["成功","再来あり"]:
                action_types[at]["success"] += 1
        # AI洞察生成
        logs_str = "\n".join([
            f"- {l.get('action_type','')}：{l.get('ai_proposal','')} → 実行:{l.get('executed','')} 結果:{l.get('result','')} 再来:{l.get('revisit','')}"
            for l in logs[-30:]
        ])
        prompt = (
            f"以下のAI提案フィードバックログを分析し、学習洞察をJSON形式で出力してください。\n"
            f"【重要】JSONのみ出力。\n\n"
            f"総件数:{total} 実行率:{round(executed/total*100)}% 成功率:{round(success/total*100) if executed>0 else 0}% 再来率:{round(revisit/total*100)}%\n\n"
            f"【ログ】\n{logs_str}\n\n"
            "OUTPUT:\n"
            "{\n"
            '  "effective_actions": ["効果的だったアクション1","アクション2"],\n'
            '  "ineffective_actions": ["効果なかったアクション1"],\n'
            '  "patterns": "成功パターンの分析",\n'
            '  "recommendations": ["今後の推奨方針1","推奨方針2","推奨方針3"]\n'
            "}"
        )
        raw = call_llm(
            system_prompt="フィードバック学習分析AIです。JSON形式のみで出力してください。",
            messages=[{"role":"user","content":prompt}],
            ai_tier="core", max_tokens=1500,
        )
        text = str(raw).strip()
        import re as _re5b
        text = _re5b.sub(r"^```json\s*","",text,flags=_re5b.MULTILINE).strip()
        text = _re5b.sub(r"^```\s*","",text,flags=_re5b.MULTILINE).strip()
        text = _re5b.sub(r"```\s*$","",text,flags=_re5b.MULTILINE).strip()
        s = text.find("{"); e2 = text.rfind("}")
        insight = _crm_json.loads(text[s:e2+1]) if s!=-1 and e2!=-1 else {}
        return {
            "ok": True,
            "stats": {
                "total": total,
                "executed": executed,
                "execution_rate": round(executed/total*100),
                "success_rate": round(success/executed*100) if executed>0 else 0,
                "revisit_rate": round(revisit/total*100),
                "action_types": action_types,
                "insight": insight,
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ========== CRM: イベントストリーム ==========
class CustomerEventRequest(BaseModel):
    customer_id: str = ""
    customer_name: str = ""
    event_category: str = ""
    event_type: str = ""
    event_value: str = ""
    cast_id: str = ""
    cast_name: str = ""
    metadata: str = ""

def _event_col(uid: str):
    from api.core.firestore_client import get_db as _gdb
    return _gdb().collection("users").document(uid).collection("customer_event_stream")

@router.post("/crm_event")
def crm_event_add(req: CustomerEventRequest, payload: dict = Depends(verify_token)):
    uid = payload["uid"]
    if not _crm_feat(uid, "diag_crm"):
        raise HTTPException(status_code=403, detail="この機能はAPEX・ULTRAプラン限定です")
    try:
        import datetime as _dt6
        data = req.dict()
        data["timestamp"] = _dt6.datetime.utcnow().isoformat()
        data["uid"] = uid
        _event_col(uid).add(data)
        if req.customer_id:
            from api.core.firestore_client import get_db as _gdb
            _upd = {
                "last_event_at": data["timestamp"],
                "last_contact_at": data["timestamp"],
                "updated_at": data["timestamp"],
            }
            _gdb().collection("users").document(uid).collection("crm_customers").document(req.customer_id).set(
                _upd, merge=True
            )
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/crm_event_list")
def crm_event_list(customer_id: str = "", limit: int = 50, payload: dict = Depends(verify_token)):
    uid = payload["uid"]
    if not _crm_feat(uid, "diag_crm"):
        raise HTTPException(status_code=403, detail="この機能はAPEX・ULTRAプラン限定です")
    try:
        col = _event_col(uid)
        if customer_id:
            docs = list(col.where("customer_id", "==", customer_id).stream())
        else:
            docs = list(col.stream())
        events = []
        for d in docs:
            row = d.to_dict() or {}
            row["id"] = d.id
            events.append(row)
        events.sort(key=lambda x: x.get("timestamp",""), reverse=True)
        return {"ok": True, "events": events[:limit]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.delete("/crm_event/{event_id}")
def crm_event_delete(event_id: str, payload: dict = Depends(verify_token)):
    uid = payload["uid"]
    if not _crm_feat(uid, "diag_crm"):
        raise HTTPException(status_code=403, detail="この機能はAPEX・ULTRAプラン限定です")
    try:
        _event_col(uid).document(event_id).delete()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ========== CRM: マルチRAG構造 ==========
class CrmRagChunkRequest(BaseModel):
    rag_type: str = ""
    title: str = ""
    content: str = ""
    tags: str = ""

def _rag_col(uid: str):
    from api.core.firestore_client import get_db as _gdb
    return _gdb().collection("users").document(uid).collection("crm_rag_chunks")

@router.post("/crm_rag_add")
def crm_rag_add(req: CrmRagChunkRequest, payload: dict = Depends(verify_token)):
    uid = payload["uid"]
    if not _crm_feat(uid, "diag_crm"):
        raise HTTPException(status_code=403, detail="この機能はAPEX・ULTRAプラン限定です")
    if not req.content.strip():
        raise HTTPException(status_code=400, detail="内容は必須です")
    try:
        import datetime as _dt7
        from api.core.rag import embed_text
        embedding = embed_text(req.content)
        import struct
        embedding_bytes = struct.pack(f"{len(embedding)}f", *embedding)
        data = {
            "rag_type": req.rag_type,
            "title": req.title,
            "content": req.content,
            "tags": req.tags,
            "embedding_bytes": embedding_bytes,
            "created_at": _dt7.datetime.utcnow().isoformat(),
            "uid": uid,
        }
        _rag_col(uid).add(data)
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/crm_rag_list")
def crm_rag_list(rag_type: str = "", payload: dict = Depends(verify_token)):
    uid = payload["uid"]
    if not _crm_feat(uid, "diag_crm"):
        raise HTTPException(status_code=403, detail="この機能はAPEX・ULTRAプラン限定です")
    try:
        col = _rag_col(uid)
        if rag_type:
            docs = list(col.where("rag_type", "==", rag_type).stream())
        else:
            docs = list(col.stream())
        chunks = []
        for d in docs:
            row = d.to_dict() or {}
            row.pop("embedding_bytes", None)
            row["id"] = d.id
            chunks.append(row)
        chunks.sort(key=lambda x: x.get("created_at",""), reverse=True)
        return {"ok": True, "chunks": chunks}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.delete("/crm_rag/{chunk_id}")
def crm_rag_delete(chunk_id: str, payload: dict = Depends(verify_token)):
    uid = payload["uid"]
    if not _crm_feat(uid, "diag_crm"):
        raise HTTPException(status_code=403, detail="この機能はAPEX・ULTRAプラン限定です")
    try:
        _rag_col(uid).document(chunk_id).delete()
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/crm_rag_search")
def crm_rag_search(req: dict, payload: dict = Depends(verify_token)):
    uid = payload["uid"]
    if not _crm_feat(uid, "diag_crm"):
        raise HTTPException(status_code=403, detail="この機能はAPEX・ULTRAプラン限定です")
    query = req.get("query","")
    rag_types = req.get("rag_types", [])
    top_k = req.get("top_k", 5)
    if not query.strip():
        raise HTTPException(status_code=400, detail="クエリは必須です")
    try:
        import struct, math
        from api.core.rag import embed_text
        q_vec = embed_text(query)
        col = _rag_col(uid)
        docs = list(col.stream())
        results = []
        for d in docs:
            row = d.to_dict() or {}
            if rag_types and row.get("rag_type","") not in rag_types:
                continue
            emb_bytes = row.get("embedding_bytes")
            if not emb_bytes:
                continue
            try:
                n = len(emb_bytes) // 4
                vec = list(struct.unpack(f"{n}f", emb_bytes))
                dot = sum(a*b for a,b in zip(q_vec,vec))
                na = math.sqrt(sum(a*a for a in q_vec))
                nb = math.sqrt(sum(b*b for b in vec))
                score = dot/(na*nb) if na*nb > 0 else 0
                results.append({
                    "id": d.id,
                    "rag_type": row.get("rag_type",""),
                    "title": row.get("title",""),
                    "content": row.get("content",""),
                    "tags": row.get("tags",""),
                    "score": score,
                })
            except Exception:
                continue
        results.sort(key=lambda x: x["score"], reverse=True)
        return {"ok": True, "results": results[:top_k]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ========== CRM: リアルタイム推論パイプライン ==========
class RealtimeInferenceRequest(BaseModel):
    customer_id: str = ""
    trigger_event: str = ""
    customer: dict = {}
    industry: str = "nightlife"

@router.post("/crm_realtime_inference")
def crm_realtime_inference(req: RealtimeInferenceRequest, payload: dict = Depends(verify_token)):
    uid = payload["uid"]
    if not _crm_feat(uid, "diag_crm"):
        raise HTTPException(status_code=403, detail="この機能はAPEX・ULTRAプラン限定です")
    import re as _re8, datetime as _dt8
    from api.core.firestore_client import get_db as _gdb
    c = req.customer
    trigger = req.trigger_event
    customer_id = req.customer_id
    print(f"[CRM_REALTIME] customer_id={repr(customer_id)} trigger={trigger[:30]}", flush=True)
    try:
        line_rate = float(str(c.get("line_reply_rate","0")).replace("%","") or 0)
        cancel_rate = float(str(c.get("cancel_rate","0")).replace("%","") or 0)
        late_night = float(str(c.get("late_night_rate","0")).replace("%","") or 0)
        visit_count = int(str(c.get("visit_count","0")) or 0)
        visit_cycle = int(str(c.get("visit_cycle","14")) or 14)
        spend_total = int(str(c.get("spend_total","0")) or 0)
        avg_spend = spend_total // visit_count if visit_count > 0 else 0
        churn_risk = c.get("churn_risk","低")
        temp = c.get("temp","B")
        psych_type = c.get("psych_type","")
        current_desire = c.get("current_desire","")
    except Exception:
        line_rate = cancel_rate = late_night = 0
        visit_count = avg_spend = 0
        visit_cycle = 14
        churn_risk = "低"; temp = "B"; psych_type = ""; current_desire = ""

    _industry_rt = req.industry or "nightlife"
    _cfg_rt = _get_crm_cfg(_industry_rt)
    _kpi_rt = "、".join(_cfg_rt["kpi"])
    _contact_lbl_rt = _cfg_rt.get("contact_label", "LINE") if isinstance(_cfg_rt, dict) else "LINE"
    _usage_lbl_rt = _cfg_rt.get("visit_label", "来店") if isinstance(_cfg_rt, dict) else "来店"
    _revenue_lbl_rt = _cfg_rt.get("spend_label", "利用金額") if isinstance(_cfg_rt, dict) else "利用金額"
    if _industry_rt == "nightlife":
        _rt_extra = f"深夜返信率: {late_night}% / ドタキャン率: {cancel_rate}%\n"
    elif _industry_rt == "fitness":
        _rt_extra = f"欠席率: {cancel_rate}% / 継続率: {100 - cancel_rate:.0f}%\n"
    elif _industry_rt == "b2b":
        _rt_extra = f"商談進捗: {c.get('consideration_phase','')} / 見積反応: {c.get('sale_response','')}\n"
    elif _industry_rt == "realestate":
        _rt_extra = f"内見反応: {c.get('visit_reaction','')} / 申込状況: {c.get('application_status','')}\n"
    elif _industry_rt == "beauty":
        _rt_extra = f"予約反応: {c.get('booking_response','')} / 口コミ傾向: {c.get('review_tendency','')}\n"
    else:
        _rt_extra = f"ドタキャン率: {cancel_rate}%\n"
    prompt = (
        f"【顧客状態推論】以下のデータを元に状態を更新してください。\n"
        f"《業種》{_industry_rt} / {_cfg_rt['staff_noun']} / KPI:{_kpi_rt} / リスク:{_cfg_rt['risk_focus']}\n"
        f"トリガーイベント: {trigger[:30]}\n"
        f"現在温度: {temp} / 現在欲求: {current_desire[:20]} / 離脱リスク: {churn_risk}\n"
        f"心理タイプ: {psych_type}\n"
        f"{_contact_lbl_rt}反応率: {line_rate}% / " + _rt_extra +
        f"{_usage_lbl_rt}回数: {visit_count}回 / {_usage_lbl_rt}周期: {visit_cycle}日 / 平均{_revenue_lbl_rt}: {avg_spend}円\n"
        "以下JSON1行のみ出力。値は全て10字以内。他出力禁止:\n"
        "{\"t\":\"B\",\"d\":\"承認不足\",\"r\":\"低\",\"why\":\"理由\",\"act\":\"行動\",\"c\":0.8}"
    )
    try:
        raw = call_llm(
            system_prompt=f"あなたは{_industry_rt}業種のリアルタイム顧客状態推論AIです。{_cfg_rt['staff_noun']}の視点で分析し、指定JSON形式のみで出力してください。説明文・コードブロック不要。",
            messages=[{"role":"user","content":prompt}],
            ai_tier="core", max_tokens=1200,
        )
        text = str(raw).strip()
        text = _re8.sub(r"^```json\s*","",text,flags=_re8.MULTILINE).strip()
        text = _re8.sub(r"^```\s*","",text,flags=_re8.MULTILINE).strip()
        text = _re8.sub(r"```\s*$","",text,flags=_re8.MULTILINE).strip()
        s = text.find("{"); e2 = text.rfind("}")
        if s==-1:
            raise ValueError(f"no JSON in: {text[:200]}")
        json_str = text[s:e2+1] if e2!=-1 else text[s:] + "}"
        raw_result = _crm_json.loads(json_str)
        # 短縮キー→標準キーに変換
        result = {
            "updated_temp": raw_result.get("t", raw_result.get("updated_temp", temp)),
            "updated_desire": raw_result.get("d", raw_result.get("updated_desire", current_desire)),
            "updated_churn_risk": raw_result.get("r", raw_result.get("updated_churn_risk", churn_risk)),
            "inference_reason": raw_result.get("why", raw_result.get("inference_reason", "")),
            "immediate_action": raw_result.get("act", raw_result.get("immediate_action", "")),
            "confidence": raw_result.get("c", raw_result.get("confidence", 0.7)),
        }
        now = _dt8.datetime.utcnow().isoformat()
        if customer_id:
            _gdb().collection("users").document(uid).collection("crm_customers").document(customer_id).collection("ai_states").add({
                "trigger_event": trigger,
                "updated_temp": result.get("updated_temp", temp),
                "updated_desire": result.get("updated_desire", current_desire),
                "updated_churn_risk": result.get("updated_churn_risk", churn_risk),
                "inference_reason": result.get("inference_reason",""),
                "immediate_action": result.get("immediate_action",""),
                "confidence": result.get("confidence", 0.7),
                "inferred_at": now,
            })
            _gdb().collection("users").document(uid).collection("crm_customers").document(customer_id).set({
                "temp": result.get("updated_temp", temp),
                "current_desire": result.get("updated_desire", current_desire),
                "churn_risk": result.get("updated_churn_risk", churn_risk),
                "last_inferred_at": now,
            }, merge=True)
        return {"ok": True, "result": result}
    except _crm_json.JSONDecodeError as je:
        print(f"[CRM_JSON_ERR] raw={text[:1000]}", flush=True)
        raise HTTPException(status_code=502, detail=f"JSON解析失敗: {str(je)[:120]}")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"推論失敗: {str(e)}")

@router.get("/crm_ai_states/{customer_id}")
def crm_ai_states(customer_id: str, payload: dict = Depends(verify_token)):
    uid = payload["uid"]
    if not _crm_feat(uid, "diag_crm"):
        raise HTTPException(status_code=403, detail="この機能はAPEX・ULTRAプラン限定です")
    try:
        from api.core.firestore_client import get_db as _gdb
        docs = list(
            _gdb().collection("users").document(uid)
            .collection("crm_customers").document(customer_id)
            .collection("ai_states").stream()
        )
        states = []
        for d in docs:
            row = d.to_dict() or {}
            row["id"] = d.id
            states.append(row)
        states.sort(key=lambda x: x.get("inferred_at",""), reverse=True)
        return {"ok": True, "states": states[:20]}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# ========== CRM: 7日接触なし自動推論 ==========
@router.post("/crm_auto_inference_check")
def crm_auto_inference_check(payload: dict = Depends(verify_token)):
    uid = payload["uid"]
    if not _crm_feat(uid, "diag_crm"):
        raise HTTPException(status_code=403, detail="この機能はAPEX・ULTRAプラン限定です")
    import re as _re9, datetime as _dt9
    from api.core.firestore_client import get_db as _gdb
    try:
        today = _dt9.date.today()
        docs = list(_gdb().collection("users").document(uid).collection("crm_customers").stream())
        triggered = []
        for d in docs:
            c = d.to_dict() or {}
            customer_id = d.id
            name = c.get("name","")
            last_visit = c.get("last_visit","")
            last_inferred = c.get("last_inferred_at","")
            # 最終推論から24時間以上経過しているか確認
            if last_inferred:
                try:
                    li = _dt9.datetime.fromisoformat(last_inferred)
                    hours_since = ((_dt9.datetime.utcnow()) - li).total_seconds() / 3600
                    if hours_since < 24:
                        continue
                except Exception:
                    pass
            # 7日以上接触なしか確認（last_contact_at > last_event_at > last_line_at > last_visit の優先順）
            days_since = 999
            _contact_candidates = [
                c.get("last_contact_at",""),
                c.get("last_event_at",""),
                c.get("last_line_at",""),
                last_visit,
            ]
            for _cdate in _contact_candidates:
                if not _cdate:
                    continue
                try:
                    _cd = _dt9.date.fromisoformat(str(_cdate)[:10])
                    days_since = (today - _cd).days
                    break
                except Exception:
                    continue
            if days_since < 7:
                continue
            # 7日以上接触なし → 自動推論実行
            trigger = f"7日間接触なし（{days_since}日経過）"
            temp = c.get("temp","B")
            churn_risk = c.get("churn_risk","低")
            current_desire = c.get("current_desire","")
            prompt = (
                f"イベント:{trigger} 温度:{temp} 欲求:{current_desire[:120]} 離脱:{churn_risk}\n"
                "以下JSON1行のみ出力。値は全て10字以内。他出力禁止:\n"
                "{\"t\":\"B\",\"d\":\"承認不足\",\"r\":\"低\",\"why\":\"理由\",\"act\":\"行動\",\"c\":0.8}"
            )
            try:
                _industry_ac = c.get('industry') or 'other'
                _cfg_ac = _get_crm_cfg(_industry_ac)
                raw = call_llm(
                    system_prompt=f"あなたは{_industry_ac}業種のリアルタイム顧客状態推論AIです。{_cfg_ac['staff_noun']}の視点で分析し、指定JSON形式のみで出力してください。",
                    messages=[{"role":"user","content":prompt}],
                    ai_tier="core", max_tokens=1200,
                )
                text = str(raw).strip()
                text = _re9.sub(r"^```json\s*","",text,flags=_re9.MULTILINE).strip()
                text = _re9.sub(r"^```\s*","",text,flags=_re9.MULTILINE).strip()
                text = _re9.sub(r"```\s*$","",text,flags=_re9.MULTILINE).strip()
                s = text.find("{"); e2 = text.rfind("}")
                json_str = text[s:e2+1] if s!=-1 and e2!=-1 else text[s:]+"}" if s!=-1 else "{}"
                raw_result = _crm_json.loads(json_str)
                # sanitize
                _allowed_temp = ["S","A","B","C","D"]
                _allowed_risk = ["低","中","高"]
                _upd_temp = raw_result.get("t", temp)
                if _upd_temp not in _allowed_temp:
                    _upd_temp = temp
                _upd_risk = raw_result.get("r", churn_risk)
                if _upd_risk not in _allowed_risk:
                    _upd_risk = churn_risk
                _raw_conf = raw_result.get("c", None)
                if isinstance(_raw_conf, (int, float)) and 0.0 <= float(_raw_conf) <= 1.0:
                    _confidence = float(_raw_conf)
                else:
                    _confidence = None
                result = {
                    "updated_temp": _upd_temp,
                    "updated_desire": raw_result.get("d", current_desire),
                    "updated_churn_risk": _upd_risk,
                    "inference_reason": raw_result.get("why","7日接触なし"),
                    "immediate_action": raw_result.get("act",""),
                    "confidence": _confidence,
                }
                now = _dt9.datetime.utcnow().isoformat()
                _gdb().collection("users").document(uid).collection("crm_customers").document(customer_id).collection("ai_states").add({
                    "trigger_event": trigger,
                    "updated_temp": result["updated_temp"],
                    "updated_desire": result["updated_desire"],
                    "updated_churn_risk": result["updated_churn_risk"],
                    "inference_reason": result["inference_reason"],
                    "immediate_action": result["immediate_action"],
                    "confidence": result["confidence"],
                    "inferred_at": now,
                })
                _gdb().collection("users").document(uid).collection("crm_customers").document(customer_id).set({
                    "temp": result["updated_temp"],
                    "current_desire": result["updated_desire"],
                    "churn_risk": result["updated_churn_risk"],
                    "last_inferred_at": now,
                }, merge=True)
                triggered.append({"name":name,"customer_id":customer_id,"days_since":days_since,"result":result})
            except Exception as e:
                print(f"[crm_auto_inference_check] failed customer_id={customer_id}: {e}", flush=True)
                continue
        return {"ok": True, "triggered": triggered, "count": len(triggered)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/crm_priority_batch")
async def crm_priority_batch(request: Request):
    """Cloud Schedulerから毎分叩かれる。priority_action_timeが現在時刻に一致するユーザーを処理。"""
    import datetime as _bdt
    import os as _bos
    body = await request.json()
    if body.get("admin_password") != _bos.environ.get("ADMIN_PASSWORD",""):
        raise HTTPException(status_code=403, detail="forbidden")
    try:
        from api.core.firestore_client import get_db as _bgdb
        db = _bgdb()
        now_jst = _bdt.datetime.utcnow() + _bdt.timedelta(hours=9)
        now_hm = now_jst.strftime("%H:%M")
        today_str = now_jst.strftime("%Y-%m-%d")

        # CRM機能有効ユーザーを全取得
        users_snap = list(db.collection("users").stream())
        processed = []
        for udoc in users_snap:
            try:
                uid = udoc.id
                udata = udoc.to_dict() or {}
                # feature確認
                features = udata.get("feature_overrides", {})
                plan = udata.get("plan", "starter")
                if plan not in ("apex","ultra_admin","ultra_member") and not features.get("diag_crm"):
                    continue
                # 通知設定確認
                notif_settings = udata.get("notification_settings", {})
                if notif_settings.get("notify_priority_action", True) is False:
                    continue
                # 時刻確認
                action_time = notif_settings.get("priority_action_time", "09:00")
                if action_time != now_hm:
                    continue
                # 既に今日処理済みか確認
                cache_ref = db.collection("users").document(uid).collection("crm_priority_cache").document(today_str)
                cache_snap = cache_ref.get()
                if cache_snap.exists:
                    _cdata = cache_snap.to_dict() or {}
                    _gen_at = _cdata.get("generated_at","")
                    _cached_cn = _cdata.get("priority",{}).get("contact_now",[])
                    _valid_count = len([x for x in _cached_cn if x.get("line_direction") or x.get("line_timing")])
                    try:
                        import datetime as _ttl_dt
                        if _gen_at:
                            _gen_dt = _ttl_dt.datetime.fromisoformat(_gen_at.replace("Z","+00:00"))
                            if _gen_dt.tzinfo is None:
                                _gen_dt = _gen_dt.replace(tzinfo=now_jst.tzinfo)
                            _age_hours = (now_jst - _gen_dt).total_seconds() / 3600
                            if _age_hours < 2 and _valid_count > 0:
                                continue
                    except Exception:
                        pass
                # 顧客データ取得
                docs = list(db.collection("users").document(uid).collection("crm_customers").stream())
                customers = [dict(d.to_dict() or {}, id=d.id) for d in docs]
                if not customers:
                    continue

                # 優先アクション計算（crm_priorityと同一ロジック）
                import datetime as _dt3
                import json as _json3
                today = _bdt.date.today()
                def _safe_float(v, default=0.0):
                    try: return float(str(v).replace("%","").strip() or default)
                    except: return default
                def _safe_int(v, default=0):
                    try: return int(str(v).replace("日","").strip() or default)
                    except: return default
                def _days_since(date_str):
                    if not date_str: return 999
                    try: return (today - _dt3.date.fromisoformat(str(date_str)[:10])).days
                    except: return 999

                contact_now, churn_danger, revenue_expect, leave_alone, mental_danger = [], [], [], [], []
                for c in customers:
                    name = c.get("name","")
                    days_since = _days_since(c.get("last_visit",""))
                    cycle_days = _safe_int(c.get("visit_cycle",0))
                    line_rate = _safe_float(c.get("line_reply_rate",0))
                    late_night = _safe_float(c.get("late_night_rate",0))
                    cancel_rate = _safe_float(c.get("cancel_rate",0))
                    spend_total = _safe_int(c.get("spend_total",0))
                    visit_count = _safe_int(c.get("visit_count",0))
                    avg_spend = spend_total // visit_count if visit_count > 0 else 0
                    obsession = c.get("obsession_tendency","")
                    mental = c.get("mental_instability","")
                    pseudo_love = c.get("pseudo_love_tendency","")
                    claim = c.get("claim_history","")
                    cycle_lag = days_since - cycle_days if cycle_days > 0 else 0
                    churn_score = 0
                    if cycle_days > 0 and cycle_lag > 7: churn_score += 30
                    if cycle_days > 0 and cycle_lag > 14: churn_score += 20
                    if line_rate < 30: churn_score += 25
                    if line_rate < 10: churn_score += 15
                    if cancel_rate > 30: churn_score += 15
                    if days_since > 60: churn_score += 20
                    if c.get("temp") == "C": churn_score += 10
                    if c.get("churn_risk") == "高": churn_score += 10
                    churn_score = min(100, churn_score)
                    dep_score = 0
                    if late_night > 60: dep_score += 25
                    if obsession: dep_score += 25
                    if pseudo_love and "強" in str(pseudo_love): dep_score += 20
                    if mental: dep_score += 15
                    if c.get("temp") == "S": dep_score += 15
                    dep_score = min(100, dep_score)
                    rev_score = 0
                    if avg_spend > 50000: rev_score += 35
                    elif avg_spend > 20000: rev_score += 20
                    if visit_count > 10: rev_score += 20
                    if visit_count > 20: rev_score += 10
                    if line_rate > 70: rev_score += 15
                    if cancel_rate < 10: rev_score += 10
                    if churn_score < 30: rev_score += 10
                    rev_score = min(100, rev_score)
                    revisit_prob = 0
                    if cycle_days > 0 and 0 <= cycle_lag <= 3: revisit_prob += 40
                    if cycle_days > 0 and cycle_lag <= 0: revisit_prob += 20
                    if line_rate > 70: revisit_prob += 20
                    if cancel_rate < 10: revisit_prob += 15
                    if visit_count > 5: revisit_prob += 5
                    revisit_prob = min(100, revisit_prob)
                    vip_prob = 0
                    if avg_spend > 50000: vip_prob += 35
                    if visit_count > 20: vip_prob += 25
                    if line_rate > 80: vip_prob += 20
                    if cancel_rate < 5: vip_prob += 10
                    if churn_score < 20: vip_prob += 10
                    vip_prob = min(100, vip_prob)
                    mental_score = 0
                    if dep_score > 60: mental_score += 40
                    if obsession: mental_score += 25
                    if mental: mental_score += 25
                    if claim: mental_score += 10
                    mental_score = min(100, mental_score)
                    score_entry = {
                        "name": name,
                        "churn_risk_score": churn_score,
                        "revisit_probability": revisit_prob,
                        "revenue_score": rev_score,
                        "vip_probability": vip_prob,
                        "mental_danger_score": mental_score,
                        "line_direction": c.get("ai_line_direction",""),
                        "line_forbidden": c.get("ai_line_forbidden","なし"),
                        "line_timing": c.get("ai_line_timing",""),
                    }
                    _has_action = bool(c.get("ai_line_direction","").strip()) or bool(c.get("ai_line_timing","").strip())
                    if churn_score >= 50:
                        churn_danger.append(score_entry)
                    if _has_action:
                        if cycle_days > 0 and 0 <= cycle_lag <= 5:
                            contact_now.append(score_entry)
                        elif cycle_days > 0 and cycle_lag > 0 and churn_score < 50:
                            contact_now.append(score_entry)
                        elif cycle_days == 0 and days_since <= 30:
                            contact_now.append(score_entry)
                    if rev_score >= 50 and churn_score < 50:
                        revenue_expect.append(score_entry)
                    if cycle_days > 0 and days_since > cycle_days * 2 and churn_score < 50:
                        leave_alone.append(score_entry)
                    elif cycle_days == 0 and days_since > 60 and churn_score < 50:
                        leave_alone.append(score_entry)
                    if mental_score >= 50:
                        mental_danger.append(score_entry)

                churn_danger.sort(key=lambda x: x["churn_risk_score"], reverse=True)
                contact_now.sort(key=lambda x: x["revisit_probability"], reverse=True)
                revenue_expect.sort(key=lambda x: x["revenue_score"], reverse=True)
                mental_danger.sort(key=lambda x: x["mental_danger_score"], reverse=True)

                priority_data = {
                    "contact_now": contact_now[:5],
                    "churn_danger": churn_danger[:5],
                    "revenue_expect": revenue_expect[:5],
                    "leave_alone": leave_alone[:5],
                    "mental_danger": mental_danger[:5],
                }

                # Firestoreにキャッシュ保存（actionゼロなら保存しない）
                _valid_cn = [x for x in contact_now[:5] if x.get("line_direction") or x.get("line_timing")]
                if _valid_cn or churn_danger:
                    cache_ref.set({
                        "priority": priority_data,
                        "generated_at": now_jst.isoformat(),
                        "generated_action_count": len(_valid_cn),
                    })

                # 通知生成
                pa_total = len(contact_now[:5]) + len(churn_danger[:5])
                if pa_total > 0:
                    import uuid as _uuid
                    pa_parts = []
                    if contact_now: pa_parts.append(f"今日連絡すべき顧客 {len(contact_now[:5])}名")
                    if churn_danger: pa_parts.append(f"離脱危険顧客 {len(churn_danger[:5])}名")
                    nid = _uuid.uuid4().hex
                    db.collection("notifications").document(uid).collection("items").document(nid).set({
                        "notif_id": nid,
                        "uid": uid,
                        "type": "priority_action",
                        "title": "🚨 今日の優先アクションがあります",
                        "body": "、".join(pa_parts),
                        "link_tab": "crm",
                        "read": False,
                        "created_at": now_jst.isoformat(),
                    })
                processed.append(uid)
            except Exception:
                continue
        return {"ok": True, "processed": len(processed), "time": now_hm}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
