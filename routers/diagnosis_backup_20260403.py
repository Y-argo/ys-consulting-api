# api/routers/diagnosis.py
import datetime
_JST = datetime.timezone(datetime.timedelta(hours=9))
def _now_jst(): return datetime.datetime.now(_JST)
import uuid
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Body, Form
from pydantic import BaseModel
from typing import List, Optional
from google.cloud import firestore as fs

from api.routers.auth import verify_token
from api.core.firestore_client import get_db, DEFAULT_TENANT
from api.core.llm_client import call_llm

router = APIRouter(prefix="/api/diagnosis", tags=["diagnosis"])

def _load_rank_config(tenant_id: str) -> dict:
    try:
        db = get_db()
        doc = db.collection("tenant_settings").document(tenant_id).get()
        if doc.exists:
            d = doc.to_dict() or {}
            return {
                "rank_1_name": d.get("rank_1_name", "D"),
                "rank_2_name": d.get("rank_2_name", "C"),
                "rank_3_name": d.get("rank_3_name", "B"),
                "rank_4_name": d.get("rank_4_name", "A"),
            }
    except Exception:
        pass
    return {"rank_1_name": "D", "rank_2_name": "C", "rank_3_name": "B", "rank_4_name": "A"}

def _load_chat_history_across_sessions(uid: str, tenant_id: str, limit: int = 30) -> list:
    db = get_db()
    try:
        sessions_raw = list(
            db.collection("chat_sessions")
            .where("uid", "==", uid)
            .limit(50)
            .stream()
        )
        sessions_raw.sort(key=lambda s: str((s.to_dict() or {}).get("updated_at", "")), reverse=True)
        sessions = sessions_raw[:10]
        msgs = []
        for s in sessions:
            ref = db.collection("chat_sessions").document(s.id).collection("messages")
            for m in ref.order_by("ts").limit_to_last(10).get():
                d = m.to_dict() or {}
                msgs.append({"role": d.get("role", "user"), "content": d.get("content", "")})
            if len(msgs) >= limit:
                break
        return msgs[-limit:]
    except Exception:
        return []

def _load_score_config(tenant_id: str) -> dict:
    db = get_db()
    DEFAULT = {
        "struct_words": "構造,資本,市場,制度,最適,期待値,確率,アーキテクチャ,設計,フレームワーク",
        "strategy_words": "戦略,施策,優先,差別化,競合,ポジショニング,KPI,ROI,目標",
        "exec_words": "実行,手順,タスク,スケジュール,チェック,改善,運用,効率",
        "emotion_words": "不安,ムカつく,なぜ俺,怖い,どうせ,無理,クソ,無能,イライラ,最悪",
    }
    try:
        for tid in [tenant_id, "default"]:
            doc = db.collection("system_settings").document(f"score_config_{tid}").get()
            if doc.exists:
                d = doc.to_dict() or {}
                return {**DEFAULT, **d}
    except Exception:
        pass
    return DEFAULT

def _generate_diagnosis(uid: str, tenant_id: str, n_chats: int = 30) -> str:
    rank_cfg = _load_rank_config(tenant_id)
    msgs = _load_chat_history_across_sessions(uid, tenant_id, limit=n_chats)
    chat_text = "\n".join([f"{m['role']}: {m.get('content','')}" for m in msgs if m.get("content")])
    if not chat_text.strip():
        return ""
    timestamp_str = _now_jst().strftime("%Y-%m-%d %H:%M")
    dr1 = rank_cfg["rank_1_name"]
    dr2 = rank_cfg["rank_2_name"]
    dr3 = rank_cfg["rank_3_name"]
    dr4 = rank_cfg["rank_4_name"]
    prompt = f"""以下はユーザー「{uid}」の直近チャット履歴（{n_chats}件）です。

{chat_text}

---
【ランク体系（低→高）】: {dr1} → {dr2} → {dr3} → {dr4}

上記の履歴を踏まえ、以下のフォーマットを厳守して「現状課題診断レポート」を生成してください。

# 現状課題診断レポート
生成日時: {timestamp_str}
対象ユーザー: {uid}
解析範囲: 直近チャット {n_chats} 件
---
## 総合評価
評価ランク: [S/A/B/C/D]
現状状態: [1〜3行で要約]
優先改善度: [低/中/高/緊急]
---
## 主要課題（最大3件）
### 課題1: [課題名]
影響度: [高/中/低]
原因: [構造的原因]
推奨行動: [具体行動]
---
### 課題2: [課題名]
影響度: [高/中/低]
原因: [構造的原因]
推奨行動: [具体行動]
---
### 課題3: [課題名]
影響度: [高/中/低]
原因: [構造的原因]
推奨行動: [具体行動]
---
## 強み・弱点
強み: [維持すべき行動]
弱点: [改善が必要な行動]
"""
    try:
        return call_llm(
            system_prompt="あなたは戦略コンサルタントです。与えられたチャット履歴を分析し、構造的な課題診断レポートを生成してください。",
            messages=[{"role": "user", "content": prompt}],
            ai_tier="core",
            max_tokens=4096,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"診断生成エラー: {e}")

def _save_diagnosis(uid: str, tenant_id: str, report_md: str, rank: str = None, n_chats: int = None) -> str:
    db = get_db()
    doc_id = str(uuid.uuid4())
    data = {
        "uid": uid,
        "tenant_id": tenant_id,
        "report_md": report_md,
        "created_at": fs.SERVER_TIMESTAMP,
    }
    if rank: data["rank"] = rank
    if n_chats: data["n_chats"] = n_chats
    db.collection("user_diagnoses").document(doc_id).set(data)
    return doc_id

def _load_diagnoses(uid: str, tenant_id: str, limit: int = 5) -> list:
    db = get_db()
    try:
        docs = list(
            db.collection("user_diagnoses")
            .where("uid", "==", uid)
            .limit(50)
            .stream()
        )
        result = []
        for d in docs:
            data = d.to_dict() or {}
            result.append({
                "doc_id": d.id,
                "report_md": data.get("report_md", ""),
                "created_at": str(data.get("created_at", "")),
                "rank": data.get("rank"),
                "n_chats": data.get("n_chats"),
            })
        result.sort(key=lambda x: x["created_at"], reverse=True)
        return result[:limit]
    except Exception:
        return []

class DiagnosisRequest(BaseModel):
    n_chats: int = 30

@router.post("/generate")
def generate_diagnosis(req: DiagnosisRequest, payload: dict = Depends(verify_token)):
    from api.core.features import is_feature_enabled
    uid = payload["uid"]; tenant_id = payload.get("tenant_id","default")
    if not is_feature_enabled(uid, "current_issue_diagnosis"):
        raise HTTPException(status_code=403, detail="この機能は利用できません")
    uid = payload["uid"]
    tenant_id = payload.get("tenant_id", DEFAULT_TENANT)
    report_md = _generate_diagnosis(uid, tenant_id, req.n_chats)
    if not report_md:
        raise HTTPException(status_code=400, detail="チャット履歴が不足しています")
    # ランク取得
    try:
        from api.routers.user_stats import _load_rank_config, _score_to_rank
        from api.core.firestore_client import get_db as _gdb2
        _snap = _gdb2().collection("users").document(uid).get()
        _score = int((_snap.to_dict() or {}).get("level_score", 0)) if _snap.exists else 0
        _cfg = _load_rank_config(tenant_id)
        _rank = _score_to_rank(_score, _cfg)
    except Exception:
        _rank = None
    doc_id = _save_diagnosis(uid, tenant_id, report_md, rank=_rank, n_chats=req.n_chats)
    return {"doc_id": doc_id, "report_md": report_md}

@router.get("/list")
def list_diagnoses(payload: dict = Depends(verify_token)):
    uid = payload["uid"]
    tenant_id = payload.get("tenant_id", DEFAULT_TENANT)
    return {"diagnoses": _load_diagnoses(uid, tenant_id)}

@router.get("/thought_map")
def get_thought_map(payload: dict = Depends(verify_token)):
    uid = payload["uid"]
    tenant_id = payload.get("tenant_id", DEFAULT_TENANT)
    db = get_db()
    try:
        all_sessions = list(db.collection("chat_sessions").limit(200).stream())
        sessions_raw = [s for s in all_sessions if f"__{uid}__" in s.id]
        sessions_raw.sort(key=lambda s: str((s.to_dict() or {}).get("updated_at","")), reverse=True)

        # トピック分類キーワード
        TOPICS = {
            "戦略・競合": ["戦略","競合","差別化","ポジション","市場","シェア","ブランド"],
            "集客・SNS": ["集客","SNS","Instagram","Twitter","広告","フォロワー","投稿"],
            "売上・財務": ["売上","収益","利益","資金","コスト","価格","単価"],
            "組織・人材": ["採用","チーム","組織","スタッフ","教育","マネジメント"],
            "投資・株": ["投資","株","銘柄","シグナル","相場","資産","ポートフォリオ"],
            "診断・分析": ["診断","分析","レポート","スコア","評価","課題"],
            "指名・接客": ["指名","接客","お客様","キャスト","リピート","コミュニケーション"],
        }

        def classify(text):
            for topic, keywords in TOPICS.items():
                if any(k in text for k in keywords):
                    return topic
            return "その他"

        # ノード収集
        raw_nodes = []
        node_set = set()
        for s in sessions_raw[:15]:
            try:
                msgs = list(db.collection("chat_sessions").document(s.id).collection("messages").limit(40).stream())
                msgs.sort(key=lambda m: str((m.to_dict() or {}).get("ts","")))
                for m in msgs:
                    d = m.to_dict() or {}
                    if d.get("role") == "user":
                        content = (d.get("content","") or "")[:35]
                        if content and content not in node_set and len(content) > 3:
                            node_set.add(content)
                            topic = classify(content)
                            raw_nodes.append({"id": content, "label": content, "group": topic})
            except Exception:
                continue

        raw_nodes = raw_nodes[:30]

        # トピック間のエッジ（同トピック内で繋ぐ）
        edges = []
        topic_last = {}
        for n in raw_nodes:
            g = n["group"]
            if g in topic_last:
                edges.append({"from": topic_last[g], "to": n["id"], "topic": g})
            topic_last[g] = n["id"]

        # トピック中心ノードを追加
        topics_used = list(set(n["group"] for n in raw_nodes))
        center_nodes = [{"id": f"__topic_{t}__", "label": t, "group": t, "is_center": True} for t in topics_used]

        # 各ノードをトピック中心に繋ぐ
        center_edges = [{"from": f"__topic_{n['group']}__", "to": n["id"], "topic": n["group"]} for n in raw_nodes]

        all_nodes = center_nodes + raw_nodes
        all_edges = center_edges

        return {"nodes": all_nodes, "edges": all_edges, "topics": topics_used}
    except Exception as e:
        return {"nodes": [], "edges": [], "error": str(e)}

from fastapi import Body as _Body

class ConsultRequest(BaseModel):
    analysis_type: str
    input_text: str
    supplement: str = ""
    options: str = ""
    strategy: str = ""
    policy: str = ""

@router.post("/consult")
def run_consult(req: ConsultRequest, payload: dict = Depends(verify_token)):
    from api.core.llm_client import call_llm
    uid = payload["uid"]
    tenant_id = payload.get("tenant_id", "default")
    db = get_db()

    frameworks = []
    try:
        docs = list(db.collection("tenants").document(tenant_id).collection("consulting_frameworks").where("active","==",True).stream())
        frameworks = [d.to_dict().get("name","") for d in docs if d.to_dict().get("name")]
    except Exception:
        pass

    fw_str = "、".join(frameworks[:5]) if frameworks else "MECE・SWOT・3C・ロジックツリー・Issue Tree"

    if req.analysis_type == "structure":
        prompt = f"""以下の入力を構造診断せよ。JSONで返せ。
入力: {req.input_text}
補足: {req.supplement}
フレームワーク候補: {fw_str}
出力形式: {{"summary":"全体要約","structure_layers":[{{"layer":"層名","content":"内容","strength":0.8}}],"key_bottleneck":"主要ボトルネック","recommended_framework":"推奨フレームワーク","next_actions":["アクション1"]}}"""

    elif req.analysis_type == "issue":
        prompt = f"""以下の状況から課題仮説を生成せよ。JSONで返せ。
入力: {req.input_text}
フレームワーク: {fw_str}
出力形式: {{"hypotheses":[{{"hypothesis":"仮説","priority":"high/mid/low","evidence":"根拠","verification":"検証方法"}}],"root_cause":"根本原因","quick_wins":["即効策"]}}"""

    elif req.analysis_type == "comparison":
        prompt = f"""以下の選択肢を多軸比較せよ。JSONで返せ。
選択肢: {req.options or req.input_text}
評価軸: コスト・リスク・効果・実現性・速度
出力形式: {{"options":[{{"name":"選択肢名","scores":{{"cost":80,"risk":60,"effect":90,"feasibility":70,"speed":75}},"summary":"評価コメント"}}],"recommendation":"推奨選択肢","rationale":"理由"}}"""

    elif req.analysis_type == "contradiction":
        prompt = f"""以下の戦略と方針の矛盾を検知せよ。JSONで返せ。
戦略: {req.strategy or req.input_text}
方針: {req.policy or req.supplement}
出力形式: {{"contradictions":[{{"point":"矛盾点","severity":"high/mid/low","resolution":"解決策"}}],"consistency_score":70,"overall_assessment":"総合評価"}}"""

    elif req.analysis_type == "execution":
        prompt = f"""以下の目標に対する実行計画を生成せよ。JSONで返せ。
目標・背景: {req.input_text}
フレームワーク: {fw_str}
出力形式: {{"phases":[{{"phase":"フェーズ名","duration":"期間","actions":["アクション"],"kpi":"KPI","risks":["リスク"]}}],"critical_path":"クリティカルパス","success_criteria":["成功条件"]}}"""
    else:
        return {"ok": False, "error": f"不明なanalysis_type: {req.analysis_type}"}

    try:
        import re, json as _json
        res = call_llm(
            system_prompt="戦略コンサルタント。JSONのみ出力。余計なテキスト禁止。",
            messages=[{"role":"user","content":prompt}],
            ai_tier="core", max_tokens=2048
        )
        m = re.search(r"\{.*\}", res, re.DOTALL)
        result = _json.loads(m.group(0)) if m else {"raw": res}

        db.collection("tenants").document(tenant_id).collection("consulting_analyses").add({
            "uid": uid, "tenant_id": tenant_id,
            "analysis_type": req.analysis_type,
            "input_text": req.input_text[:500],
            "result": result,
            "created_at": _now_jst().isoformat(),
        })
        return {"ok": True, "result": result, "analysis_type": req.analysis_type}
    except Exception as e:
        return {"ok": False, "error": str(e)}

@router.get("/consult/history")
def get_consult_history(analysis_type: str = "", payload: dict = Depends(verify_token)):
    uid = payload["uid"]
    tenant_id = payload.get("tenant_id", "default")
    db = get_db()
    try:
        docs = list(
            db.collection("tenants").document(tenant_id)
            .collection("consulting_analyses")
            .where("uid","==",uid)
            .limit(20)
            .stream()
        )
        analyses = [d.to_dict() for d in docs]
        if analysis_type:
            analyses = [a for a in analyses if a.get("analysis_type") == analysis_type]
        analyses.sort(key=lambda x: x.get("created_at",""), reverse=True)
        return {"analyses": analyses}
    except Exception:
        return {"analyses": []}

@router.get("/frameworks")
def get_frameworks(payload: dict = Depends(verify_token)):
    tenant_id = payload.get("tenant_id", "default")
    db = get_db()
    try:
        docs = list(db.collection("tenants").document(tenant_id).collection("consulting_frameworks").stream())
        fw = [d.to_dict() for d in docs]
        if not fw:
            fw = [
                {"name":"MECE","description":"ダブりなく、漏れなく","active":True},
                {"name":"SWOT","description":"強み・弱み・機会・脅威","active":True},
                {"name":"3C","description":"顧客・競合・自社","active":True},
                {"name":"ロジックツリー","description":"問題を論理的に分解","active":True},
                {"name":"Issue Tree","description":"課題を階層的に整理","active":True},
            ]
        return {"frameworks": fw}
    except Exception:
        return {"frameworks": []}

@router.post("/weekly_report")
def generate_weekly_report(body: dict = {}, payload: dict = Depends(verify_token)):
    from api.core.llm_client import call_llm_pro as _clp2
    uid = payload["uid"]
    tenant_id = payload.get("tenant_id", "default")
    db = get_db()
    n_chats = int((body or {}).get("n_chats", 30))
    try:
        history = _load_chat_history_across_sessions(uid, tenant_id, n_chats)
        if not history:
            return {"ok": False, "error": "チャット履歴がありません"}
        chat_text = "\n".join([f"{m['role']}: {m.get('content','')}" for m in history if m.get("content")])[:4000]
        ts = _now_jst().strftime("%Y-%m-%d %H:%M")
        prompt = f"""以下はユーザーのチャット履歴です。週次戦術レポートを生成してください。

{chat_text}

# 週次戦術レポート
生成日時: {ts}
---
## 今週の主要相談テーマ
[最も多かった相談テーマ上位3件]
---
## 意思決定パターン分析
[どのような意思決定が多かったか]
---
## 成長トレンド
[前週比での変化・改善点]
---
## 来週の優先アクション
[最重要アクション3件]
---
## 戦略的提言
[中長期視点での提言]
"""
        report = _clp2(
            system_prompt="あなたは戦略コンサルタント。週次レポートを指定フォーマット厳守で生成せよ。",
            messages=[{"role":"user","content":prompt}],
            max_tokens=3000,
        )
        db.collection("weekly_reports").add({
            "uid": uid, "tenant_id": tenant_id,
            "report_md": report,
            "created_at": _now_jst().isoformat(),
        })
        return {"ok": True, "report_md": report}
    except Exception as e:
        return {"ok": False, "error": str(e)}

@router.get("/weekly_report/list")
def list_weekly_reports(payload: dict = Depends(verify_token)):
    uid = payload["uid"]
    tenant_id = payload.get("tenant_id", "default")
    db = get_db()
    try:
        docs = list(
            db.collection("weekly_reports")
            .where("uid","==",uid)
            .where("tenant_id","==",tenant_id)
            .limit(10)
            .stream()
        )
        reports = [d.to_dict() for d in docs]
        reports.sort(key=lambda x: x.get("created_at",""), reverse=True)
        return {"reports": reports}
    except Exception:
        return {"reports": []}


@router.post("/file_diagnosis")
async def file_diagnosis(
    file: UploadFile = File(...),
    answer_context: str = Form(""),
    payload: dict = Depends(verify_token)
):
    """ファイル全タブ横断診断→構造診断・課題仮説・実行計画を一括生成"""
    from api.core.llm_client import call_llm as _cllm
    import io, re as _re, json as _json

    filename = file.filename or "file"
    content = await file.read()
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    # 全シート抽出
    sheets_text = {}
    try:
        if ext in ("xlsx", "xls"):
            import pandas as pd
            xf = pd.ExcelFile(io.BytesIO(content))
            for sheet in xf.sheet_names:
                try:
                    df_raw = xf.parse(sheet, header=None).dropna(how="all").dropna(axis=1, how="all")
                    sheets_text[sheet] = df_raw.to_csv(index=False, header=False)[:4000] if not df_raw.empty else f"({sheet}:データなし)"
                except Exception as _e:
                    sheets_text[sheet] = f"({sheet}:読み込みエラー:{_e})"
        elif ext == "ods":
            import pandas as pd
            xf = pd.ExcelFile(io.BytesIO(content), engine="odf")
            for sheet in xf.sheet_names:
                try:
                    df_raw = xf.parse(sheet, header=None).dropna(how="all").dropna(axis=1, how="all")
                    sheets_text[sheet] = df_raw.to_csv(index=False, header=False)[:4000] if not df_raw.empty else f"({sheet}:データなし)"
                except Exception as _e:
                    sheets_text[sheet] = f"({sheet}:読み込みエラー:{_e})"
        elif ext == "csv":
            import pandas as pd
            df = pd.read_csv(io.BytesIO(content))
            sheets_text["Sheet1"] = df.to_csv(index=False)[:6000]
        elif ext == "pdf":
            try:
                import pypdf
                reader = pypdf.PdfReader(io.BytesIO(content))
                text = "\n".join(p.extract_text() or "" for p in reader.pages)
                sheets_text["PDF"] = text[:6000]
            except Exception:
                sheets_text["PDF"] = content.decode("utf-8", errors="ignore")[:6000]
        else:
            sheets_text["TEXT"] = content.decode("utf-8", errors="ignore")[:6000]
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"ファイル読み込みエラー: {e}")

    if not sheets_text:
        raise HTTPException(status_code=400, detail="ファイルの内容を読み取れませんでした")

    # 数式を読み込む（xlsx/xls/ods対応）
    formula_summary = ""
    if ext in ("xlsx", "xls"):
        try:
            import openpyxl as _opxl
            wb = _opxl.load_workbook(io.BytesIO(content), data_only=False)
            for sheet_name in wb.sheetnames[:6]:
                ws = wb[sheet_name]
                formulas = []
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                            formulas.append(f"{cell.coordinate}: {cell.value}")
                if formulas:
                    formula_summary += f"\n【{sheet_name}シートの数式（最初の10件）】\n" + "\n".join(formulas[:10])
        except Exception as _fe:
            formula_summary = f"（xlsx数式読み込みエラー: {_fe}）"
    elif ext == "ods":
        try:
            from odf.opendocument import load as _odf_load
            from odf.table import Table, TableRow, TableCell
            from odf.text import P
            doc = _odf_load(io.BytesIO(content))
            for sheet in doc.spreadsheet.getElementsByType(Table)[:6]:
                sheet_name = sheet.getAttribute("name")
                formulas = []
                for row in sheet.getElementsByType(TableRow):
                    for cell in row.getElementsByType(TableCell):
                        formula = cell.getAttribute("formula")
                        if formula:
                            formulas.append(f"{sheet_name}: {formula}")
                if formulas:
                    formula_summary += f"\n【{sheet_name}シートの数式（最初の10件）】\n" + "\n".join(formulas[:10])
        except Exception as _fe:
            formula_summary = f"（ods数式読み込みエラー: {_fe}）"


    # 全シートを結合
    combined = ""
    for sheet, text in sheets_text.items():
        combined += f"\n\n【シート: {sheet}】\n{text}"
    combined = combined[:12000]

    # ===== ③ Python先行数値分析 =====
    import numpy as _np
    import re as _re2, json as _json2

    numeric_analysis = {}
    for sheet, text in sheets_text.items():
        try:
            import pandas as _pd, io as _io
            lines = text.strip().split("\n")
            if len(lines) < 2:
                continue
            df = _pd.read_csv(_io.StringIO(text))

            # Unnamed列・変化率計算不要列を除外
            valid_cols = [c for c in df.columns if "Unnamed" not in str(c)]
            df = df[valid_cols]

            sheet_stats = {"行数": len(df), "列数": len(df.columns), "有効列": valid_cols}

            # 数値列の基本統計（意味のある列のみ）
            num_cols = [c for c in df.select_dtypes(include=["number"]).columns.tolist()
                       if "Unnamed" not in str(c)]

            for col in num_cols[:15]:
                s = df[col].dropna()
                if len(s) > 0 and s.sum() > 0:  # 全0列はスキップ
                    sheet_stats[col] = {
                        "合計": round(float(s.sum()), 2),
                        "平均": round(float(s.mean()), 2),
                        "最大": round(float(s.max()), 2),
                        "最小": round(float(s.min()), 2),
                        "データ数": int(s.count()),
                    }

            # 上位・下位パフォーマー検出（異常値ではなく業界文脈で解釈）
            performers = []
            for col in num_cols[:8]:
                s = df[col].dropna()
                if len(s) >= 4 and s.sum() > 0:
                    top = s.nlargest(3)
                    if top.iloc[0] > s.mean() * 2:
                        performers.append(f"{col}: 上位集中={top.values.tolist()}（平均{round(float(s.mean()),1)}の{round(top.iloc[0]/s.mean(),1)}倍）")
            if performers:
                sheet_stats["上位集中検出"] = performers

            numeric_analysis[sheet] = sheet_stats
        except Exception as _e:
            numeric_analysis[sheet] = {"エラー": str(_e)}

    numeric_summary = _json2.dumps(numeric_analysis, ensure_ascii=False, indent=2)

    # ===== ② Chain of Thought: 4段階順次分析 =====
    # 確認済み内容から業種を抽出してシステムプロンプトに動的反映
    industry_hint = ""
    if answer_context:
        industry_hint = f"\n【確認済みの業種・業界背景・用語定義（診断に必ず反映せよ）】\n{answer_context[:3000]}"

    system = f"""あなたは超一流の経営コンサルタントかつデータ分析の専門家である。{industry_hint}

以下のルールを必ず守れ：
- 確認済みの業種・業界の文脈で全ての数値・用語を解釈せよ
- 提供された数値分析結果を必ず引用して根拠とせよ
- 「具体的数値は集計が必要」「仮定」等の逃げ回答は絶対禁止
- 推測・一般論ではなくデータから読み取れる事実のみを述べよ
- KPIは実データから算出した根拠ある数値のみ使え。データにない数字を作るな
- 業界の慣習・用語・ビジネスモデルを踏まえた専門的な解釈をせよ
- IQR異常値検出はデータ構造上の偏りである場合が多い。業界文脈で正常か異常か判断せよ
- 「上位集中」として検出された数値はトップパフォーマーの正常な実績である。異常値として扱うな
- NET.FREEシートの「R」列はランキング順位であり、リピート数ではない
- グレード表はR4〜R8の年度別の昇格・降格・ゴールド記録であり、読み込みエラーでも「データなし」でもない
- KPIは実データから算出した数値のみ使え。「現状データなし」と書く場合は実データから推計せよ"""

    # answer_contextからユーザー回答のみ抽出
    user_answers = ""
    if answer_context:
        for _line in answer_context.split("\n"):
            if _line.startswith("ユーザー:") or (_line.startswith("Q") and ":" in _line):
                user_answers += _line + "\n"

    context_str = f"""【最重要：事前確認済みの用語・業界定義（必ず全て診断に反映せよ）】
{user_answers if user_answers else answer_context[:2000]}

【Excelの数式（条件・計算ロジック自動読み取り）】
{formula_summary if formula_summary else "（数式なし/odsファイル）"}

【Python数値分析結果】
{numeric_summary}

【生データ（全シート）】
{combined[:8000]}"""

    # Step1: 現状把握
    step1 = _cllm(
        system_prompt=system,
        messages=[{"role":"user","content":f"""{context_str}

【指示】上記データの現状を把握せよ。
- 数値分析結果を必ず具体的数値で引用せよ
- 業界（確認済みの業種）の文脈で解釈せよ
- キャスト名・シート名・具体的数値を使って述べよ
- 「〜の可能性」「集計が必要」等の逃げ表現禁止
- 箇条書きで簡潔に・数値必須で300字以内"""}],
        ai_tier="core", max_tokens=1000
    )

    # Step2: 構造診断
    step2 = _cllm(
        system_prompt=system,
        messages=[
            {"role":"user","content":f"{context_str}\n\n現状把握結果：\n{step1}"},
            {"role":"assistant","content":"現状把握完了。"},
            {"role":"user","content":"""データ構造を診断せよ。
- 各シートの列構成・キャスト配置・前期後期の構造を具体的に説明せよ
- シート間の関係性（ランキング←各シート集計・グレード表←昇格条件）を明示せよ
- 実際に検出された異常値・欠損・入力ミスを数値で指摘せよ
- 改善提案を1つずつ具体的に述べよ"""}
        ],
        ai_tier="core", max_tokens=1200
    )

    # Step3: 課題仮説
    step3 = _cllm(
        system_prompt=system,
        messages=[
            {"role":"user","content":f"{context_str}\n\n現状把握：\n{step1}\n\n構造診断：\n{step2}"},
            {"role":"assistant","content":"現状把握・構造診断完了。"},
            {"role":"user","content":"""業界・店舗運営の観点から課題仮説を3〜5個生成せよ。
各課題は以下の形式で出力せよ：
## 仮説N: [タイトル]
**根拠**: [数値を必ず引用]
**影響**: [具体的な売上・人材・運営への影響]
**優先度**: 高/中/低
**推奨アクション**: [即実行できる具体的施策]

「データ品質が低い」「手入力が問題」等の一般論のみの仮説は禁止。業界特有の課題に踏み込め。"""}
        ],
        ai_tier="core", max_tokens=1500
    )

    # Step4: 実行計画
    step4 = _cllm(
        system_prompt=system,
        messages=[
            {"role":"user","content":f"{context_str}\n\n現状把握：\n{step1}\n\n構造診断：\n{step2}\n\n課題仮説：\n{step3}"},
            {"role":"assistant","content":"現状把握・構造診断・課題仮説完了。"},
            {"role":"user","content":"""上記の分析全体を踏まえ、実行計画を優先度順に3〜5件生成せよ。

各計画を以下の形式で：
## アクションN: [タイトル]
**内容**: [具体的なアクション]
**期限**: [X日以内/X週間以内]
**担当**: [誰が]
**KPI**: [数値目標・現状→目標値]
**期待効果**: [売上・人材・効率への具体的効果]

数値目標は現状データから算出した根拠ある数字を使え。"""}
        ],
        ai_tier="core", max_tokens=1500
    )

    # key_metrics・risks（最終統合）
    step5_prompt = f"""現状把握：{step1}\n構造診断：{step2}\n課題仮説：{step3}\n実行計画：{step4}\n\n以上の分析から、JSONのみで返せ：\n{{"key_metrics":"注目すべき重要指標・数値（箇条書き5件以内・数値必須）","risks":"見逃せないリスク・警告事項（箇条書き3件以内）"}}"""
    step5_raw = _cllm(
        system_prompt=system,
        messages=[{"role":"user","content":step5_prompt}],
        ai_tier="core", max_tokens=512
    )
    try:
        m5 = _re2.search(r'\{.*\}', step5_raw, _re2.DOTALL)
        step5 = _json2.loads(m5.group(0)) if m5 else {}
    except Exception:
        step5 = {}

    return {
        "ok": True,
        "filename": filename,
        "sheets": list(sheets_text.keys()),
        "overview": step1,
        "structure": step2,
        "issues": step3,
        "action_plan": step4,
        "key_metrics": step5.get("key_metrics", ""),
        "risks": step5.get("risks", ""),
        "numeric_analysis": numeric_analysis,
    }


@router.post("/file_followup")
async def file_followup(body: dict = Body(...), payload: dict = Depends(verify_token)):
    """ファイル診断結果への追加質問・深掘り分析"""
    from api.core.llm_client import call_llm as _cllm
    question = body.get("question", "").strip()
    context = body.get("context", "")
    filename = body.get("filename", "")
    if not question:
        raise HTTPException(status_code=400, detail="質問が必要です")

    system = """あなたは超一流の経営コンサルタントかつデータ分析の専門家である。
以下のルールを必ず守れ：

【最重要ルール】
1. 質問に業界特有の専門用語・略語・固有名詞が含まれており、その意味がデータから判断できない場合は、
   回答する前に必ず「〇〇とはどういう意味ですか？」と質問し、確認してから回答せよ。
2. わかったふりをして回答することは絶対禁止。不明な点は必ず確認せよ。
3. データに基づいた具体的な回答のみ出力せよ。拒否・曖昧回答・一般論のみは禁止。
4. 数値は必ず引用し、比較・変化率・傾向を明示せよ。"""

    prompt = f"""ファイル「{filename}」の診断結果：
{context[:6000]}

ユーザーの追加質問：
{question}

【指示】
- 質問に不明な専門用語・略語・業界用語が含まれる場合は、まず「〇〇とはどういう意味ですか？」と確認せよ
- 意味が明確な場合は、診断結果のデータを根拠に具体的・詳細に回答せよ
- 数値は必ず引用すること"""

    try:
        answer = _cllm(
            system_prompt=system,
            messages=[{"role":"user","content":prompt}],
            ai_tier="core", max_tokens=2048
        )
        return {"ok": True, "answer": answer}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/file_diagnosis_check")
async def file_diagnosis_check(
    file: UploadFile = File(...),
    payload: dict = Depends(verify_token)
):
    """ファイルをスキャンして不明な専門用語・業界用語があれば質問を返す"""
    from api.core.llm_client import call_llm as _cllm
    import io, re as _re, json as _json

    filename = file.filename or "file"
    content = await file.read()
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    sheets_text = {}
    try:
        if ext in ("xlsx", "xls"):
            import pandas as pd
            xf = pd.ExcelFile(io.BytesIO(content))
            for sheet in xf.sheet_names:
                df = xf.parse(sheet)
                sheets_text[sheet] = df.to_csv(index=False)[:2000]
        elif ext == "ods":
            import pandas as pd
            xf = pd.ExcelFile(io.BytesIO(content), engine="odf")
            for sheet in xf.sheet_names:
                df = xf.parse(sheet)
                sheets_text[sheet] = df.to_csv(index=False)[:2000]
        elif ext == "csv":
            import pandas as pd
            df = pd.read_csv(io.BytesIO(content))
            sheets_text["Sheet1"] = df.to_csv(index=False)[:4000]
        else:
            sheets_text["TEXT"] = content.decode("utf-8", errors="ignore")[:4000]
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"ファイル読み込みエラー: {e}")

    combined = ""
    for sheet, text in sheets_text.items():
        combined += f"\n\n【シート: {sheet}】\n{text}"
    combined = combined[:6000]

    # Unnamed列を自動解釈してAIに渡す
    unnamed_note = ""
    try:
        import pandas as _pd2, io as _io2
        for sheet, text in sheets_text.items():
            df2 = _pd2.read_csv(_io2.StringIO(text))
            unnamed_cols = [c for c in df2.columns if "Unnamed" in str(c)]
            if unnamed_cols:
                unnamed_note += f"\n※{sheet}シートの'Unnamed:数字'列はExcelの結合セル・空白ヘッダーが自動変換されたものです。スタッフごとのサブ項目列（コース時間/フラグ/日計等）を表します。ユーザーへの質問対象から除外してください。"
    except Exception:
        pass

    check_prompt = f"""以下のファイルデータを分析する前に、データ内に含まれる業界固有の専門用語・略語・独自の記号・不明なコードがあれば抽出せよ。

【ファイル: {filename}】
{combined}

【自動解析済み情報】
{unnamed_note if unnamed_note else "特になし"}

【指示】
- 'Unnamed: 数字'列はExcel構造上の自動変換であるため絶対に質問するな
- データ内に意味が不明確な業界固有の専門用語・略語・記号・独自コードがあれば特定せよ
- それらの意味がわからないと正確な分析ができない場合のみ、ユーザーへの質問リストを作成せよ
- 一般的なビジネス用語・数字・日付は含めるな
- 質問が不要な場合は空リストを返せ

出力形式（JSONのみ）：
{{"need_clarification": true/false, "questions": ["Q1: 〇〇とはどういう意味ですか？", "Q2: ..."], "unknown_terms": ["用語1", "用語2"]}}"""

    try:
        raw = _cllm(
            system_prompt="データ分析の専門家。JSONのみ出力。",
            messages=[{"role":"user","content":check_prompt}],
            ai_tier="core", max_tokens=1024
        )
        m = _re.search(r'\{.*\}', raw, _re.DOTALL)
        result = _json.loads(m.group(0)) if m else {"need_clarification": False, "questions": []}
        result["filename"] = filename
        result["sheets"] = list(sheets_text.keys())
        return result
    except Exception:
        return {"need_clarification": False, "questions": [], "filename": filename, "sheets": list(sheets_text.keys())}


@router.post("/file_clarify")
async def file_clarify(body: dict = Body(...), payload: dict = Depends(verify_token)):
    """専門用語確認・双方向チャットフロー
    AIがユーザーに質問し、ユーザーもAIに質問できる。
    両方が納得したら診断開始。
    """
    from api.core.llm_client import call_llm as _cllm
    import json as _json, re as _re

    messages = body.get("messages", [])
    file_summary = body.get("file_summary", "")
    user_message = body.get("user_message", "")

    system = """あなたは超一流の経営コンサルタントであり、ファイル分析の専門家である。

【役割】
ユーザーからアップロードされたファイルを分析する前に、正確な診断に必要な情報を収集する。

【最重要ルール】
1. データ内に業界固有の専門用語・略語・独自コード・記号が含まれる場合、必ず意味を確認せよ
2. ユーザーから質問された場合は、誠実・具体的に回答せよ
3. わかったふりは絶対禁止。不明点は必ず確認せよ
4. 全ての不明点が解消されたと判断したら、最後に「[診断準備完了]」と出力せよ
5. まだ不明点がある場合は質問を続けよ
6. ユーザーの回答から業界・業種を推測し、それに基づいた専門的な確認を行え

【出力形式】
- 通常の確認・回答: そのまま日本語で出力
- 全ての不明点解消時: 回答の末尾に必ず「[診断準備完了]」を付加
"""

    # 会話履歴を構築
    chat_messages = []
    for msg in messages:
        chat_messages.append({"role": msg["role"], "content": msg["content"]})

    # 最新のユーザーメッセージを追加
    if user_message:
        chat_messages.append({"role": "user", "content": user_message})

    # ファイル概要をコンテキストとして最初のメッセージに追加
    context = f"""【分析対象ファイルの概要】
{file_summary}

【会話の目的】
このファイルを正確に診断するため、業界固有の専門用語・略語・独自ルールを確認する。
ユーザーへの質問と、ユーザーからの質問への回答を行う。"""

    if chat_messages and chat_messages[0]["role"] == "user":
        chat_messages[0]["content"] = context + "\n\n" + chat_messages[0]["content"]
    else:
        chat_messages.insert(0, {"role": "user", "content": context})

    try:
        response = _cllm(
            system_prompt=system,
            messages=chat_messages,
            ai_tier="core",
            max_tokens=1024
        )
        is_ready = "[診断準備完了]" in response
        clean_response = response.replace("[診断準備完了]", "").strip()

        return {
            "ok": True,
            "message": clean_response,
            "is_ready": is_ready
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/file_clarify_save")
async def file_clarify_save(body: dict = Body(...), payload: dict = Depends(verify_token)):
    """ファイル種別ごとの確認済み用語・背景情報をFirestoreに保存"""
    from api.core.firestore_client import get_db
    uid = payload["uid"]
    file_key = body.get("file_key", "")  # ファイル名ベースのキー
    context = body.get("context", {})    # {term: explanation} の辞書

    if not file_key or not context:
        raise HTTPException(status_code=400, detail="file_keyとcontextが必要です")

    try:
        db = get_db()
        db.collection("users").document(uid).collection("file_contexts").document(file_key).set({
            "context": context,
            "updated_at": __import__("datetime").datetime.utcnow().isoformat()
        })
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/file_clarify_load")
async def file_clarify_load(file_key: str, payload: dict = Depends(verify_token)):
    """保存済みの確認内容を読み込む"""
    from api.core.firestore_client import get_db
    uid = payload["uid"]
    try:
        db = get_db()
        doc = db.collection("users").document(uid).collection("file_contexts").document(file_key).get()
        if doc.exists:
            return {"ok": True, "context": doc.to_dict().get("context", {}), "found": True}
        return {"ok": True, "context": {}, "found": False}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
